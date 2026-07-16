from __future__ import annotations

import html
import re
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Mapping, Optional, Tuple

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from streamlit_gsheets import GSheetsConnection


# ==========================================================
# CONFIGURAÇÃO
# ==========================================================
st.set_page_config(
    page_title="Gestão de Volumetria",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

URL_GSHEETS = (
    "https://docs.google.com/spreadsheets/d/"
    "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
)

META_EXECUCAO = 0.70
META_EXECUTADAS_TECNICO = 7

COL_STATUS = "STATUS CONTRATO"
COL_TOTAL = "TOTAL DE TAREFAS"
COL_TECNICO = "TÉCNICO"
COL_MONITOR = "MONITOR"
COL_REGIAO = "REGIÃO"

STATUS_ORDEM = ["Executada", "Não Executada", "Pendente"]

CORES_STATUS = {
    "Executada": "#10B981",
    "Não Executada": "#EF4444",
    "Pendente": "#F59E0B",
}


# ==========================================================
# ESTILO VISUAL
# ==========================================================
def aplicar_estilo() -> None:
    st.markdown(
        """
        <style>
            .block-container {
                padding-top: 1.4rem;
                padding-bottom: 2rem;
                max-width: 1500px;
            }

            [data-testid="stSidebar"] {
                background: #F8FAFC;
                border-right: 1px solid #E2E8F0;
            }

            [data-testid="stSidebar"] h1,
            [data-testid="stSidebar"] h2,
            [data-testid="stSidebar"] h3 {
                color: #0F172A;
            }

            .hero {
                background: linear-gradient(135deg, #0F172A 0%, #1E3A5F 60%, #075985 100%);
                padding: 26px 30px;
                border-radius: 18px;
                color: white;
                margin-bottom: 18px;
                box-shadow: 0 12px 28px rgba(15, 23, 42, 0.18);
            }

            .hero h1 {
                margin: 0;
                font-size: 30px;
                line-height: 1.2;
                font-weight: 800;
                color: white;
            }

            .hero p {
                margin: 8px 0 0 0;
                font-size: 14px;
                color: #CBD5E1;
            }

            .context-strip {
                background: #F8FAFC;
                border: 1px solid #E2E8F0;
                border-left: 5px solid #0EA5E9;
                border-radius: 10px;
                padding: 12px 16px;
                margin-bottom: 18px;
                color: #334155;
            }

            .context-strip strong {
                color: #0F172A;
                font-size: 15px;
            }

            .context-strip span {
                display: block;
                color: #64748B;
                font-size: 12px;
                margin-top: 4px;
            }

            .kpi-card {
                min-height: 132px;
                border-radius: 16px;
                padding: 18px;
                border: 1px solid #E2E8F0;
                box-shadow: 0 6px 16px rgba(15, 23, 42, 0.07);
                margin-bottom: 14px;
                transition: 0.2s ease-in-out;
            }

            .kpi-card:hover {
                transform: translateY(-2px);
                box-shadow: 0 10px 22px rgba(15, 23, 42, 0.11);
            }

            .kpi-card.blue {
                background: linear-gradient(135deg, #EFF6FF, #FFFFFF);
                border-left: 6px solid #0EA5E9;
            }

            .kpi-card.green {
                background: linear-gradient(135deg, #F0FDF4, #FFFFFF);
                border-left: 6px solid #22C55E;
            }

            .kpi-card.red {
                background: linear-gradient(135deg, #FEF2F2, #FFFFFF);
                border-left: 6px solid #EF4444;
            }

            .kpi-card.dark {
                background: linear-gradient(135deg, #0F172A, #1E293B);
                border-left: 6px solid #38BDF8;
            }

            .kpi-label {
                color: #64748B;
                font-size: 11px;
                font-weight: 800;
                letter-spacing: 0.7px;
                text-transform: uppercase;
            }

            .kpi-value {
                color: #0F172A;
                font-size: 29px;
                font-weight: 900;
                margin-top: 7px;
                line-height: 1;
            }

            .kpi-caption {
                color: #64748B;
                font-size: 12px;
                margin-top: 9px;
            }

            .dark .kpi-label {
                color: #CBD5E1;
            }

            .dark .kpi-value {
                color: #FFFFFF;
            }

            .dark .kpi-caption {
                color: #94A3B8;
            }

            div[data-testid="stDownloadButton"] button {
                border-radius: 8px;
                border: 1px solid #0EA5E9;
                color: #0369A1;
                font-weight: 700;
            }

            div[data-testid="stDownloadButton"] button:hover {
                background: #E0F2FE;
                color: #075985;
                border-color: #0284C7;
            }

            .import-box {
                border: 2px dashed #94A3B8;
                padding: 24px;
                border-radius: 14px;
                background: #F8FAFC;
                text-align: center;
                margin-top: 16px;
            }
        </style>
        """,
        unsafe_allow_html=True,
    )


# ==========================================================
# UTILITÁRIOS
# ==========================================================
def remover_acentos(valor) -> str:
    """Remove acentos de uma string, retornando string vazia para valores nulos."""
    if pd.isna(valor):
        return ""

    return (
        unicodedata.normalize("NFKD", str(valor))
        .encode("ASCII", "ignore")
        .decode("ASCII")
    )


def chave_coluna(valor: str) -> str:
    texto = remover_acentos(valor).upper()
    return re.sub(r"[^A-Z0-9]+", "", texto)


def buscar_coluna(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
    if df is None or df.empty:
        return None

    colunas = {chave_coluna(col): col for col in df.columns}

    for alias in aliases:
        chave = chave_coluna(alias)
        if chave in colunas:
            return colunas[chave]

    return None


def limpar_texto(serie: pd.Series) -> pd.Series:
    resultado = serie.astype("string").fillna("").str.strip()

    invalidos = resultado.str.upper().isin(
        ["NAN", "NONE", "<NA>", "NULO", "NULL"]
    )

    return resultado.mask(invalidos, "")


def normalizar_chave(serie: pd.Series) -> pd.Series:
    texto = limpar_texto(serie)

    return texto.map(
        lambda x: remover_acentos(x).upper() if x else ""
    )


def normalizar_login(serie: pd.Series) -> pd.Series:
    return (
        normalizar_chave(serie)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
    )


def combinar_textos(
    principal: pd.Series,
    secundario: pd.Series,
    padrao: str,
) -> pd.Series:
    p = limpar_texto(principal)
    s = limpar_texto(secundario)

    resultado = p.where(p.ne(""), s)
    resultado = resultado.where(resultado.ne(""), padrao)

    return resultado.astype(str)


def classificar_status(status_os: pd.Series) -> pd.Series:
    status = normalizar_chave(status_os)

    nao_executada = status.str.contains(
        r"NAO\s*EXECUT",
        regex=True,
        na=False,
    )

    executada = (
        status.str.contains(r"EXECUT", regex=True, na=False)
        & ~nao_executada
    )

    return pd.Series(
        np.select(
            [executada, nao_executada],
            ["Executada", "Não Executada"],
            default="Pendente",
        ),
        index=status_os.index,
    )


def formatar_numero(valor: float) -> str:
    if pd.isna(valor):
        return "—"

    return (
        f"{float(valor):,.0f}"
        .replace(",", "X")
        .replace(".", ",")
        .replace("X", ".")
    )


def formatar_percentual(valor: float) -> str:
    if pd.isna(valor):
        return "—"

    return f"{float(valor) * 100:.1f}%".replace(".", ",")


def formatar_numero_tabela(valor) -> str:
    if pd.isna(valor):
        return ""
    return formatar_numero(valor)


def formatar_percentual_tabela(valor) -> str:
    if pd.isna(valor):
        return ""
    return formatar_percentual(valor)


# ==========================================================
# LEITURA DE DADOS
# ==========================================================
@st.cache_data(show_spinner=False)
def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
    arquivo = BytesIO(file_bytes)
    nome = filename.lower()

    if nome.endswith(".xlsx"):
        return pd.read_excel(arquivo, engine="openpyxl")

    if nome.endswith(".xls"):
        return pd.read_excel(arquivo)

    return pd.read_csv(
        arquivo,
        sep=None,
        engine="python",
    )


@st.cache_data(ttl=600, show_spinner=False)
def carregar_hierarquia_gsheets() -> Tuple[pd.DataFrame, Optional[str]]:
    colunas_padrao = [
        "__LOGIN_BASE",
        "__TECNICO_GS",
        "__MONITOR_GS",
    ]

    try:
        conn = st.connection("gsheets", type=GSheetsConnection)
        raw = conn.read(spreadsheet=URL_GSHEETS)

        if raw is None or raw.empty:
            return (
                pd.DataFrame(columns=colunas_padrao),
                "A planilha de hierarquia está vazia ou indisponível.",
            )

        raw = raw.copy()
        raw.columns = raw.columns.astype(str).str.strip()

        col_login = buscar_coluna(
            raw,
            ["LOGIN", "ID", "MATRÍCULA", "MATRICULA"],
        )

        col_tecnico = buscar_coluna(
            raw,
            ["TÉCNICO", "TECNICO", "NOME", "NOME DO TÉCNICO"],
        )

        col_monitor = buscar_coluna(
            raw,
            ["MONITOR", "GESTOR", "SUPERVISOR"],
        )

        if not col_login:
            return (
                pd.DataFrame(columns=colunas_padrao),
                "A hierarquia do Google Sheets não possui uma coluna de LOGIN.",
            )

        serie_vazia = pd.Series("", index=raw.index, dtype="string")

        hierarquia = pd.DataFrame(
            {
                "__LOGIN_BASE": normalizar_login(raw[col_login]),
                "__TECNICO_GS": (
                    raw[col_tecnico] if col_tecnico else serie_vazia
                ),
                "__MONITOR_GS": (
                    raw[col_monitor] if col_monitor else serie_vazia
                ),
            }
        )

        hierarquia = hierarquia[
            hierarquia["__LOGIN_BASE"].ne("")
        ].drop_duplicates(
            subset="__LOGIN_BASE",
            keep="last",
        )

        return hierarquia, None

    except Exception as erro:
        return (
            pd.DataFrame(columns=colunas_padrao),
            f"Falha ao carregar a hierarquia do Google Sheets: {erro}",
        )


# ==========================================================
# PREPARAÇÃO DA BASE
# ==========================================================
def preparar_base(
    df_base: pd.DataFrame,
    df_hierarquia: pd.DataFrame,
) -> pd.DataFrame:
    if df_base is None or df_base.empty:
        return pd.DataFrame()

    df = df_base.copy()
    df.columns = df.columns.astype(str).str.strip().str.upper()

    diagnostico = {
        "linhas iniciais": len(df),
        "linhas removidas por contrato inválido": 0,
        "linhas finais": 0,
        "linhas com monitor identificado": 0,
    }

    # ------------------------------------------------------
    # CONTRATO
    # ------------------------------------------------------
    col_contrato = buscar_coluna(df, ["CONTRATO"])

    if col_contrato:
        antes = len(df)

        contrato = limpar_texto(df[col_contrato])
        validos = (
            contrato.notna()
            & contrato.ne("")
            & contrato.ne("0")
        )

        df = df[validos].copy()

        diagnostico["linhas removidas por contrato inválido"] = (
            antes - len(df)
        )

    if df.empty:
        df.attrs["diagnostico"] = diagnostico
        return df

    # ------------------------------------------------------
    # STATUS
    # ------------------------------------------------------
    col_status = buscar_coluna(
        df,
        [
            "STATUS DA O.S 1",
            "STATUS DA OS 1",
            "STATUS OS 1",
            "STATUS O.S. 1",
        ],
    )

    if not col_status:
        raise ValueError(
            "A coluna 'Status da O.S 1' não foi encontrada na base."
        )

    df[COL_STATUS] = classificar_status(df[col_status])

    # ------------------------------------------------------
    # TOTAL DE TAREFAS
    # ------------------------------------------------------
    col_total = buscar_coluna(
        df,
        [
            "TOTAL DE TAREFAS",
            "TOTAL TAREFAS",
            "QUANTIDADE",
            "VOLUME",
        ],
    )

    if col_total:
        df[COL_TOTAL] = pd.to_numeric(
            df[col_total],
            errors="coerce",
        ).fillna(1)
    else:
        df[COL_TOTAL] = 1

    # ------------------------------------------------------
    # HIERARQUIA: LOGIN -> TÉCNICO / MONITOR
    # ------------------------------------------------------
    col_login = buscar_coluna(
        df,
        [
            "LOGIN DO TÉCNICO",
            "LOGIN TECNICO",
            "LOGIN",
            "USUÁRIO",
            "USUARIO",
            "MATRÍCULA",
            "MATRICULA",
        ],
    )

    if (
        col_login
        and df_hierarquia is not None
        and not df_hierarquia.empty
    ):
        df["__LOGIN_BASE"] = normalizar_login(df[col_login])

        df = df.merge(
            df_hierarquia,
            on="__LOGIN_BASE",
            how="left",
        )

    serie_vazia = pd.Series("", index=df.index, dtype="string")

    col_tecnico_base = buscar_coluna(
        df,
        [
            "TÉCNICO",
            "TECNICO",
            "NOME DO TÉCNICO",
            "NOME TECNICO",
        ],
    )

    col_monitor_base = buscar_coluna(
        df,
        [
            "MONITOR",
            "GESTOR",
            "SUPERVISOR",
        ],
    )

    tecnico_base = (
        df[col_tecnico_base]
        if col_tecnico_base
        else serie_vazia
    )

    monitor_base = (
        df[col_monitor_base]
        if col_monitor_base
        else serie_vazia
    )

    tecnico_gs = (
        df["__TECNICO_GS"]
        if "__TECNICO_GS" in df.columns
        else serie_vazia
    )

    monitor_gs = (
        df["__MONITOR_GS"]
        if "__MONITOR_GS" in df.columns
        else serie_vazia
    )

    df[COL_TECNICO] = combinar_textos(
        tecnico_gs,
        tecnico_base,
        "NÃO MAPEADO",
    )

    df[COL_MONITOR] = combinar_textos(
        monitor_gs,
        monitor_base,
        "SEM MONITOR",
    )

    # ------------------------------------------------------
    # REGIÕES
    # ------------------------------------------------------
    col_cidade = buscar_coluna(
        df,
        [
            "CIDADE",
            "LOCALIDADE",
            "MUNICÍPIO",
            "MUNICIPIO",
        ],
    )

    if col_cidade:
        cidade = normalizar_chave(df[col_cidade])
    else:
        cidade = pd.Series("", index=df.index)

    cidades_spo = [
        "SAO PAULO",
    ]

    cidades_gru = [
        "GUARULHOS",
        "ARUJA",
        "MOGI DAS CRUZES",
        "SUZANO",
        "ITAQUAQUECETUBA",
        "FERRAZ DE VASCONCELOS",
        "POA",
    ]

    cidades_abcdm = [
        "SANTO ANDRE",
        "SAO BERNARDO DO CAMPO",
        "SAO CAETANO DO SUL",
        "DIADEMA",
        "MAUA",
        "RIBEIRAO PIRES",
        "RIO GRANDE DA SERRA",
    ]

    df[COL_REGIAO] = np.select(
        [
            cidade.isin(cidades_spo),
            cidade.isin(cidades_gru),
            cidade.isin(cidades_abcdm),
        ],
        [
            "LESTE",
            "GRU",
            "ABCDM",
        ],
        default="OUTRAS",
    )

    # Remove colunas auxiliares.
    df = df.drop(
        columns=[
            "__LOGIN_BASE",
            "__TECNICO_GS",
            "__MONITOR_GS",
        ],
        errors="ignore",
    )

    diagnostico["linhas finais"] = len(df)
    diagnostico["linhas com monitor identificado"] = int(
        df[COL_MONITOR].ne("SEM MONITOR").sum()
    )

    df.attrs["diagnostico"] = diagnostico

    return df


# ==========================================================
# MÉTRICAS
# ==========================================================
def calcular_kpis(df: pd.DataFrame) -> Dict[str, float]:
    executadas = df.loc[
        df[COL_STATUS] == "Executada",
        COL_TOTAL,
    ].sum()

    nao_executadas = df.loc[
        df[COL_STATUS] == "Não Executada",
        COL_TOTAL,
    ].sum()

    pendentes = df.loc[
        df[COL_STATUS] == "Pendente",
        COL_TOTAL,
    ].sum()

    baixadas = executadas + nao_executadas
    total = baixadas + pendentes

    taxa_execucao = (
        executadas / baixadas
        if baixadas > 0
        else 0
    )

    return {
        "total": float(total),
        "executadas": float(executadas),
        "nao_executadas": float(nao_executadas),
        "pendentes": float(pendentes),
        "baixadas": float(baixadas),
        "taxa_execucao": float(taxa_execucao),
        "taxa_quebra": float(1 - taxa_execucao),
        "projecao": float(
            executadas + (taxa_execucao * pendentes)
        ),
    }


def calcular_volumetria(
    df: pd.DataFrame,
    grupos: List[str],
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()

    if not all(coluna in df.columns for coluna in grupos):
        return pd.DataFrame()

    tabela = (
        df.groupby(
            grupos + [COL_STATUS],
            dropna=False,
            observed=True,
        )[COL_TOTAL]
        .sum()
        .unstack(COL_STATUS, fill_value=0)
        .reset_index()
    )

    tabela.columns.name = None

    for status in STATUS_ORDEM:
        if status not in tabela.columns:
            tabela[status] = 0

    tabela["Baixadas"] = (
        tabela["Executada"]
        + tabela["Não Executada"]
    )

    tabela["Total Alocado"] = (
        tabela["Baixadas"]
        + tabela["Pendente"]
    )

    tabela["Taxa Execução"] = np.where(
        tabela["Baixadas"] > 0,
        tabela["Executada"] / tabela["Baixadas"],
        0,
    )

    tabela["Taxa Quebra"] = 1 - tabela["Taxa Execução"]

    tabela["Projeção"] = (
        tabela["Executada"]
        + (
            tabela["Taxa Execução"]
            * tabela["Pendente"]
        )
    )

    ordem_colunas = grupos + [
        "Executada",
        "Não Executada",
        "Pendente",
        "Baixadas",
        "Total Alocado",
        "Taxa Execução",
        "Taxa Quebra",
        "Projeção",
    ]

    return (
        tabela[ordem_colunas]
        .sort_values(
            by="Total Alocado",
            ascending=False,
        )
        .reset_index(drop=True)
    )


# ==========================================================
# EXPORTAÇÃO EXCEL
# ==========================================================
def gerar_excel(df: pd.DataFrame, nome_aba: str) -> bytes:
    output = BytesIO()

    nome_aba = re.sub(r"[\[\]:*?/\\]", "", nome_aba)[:31]
    nome_aba = nome_aba or "Dados"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(
            writer,
            index=False,
            sheet_name=nome_aba,
        )

        worksheet = writer.sheets[nome_aba]

        cor_header = PatternFill(
            fill_type="solid",
            fgColor="0F172A",
        )

        for cell in worksheet[1]:
            cell.fill = cor_header
            cell.font = Font(
                color="FFFFFF",
                bold=True,
            )
            cell.alignment = Alignment(
                horizontal="center",
                vertical="center",
            )

        worksheet.freeze_panes = "A2"
        worksheet.auto_filter.ref = worksheet.dimensions

        colunas_percentuais = {
            "Taxa Execução",
            "Taxa Quebra",
        }

        colunas_numericas = {
            "Executada",
            "Não Executada",
            "Pendente",
            "Baixadas",
            "Total Alocado",
            "Projeção",
            COL_TOTAL,
        }

        for indice, coluna in enumerate(df.columns, start=1):
            letra = get_column_letter(indice)

            valores = (
                [str(coluna)]
                + df[coluna]
                .head(300)
                .fillna("")
                .astype(str)
                .tolist()
            )

            largura = min(
                max(max(len(valor) for valor in valores) + 2, 12),
                38,
            )

            worksheet.column_dimensions[letra].width = largura

            if coluna in colunas_percentuais:
                for linha in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=linha,
                        column=indice,
                    ).number_format = "0.0%"

            if coluna in colunas_numericas:
                for linha in range(2, worksheet.max_row + 1):
                    worksheet.cell(
                        row=linha,
                        column=indice,
                    ).number_format = '#,##0'

    return output.getvalue()


# ==========================================================
# COMPONENTES VISUAIS
# ==========================================================
def renderizar_card(
    container,
    titulo: str,
    valor: str,
    subtitulo: str,
    tema: str,
) -> None:
    container.markdown(
        f"""
        <div class="kpi-card {tema}">
            <div class="kpi-label">{html.escape(titulo)}</div>
            <div class="kpi-value">{html.escape(valor)}</div>
            <div class="kpi-caption">{html.escape(subtitulo)}</div>
        </div>
        """,
        unsafe_allow_html=True,
    )


def estilo_projecao(_valor) -> str:
    return (
        "background-color:#0F172A;"
        "color:#FFFFFF;"
        "font-weight:700;"
    )


def estilo_meta_executadas(valor) -> str:
    try:
        if float(valor) >= META_EXECUTADAS_TECNICO:
            return (
                "background-color:#DCFCE7;"
                "color:#166534;"
                "font-weight:700;"
            )
    except Exception:
        pass

    return ""


def aplicar_estilo_tabela(
    tabela: pd.DataFrame,
    destacar_meta_tecnico: bool = False,
):
    formatadores = {
        "Executada": formatar_numero_tabela,
        "Não Executada": formatar_numero_tabela,
        "Pendente": formatar_numero_tabela,
        "Baixadas": formatar_numero_tabela,
        "Total Alocado": formatar_numero_tabela,
        "Projeção": formatar_numero_tabela,
        "Taxa Execução": formatar_percentual_tabela,
        "Taxa Quebra": formatar_percentual_tabela,
    }

    formatos_validos: Mapping[str, Any] = {
        coluna: formato
        for coluna, formato in formatadores.items()
        if coluna in tabela.columns
    }

    styler = tabela.style.format(formatter=formatos_validos)

    if "Projeção" in tabela.columns:
        if hasattr(styler, "map"):
            styler = styler.map(
                estilo_projecao,
                subset=["Projeção"],
            )
    else:
            styler = styler.map(
                estilo_projecao,
                subset=["Projeção"],
            )

    if destacar_meta_tecnico and "Executada" in tabela.columns:
        if hasattr(styler, "map"):
            styler = styler.map(
                estilo_meta_executadas,
                subset=["Executada"],
            )
        else:
            styler = styler.map(
                estilo_meta_executadas,
                subset=["Executada"],
            )

    return styler


def renderizar_tabela(
    tabela: pd.DataFrame,
    titulo: str,
    arquivo: str,
    aba_excel: str,
    chave_download: str,
    destacar_meta_tecnico: bool = False,
) -> None:
    st.subheader(titulo)

    if tabela is None or tabela.empty:
        st.info("Sem dados para exibir.")
        return

    styler = aplicar_estilo_tabela(
        tabela,
        destacar_meta_tecnico=destacar_meta_tecnico,
    )

    st.dataframe(
        styler,
        use_container_width=True,
        hide_index=True,
    )

    st.download_button(
        label="📥 Baixar Excel",
        data=gerar_excel(tabela, aba_excel),
        file_name=arquivo,
        mime=(
            "application/vnd.openxmlformats-officedocument."
            "spreadsheetml.sheet"
        ),
        use_container_width=True,
        key=chave_download,
    )


# ==========================================================
# GRÁFICOS
# ==========================================================
def aplicar_layout_grafico(
    figura: go.Figure,
    altura: int = 360,
) -> go.Figure:
    figura.update_layout(
        height=altura,
        paper_bgcolor="rgba(0,0,0,0)",
        plot_bgcolor="rgba(0,0,0,0)",
        font={
            "family": "Roboto, sans-serif",
            "color": "#334155",
        },
        margin={
            "l": 10,
            "r": 10,
            "t": 44,
            "b": 60,  # 🔥 aumenta margem inferior
        },
        legend=dict(
            orientation="h",
            yanchor="top",
            y=-0.25,   # 🔥 coloca abaixo do gráfico
            xanchor="center",
            x=0.5
        ),
        xaxis_title=None,
        yaxis_title=None
    )

    figura.update_xaxes(
        showline=False,
        zeroline=False,
        gridcolor="#E2E8F0",
        title=None
    )

    figura.update_yaxes(
        showline=False,
        zeroline=False,
        gridcolor="#E2E8F0",
        title=None
    )

    return figura


def grafico_status(kpis: Dict[str, float]) -> go.Figure:
    dados = pd.DataFrame(
        {
            "Status": STATUS_ORDEM,
            "Volume": [
                kpis["executadas"],
                kpis["nao_executadas"],
                kpis["pendentes"],
            ],
        }
    )

    figura = go.Figure(
        data=[
            go.Pie(
                labels=dados["Status"],
                values=dados["Volume"],
                hole=0.68,
                marker={
                    "colors": [
                        CORES_STATUS[status]
                        for status in STATUS_ORDEM
                    ]
                },
                textinfo="percent",
                textfont={"size": 12},
                hovertemplate=(
                    "<b>%{label}</b><br>"
                    "Volume: %{value:,.0f}<br>"
                    "Participação: %{percent}"
                    "<extra></extra>"
                ),
            )
        ]
    )

    figura.add_annotation(
        text=(
            f"<b>{formatar_numero(kpis['total'])}</b>"
            "<br><span style='font-size:11px;color:#64748B;'>"
            "Total alocado"
            "</span>"
        ),
        x=0.5,
        y=0.5,
        showarrow=False,
        font={
            "size": 19,
            "color": "#0F172A",
        },
    )

    figura.update_layout(
        title="Composição da operação",
        showlegend=True,
    )

    return aplicar_layout_grafico(figura, altura=355)


def grafico_regioes(df: pd.DataFrame) -> go.Figure:
    tabela = calcular_volumetria(
        df,
        [COL_REGIAO],
    )

    ordem_regioes = tabela[COL_REGIAO].tolist()

    dados_longos = tabela.melt(
        id_vars=[COL_REGIAO],
        value_vars=STATUS_ORDEM,
        var_name="Status",
        value_name="Volume",
    )

    figura = px.bar(
        dados_longos,
        x=COL_REGIAO,
        y="Volume",
        color="Status",
        barmode="stack",
        color_discrete_map=CORES_STATUS,
        category_orders={
            COL_REGIAO: ordem_regioes,
            "Status": STATUS_ORDEM,
        },
        labels={
            COL_REGIAO: "",
            "Volume": "Tarefas",
            "Status": "Status",
        },
    )

    figura.update_traces(
        hovertemplate=(
            "<b>%{x}</b><br>"
            "%{fullData.name}: %{y:,.0f}"
            "<extra></extra>"
        )
    )

    figura.update_layout(
        title="Volumetria por região",
    )

    return aplicar_layout_grafico(figura, altura=355)


def grafico_ranking_monitores(
    df: pd.DataFrame,
    limite: int = 12,
) -> go.Figure:
    ranking = calcular_volumetria(
        df,
        [COL_MONITOR],
    )

    ranking = (
        ranking.nlargest(limite, "Total Alocado")
        .sort_values("Taxa Execução")
        .reset_index(drop=True)
    )

    figura = px.bar(
        ranking,
        x="Taxa Execução",
        y=COL_MONITOR,
        orientation="h",
        color="Taxa Execução",
        color_continuous_scale=[
            "#EF4444",
            "#F59E0B",
            "#10B981",
        ],
        range_color=[0, 1],
        text=ranking["Taxa Execução"].apply(
            formatar_percentual
        ),
        custom_data=[
            "Executada",
            "Não Executada",
            "Pendente",
            "Total Alocado",
        ],
        labels={
            "Taxa Execução": "Taxa de execução",
            COL_MONITOR: "",
        },
    )

    figura.update_traces(
        textposition="outside",
        cliponaxis=False,
        hovertemplate=(
            "<b>%{y}</b><br>"
            "Taxa de execução: %{x:.1%}<br>"
            "Executadas: %{customdata[0]:,.0f}<br>"
            "Não executadas: %{customdata[1]:,.0f}<br>"
            "Pendentes: %{customdata[2]:,.0f}<br>"
            "Total alocado: %{customdata[3]:,.0f}"
            "<extra></extra>"
        ),
    )

    figura.add_vline(
        x=META_EXECUCAO,
        line_width=2,
        line_dash="dash",
        line_color="#0F172A",
        annotation_text=(
            f"Meta {formatar_percentual(META_EXECUCAO)}"
        ),
        annotation_position="top",
    )

    figura.update_xaxes(
        tickformat=".0%",
        range=[0, 1.15],
    )

    figura.update_layout(
        title="Ranking de execução por monitor",
        coloraxis_showscale=False,
    )

    return aplicar_layout_grafico(figura, altura=470)


def grafico_equipes(
    tabela_equipes: pd.DataFrame,
) -> go.Figure:
    dados = tabela_equipes.nlargest(
        20,
        "Total Alocado",
    )

    figura = px.bar(
        dados,
        x=COL_MONITOR,
        y="Projeção",
        color=COL_REGIAO,
        barmode="group",
        text=dados["Projeção"].apply(formatar_numero),
        custom_data=[
            "Executada",
            "Pendente",
            "Taxa Execução",
            "Total Alocado",
        ],
        labels={
            COL_MONITOR: "",
            "Projeção": "Projeção",
            COL_REGIAO: "Região",
        },
        color_discrete_sequence=[
            "#0EA5E9",
            "#8B5CF6",
            "#10B981",
            "#F59E0B",
        ],
    )

    figura.update_traces(
        textposition="outside",
        cliponaxis=False,
        hovertemplate=(
            "<b>%{x}</b><br>"
            "Projeção: %{y:,.0f}<br>"
            "Executadas: %{customdata[0]:,.0f}<br>"
            "Pendentes: %{customdata[1]:,.0f}<br>"
            "Taxa de execução: %{customdata[2]:.1%}<br>"
            "Total alocado: %{customdata[3]:,.0f}"
            "<extra></extra>"
        ),
    )

    figura.update_xaxes(
        tickangle=-35,
    )

    figura.update_layout(
        title="Projeção de resultado por equipe",
    )

    return aplicar_layout_grafico(figura, altura=430)


def grafico_tecnicos(
    tabela_tecnicos: pd.DataFrame,
) -> go.Figure:
    dados = (
        tabela_tecnicos.nlargest(
            20,
            "Total Alocado",
        )
        .sort_values("Executada")
        .reset_index(drop=True)
    )

    figura = px.bar(
        dados,
        x="Executada",
        y=COL_TECNICO,
        orientation="h",
        color="Taxa Execução",
        color_continuous_scale=[
            "#EF4444",
            "#F59E0B",
            "#10B981",
        ],
        range_color=[0, 1],
        text=dados["Executada"].apply(formatar_numero),
        custom_data=[
            "Não Executada",
            "Pendente",
            "Taxa Execução",
            "Total Alocado",
            "Projeção",
        ],
        labels={
            "Executada": "Tarefas executadas",
            COL_TECNICO: "",
        },
    )

    figura.update_traces(
        textposition="outside",
        cliponaxis=False,
        hovertemplate=(
            "<b>%{y}</b><br>"
            "Executadas: %{x:,.0f}<br>"
            "Não executadas: %{customdata[0]:,.0f}<br>"
            "Pendentes: %{customdata[1]:,.0f}<br>"
            "Taxa de execução: %{customdata[2]:.1%}<br>"
            "Total alocado: %{customdata[3]:,.0f}<br>"
            "Projeção: %{customdata[4]:,.0f}"
            "<extra></extra>"
        ),
    )

    figura.add_vline(
        x=META_EXECUTADAS_TECNICO,
        line_width=2,
        line_dash="dash",
        line_color="#0F172A",
        annotation_text=(
            f"Meta: {META_EXECUTADAS_TECNICO} executadas"
        ),
        annotation_position="top",
    )

    figura.update_layout(
        title="Ranking técnico por tarefas executadas",
        coloraxis_showscale=False,
    )

    return aplicar_layout_grafico(figura, altura=550)


# ==========================================================
# APP
# ==========================================================
def limpar_base_memoria() -> None:
    chaves = [
        "df_processado",
        "nome_arquivo",
        "aviso_hierarquia",
    ]

    for chave in chaves:
        st.session_state.pop(chave, None)


def main() -> None:
    aplicar_estilo()

    st.markdown(
        """
        <div class="hero">
            <h1>📊 Gestão de Volumetria</h1>
            <p>
                Painel executivo para acompanhamento de execução,
                quebra, pendências e projeção operacional.
            </p>
        </div>
        """,
        unsafe_allow_html=True,
    )

    if "df_processado" not in st.session_state:
        st.session_state.df_processado = None

    # ------------------------------------------------------
    # SIDEBAR: CONTROLES
    # ------------------------------------------------------
    with st.sidebar:
        st.header("⚙️ Controles")

        if st.button(
            "🔄 Substituir base",
            use_container_width=True,
        ):
            limpar_base_memoria()
            st.rerun()

        if st.button(
            "🧹 Limpar cache completo",
            use_container_width=True,
        ):
            st.cache_data.clear()
            limpar_base_memoria()
            st.rerun()

        st.divider()

    # ------------------------------------------------------
    # IMPORTAÇÃO
    # ------------------------------------------------------
    if st.session_state.df_processado is None:
        st.markdown(
            """
            <div class="import-box">
                <h3>Importação da Base Operacional</h3>
                <p>
                    Envie um arquivo Excel ou CSV contendo a base
                    operacional para iniciar a análise.
                </p>
            </div>
            """,
            unsafe_allow_html=True,
        )

        arquivo = st.file_uploader(
            "📥 Selecione a base operacional",
            type=["xlsx", "xls", "csv"],
        )

        if arquivo:
            try:
                with st.spinner(
                    "Processando base e atualizando hierarquia..."
                ):
                    df_base = ler_arquivo(
                        arquivo.getvalue(),
                        arquivo.name,
                    )

                    df_hierarquia, aviso_hierarquia = (
                        carregar_hierarquia_gsheets()
                    )

                    df_processado = preparar_base(
                        df_base,
                        df_hierarquia,
                    )

                    st.session_state.df_processado = df_processado
                    st.session_state.nome_arquivo = arquivo.name
                    st.session_state.aviso_hierarquia = aviso_hierarquia

                st.rerun()

            except Exception as erro:
                st.error(
                    f"Não foi possível processar a base: {erro}"
                )

        return

    df_full = st.session_state.df_processado

    if df_full is None or df_full.empty:
        st.warning(
            "A base ficou vazia após as validações de contrato."
        )
        return

    aviso_hierarquia = st.session_state.get(
        "aviso_hierarquia"
    )

    if aviso_hierarquia:
        st.warning(
            f"⚠️ {aviso_hierarquia} "
            "O painel seguirá utilizando as informações disponíveis na base."
        )

    # ------------------------------------------------------
    # DIAGNÓSTICO
    # ------------------------------------------------------
    with st.sidebar.expander(
        "🔎 Diagnóstico da carga",
        expanded=False,
    ):
        diagnostico = df_full.attrs.get(
            "diagnostico",
            {},
        )

        for chave, valor in diagnostico.items():
            if isinstance(valor, (int, float)):
                valor_exibicao = formatar_numero(valor)
            else:
                valor_exibicao = valor

            st.write(
                f"**{chave.title()}**: {valor_exibicao}"
            )

    # ------------------------------------------------------
    # FILTRO DE MONITOR
    # ------------------------------------------------------
    with st.sidebar:
        st.markdown("### 🎯 Filtros")

        monitores = sorted(
            df_full[COL_MONITOR]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        monitores_selecionados = st.multiselect(
            "Monitor",
            options=monitores,
            default=monitores,
        )

        st.caption(
            "A região é utilizada nas análises e gráficos, "
            "mas não possui filtro interativo."
        )

    df = df_full[
        df_full[COL_MONITOR].isin(monitores_selecionados)
    ].copy()

    if df.empty:
        st.warning(
            "Nenhum dado encontrado para os filtros selecionados."
        )
        return

    # ------------------------------------------------------
    # CONTEXTO
    # ------------------------------------------------------
    regioes = sorted(
        df_full[COL_REGIAO]
        .dropna()
        .astype(str)
        .unique()
        .tolist()
    )

    texto_regioes = ", ".join(regioes)

    if len(monitores_selecionados) < len(monitores):
        monitores_titulo = ", ".join(
            monitores_selecionados[:3]
        )

        if len(monitores_selecionados) > 3:
            restantes = len(monitores_selecionados) - 3
            monitores_titulo += f" e mais {restantes}"

        titulo_contexto = (
            f"Análise dos monitores: {monitores_titulo}"
        )
    else:
        titulo_contexto = "Visão Geral Consolidada"

    st.markdown(
        f"""
        <div class="context-strip">
            <strong>{html.escape(titulo_contexto)}</strong>
            <span>
                Base: {html.escape(texto_regioes)}
            </span>
        </div>
        """,
        unsafe_allow_html=True,
    )

    # ------------------------------------------------------
    # KPIs
    # ------------------------------------------------------
    kpis = calcular_kpis(df)

    col1, col2, col3, col4 = st.columns(4)

    renderizar_card(
        col1,
        "Total Alocado",
        formatar_numero(kpis["total"]),
        f"{formatar_numero(kpis['pendentes'])} pendentes",
        "blue",
    )

    renderizar_card(
        col2,
        "Executadas",
        formatar_numero(kpis["executadas"]),
        f"Execução: {formatar_percentual(kpis['taxa_execucao'])}",
        "green",
    )

    renderizar_card(
        col3,
        "Não Executadas",
        formatar_numero(kpis["nao_executadas"]),
        f"Quebra: {formatar_percentual(kpis['taxa_quebra'])}",
        "red",
    )

    renderizar_card(
        col4,
        "Projeção",
        formatar_numero(kpis["projecao"]),
        "Baseada na taxa atual de execução",
        "dark",
    )

    st.markdown("### 📈 Visão Executiva")

    st.plotly_chart(
        grafico_status(kpis),
        use_container_width=True,
        config={"displayModeBar": False},
    )

    st.plotly_chart(
        grafico_ranking_monitores(df),
        use_container_width=True,
        config={"displayModeBar": False},
    )

    st.caption(
        f"Meta de referência de execução: "
        f"{formatar_percentual(META_EXECUCAO)}."
    )

    st.divider()

    # ------------------------------------------------------
    # ABAS
    # ------------------------------------------------------
    aba_equipe, aba_tecnico, aba_base = st.tabs(
        [
            "👥 Desempenho por Equipe",
            "🧑‍🔧 Desempenho por Técnico",
            "📋 Base Completa",
        ]
    )

    # ------------------------------------------------------
    # ABA EQUIPE
    # ------------------------------------------------------
    with aba_equipe:
        df_equipes = calcular_volumetria(
            df,
            [COL_REGIAO, COL_MONITOR],
        )

        if not df_equipes.empty:
            st.plotly_chart(
                grafico_equipes(df_equipes),
                use_container_width=True,
                config={"displayModeBar": False},
            )

        renderizar_tabela(
            tabela=df_equipes,
            titulo="Desempenho Consolidado por Equipe",
            arquivo="desempenho_equipes.xlsx",
            aba_excel="Por Equipe",
            chave_download="download_equipes",
        )

    # ------------------------------------------------------
    # ABA TÉCNICO
    # ------------------------------------------------------
    with aba_tecnico:
        monitores_tecnico = sorted(
            df[COL_MONITOR]
            .dropna()
            .astype(str)
            .unique()
            .tolist()
        )

        if not monitores_tecnico:
            st.info(
                "Não há monitores disponíveis nos filtros atuais."
            )
        else:
            if (
                "monitor_tecnico"
                in st.session_state
                and st.session_state.monitor_tecnico
                not in monitores_tecnico
            ):
                del st.session_state.monitor_tecnico

            monitor_selecionado = st.selectbox(
                "Selecione um monitor para detalhar",
                options=monitores_tecnico,
                key="monitor_tecnico",
            )

            df_monitor = df[
                df[COL_MONITOR] == monitor_selecionado
            ].copy()

            df_tecnicos = calcular_volumetria(
                df_monitor,
                [COL_TECNICO],
            )

            if not df_tecnicos.empty:
                df_tecnicos = (
                    df_tecnicos.sort_values(
                        by=[
                            "Executada",
                            "Total Alocado",
                        ],
                        ascending=[False, False],
                    )
                    .reset_index(drop=True)
                )

                st.plotly_chart(
                    grafico_tecnicos(df_tecnicos),
                    use_container_width=True,
                    config={"displayModeBar": False},
                )

            renderizar_tabela(
                tabela=df_tecnicos,
                titulo=(
                    f"Desempenho Técnico — "
                    f"{monitor_selecionado}"
                ),
                arquivo=(
                    f"desempenho_tecnicos_"
                    f"{str(monitor_selecionado).strip()}.xlsx"
                ),
                aba_excel="Por Tecnico",
                chave_download="download_tecnicos",
                destacar_meta_tecnico=True,
            )

    # ------------------------------------------------------
    # ABA BASE COMPLETA
    # ------------------------------------------------------
    with aba_base:
        st.subheader("Base Completa Filtrada")

        st.caption(
            f"Linhas disponíveis após filtros: "
            f"{formatar_numero(len(df))}"
        )

        st.dataframe(
            df,
            use_container_width=True,
            hide_index=True,
        )

        st.download_button(
            label="📥 Baixar Base Filtrada em Excel",
            data=gerar_excel(df, "Base Filtrada"),
            file_name="base_filtrada.xlsx",
            mime=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            ),
            use_container_width=True,
            key="download_base_completa",
        )


if __name__ == "__main__":
    main()
# rota_inicial.py

from __future__ import annotations

import re
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Literal, Optional

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from streamlit_gsheets import GSheetsConnection

# ── Componentes corporativos ──────────────────────────────────────────────────
from components.componentes import (
    aplicar_estilo,
    render_hero,
    render_kpi,
    render_kpi_sm,
    render_section,
    render_section_header,
    render_insight,
    FONTE_TEXTO,
    FONTE_TITULO,
    COR_PRIMARIA,
    COR_SECUNDARIA,
    COR_SUCESSO,
    COR_ALERTA,
    COR_NEUTRO,
    COR_TEXTO,
    COR_TEXTO_2,
    COR_TEXTO_3,
    COR_BORDA,
    COR_FUNDO,
)

# ====================================================
# 1. CONFIGURAÇÃO DA PÁGINA
# ====================================================
st.set_page_config(
    page_title="Rota Inicial",
    page_icon="🗺️",
    layout="wide",
    initial_sidebar_state="expanded",
)

if "df_master" not in st.session_state:
    st.session_state["df_master"] = None

if "total_montados_manual" not in st.session_state:
    st.session_state["total_montados_manual"] = 0

if "_reset_counter" not in st.session_state:
    st.session_state["_reset_counter"] = 0

# ====================================================
# 2. CONFIGURAÇÕES GLOBAIS
# ====================================================
URL_GSHEETS = (
    "https://docs.google.com/spreadsheets/d/"
    "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
)

CONTRATO_VALORES_VAZIOS = {"", "NAN", "NONE", "N/A", "NA", "-", "0", "NULL", "<NA>"}

MAPEAMENTO_PERIODOS = {
    "08:00 - 10:00": "Manhã",
    "08:00 - 11:00": "Manhã",
    "08:00 - 12:00": "Manhã",
    "10:00 - 12:00": "Manhã",
    "11:00 - 14:00": "Manhã",
    "12:00 - 14:00": "Tarde I",
    "12:00 - 15:00": "Tarde I",
    "12:00 - 18:00": "Tarde II",
    "14:00 - 16:00": "Tarde II",
    "14:00 - 17:00": "Tarde II",
    "15:00 - 18:00": "Tarde II",
    "16:00 - 18:00": "Tarde II",
    "17:00 - 20:00": "Tarde II",
    "Imediata": "Imediata",
}

TEMAS_CARD = {
    "amarelo":  {"fundo": "#FEF9C3", "texto": "#854D0E", "borda": "#EAB308", "titulo": "#A16207"},
    "azul":     {"fundo": "#F0F9FF", "texto": "#0369A1", "borda": "#0EA5E9", "titulo": "#075985"},
    "verde":    {"fundo": "#F0FDF4", "texto": "#15803D", "borda": "#22C55E", "titulo": "#166534"},
    "roxo":     {"fundo": "#FAF5FF", "texto": "#7E22CE", "borda": "#A855F7", "titulo": "#6B21A8"},
    "cinza":    {"fundo": "#F8FAFC", "texto": "#334155", "borda": "#94A3B8", "titulo": "#64748B"},
    "escuro":   {"fundo": "#1E293B", "texto": "#FFFFFF", "borda": "#475569", "titulo": "#E2E8F0"},
    "laranja":  {"fundo": "#FFF7ED", "texto": "#C2410C", "borda": "#F97316", "titulo": "#9A3412"},
    "vermelho": {"fundo": "#FEF2F2", "texto": "#B91C1C", "borda": "#EF4444", "titulo": "#991B1B"},
}

CORES_REGIAO = {
    "LESTE":  {"bg": "#DBEAFE", "text": "#1E40AF", "border": "#3B82F6"},
    "GRU":    {"bg": "#D1FAE5", "text": "#065F46", "border": "#10B981"},
    "ABCDM":  {"bg": "#EDE9FE", "text": "#5B21B6", "border": "#8B5CF6"},
    "OUTRAS": {"bg": "#F1F5F9", "text": "#475569", "border": "#94A3B8"},
}

RENOMEAR_COLUNAS: Dict[str, str] = {
    "Monitor": "Monitor",
    "OS": "Volume de O.S.",
    "GPON": "GPON",
    "ND": "Adesão",
    "PME": "PME",
    "Migração": "Migração",
    "Qtd_4K": "4K",
    "Ultra": "Ponto Ultra",
    "Soundbox": "Soundbox",
    "Equipe": "Equipe",
    "Média": "Média/Téc.",
    "NOME_OFICIAL": "Técnico",
    "TOTAL_TAREFAS": "Total de O.S.",
    "LOGIN_TECNICO": "Login",
    "STATUS_ATIVIDADE": "Status",
    "PERIODO_TRATADO": "Período",
    "HABILIDADE": "Habilidade",
    "TIPO_OS": "Tipo OS",
    "PRODUTO": "Produto",
    "INTERVALO": "Intervalo",
    "CIDADE": "Cidade",
    "CONTRATO": "Contrato",
    "REGIÃO": "Região",
    "VELOCIDADE_BANDA": "Velocidade",
}

# ====================================================
# 2.1  EXTRAÇÃO DE VELOCIDADE (PRIORIZA PENDENTE)
# ====================================================
_RE_PRODUTO_INTERNET = re.compile(r"\b(BL|BANDA\s*LARGA)\b", re.IGNORECASE)

_RE_Mbps_Gbps = re.compile(
    r"(\d+(?:[.,]\d+)?)\s*(GIGA|GB|G|MEGA|MB|M)\b",
    re.IGNORECASE
)

_RE_BL = re.compile(
    r"\bBL\s*(\d+(?:[.,]\d+)?)\s*(M|MEGA|G|GIGA)?\b",
    re.IGNORECASE
)


def _extrair_velocidade_e_mbps(produto: Any) -> tuple[str, float]:
    if not isinstance(produto, str) or not produto.strip():
        return "", np.nan

    produto_up = produto.upper()

    if not _RE_PRODUTO_INTERNET.search(produto_up):
        return "", np.nan

    mbps = np.nan

    m_bl = _RE_BL.search(produto_up)
    if m_bl:
        valor = float(m_bl.group(1).replace(",", "."))
        unidade = m_bl.group(2)
        if unidade and unidade.startswith("G"):
            mbps = valor * 1000
        else:
            mbps = valor
    else:
        m = _RE_Mbps_Gbps.search(produto_up)
        if m:
            valor = float(m.group(1).replace(",", "."))
            unidade = m.group(2)
            if unidade.startswith("G"):
                mbps = valor * 1000
            else:
                mbps = valor

    if pd.notna(mbps) and mbps >= 70:
        label = f"{int(mbps)} Mbps" if mbps < 1000 else f"{mbps/1000:.0f} Gbps"
        return label, mbps

    return "", np.nan


def atribuir_velocidade_por_contrato(df: pd.DataFrame) -> pd.DataFrame:
    df = df.copy()
    col_status = "STATUS_PRODUTO"
    if col_status not in df.columns or not df[col_status].notna().any():
        col_status = "STATUS_ATIVIDADE"

    status = df[col_status].astype(str).str.strip().str.upper()
    vel_extraida = df["PRODUTO"].apply(_extrair_velocidade_e_mbps)
    df["_VEL_LABEL"] = vel_extraida.str[0]
    df["_VEL_MBPS"] = vel_extraida.str[1]
    tem_vel = df["_VEL_LABEL"].ne("")

    df["_PRIORIDADE_VEL"] = np.select(
        [
            tem_vel & status.eq("PENDENTE"),
            tem_vel & status.eq("INSTALADO"),
            tem_vel,
        ],
        [0, 1, 2],
        default=9,
    )

    vel_por_contrato = (
        df.loc[tem_vel]
        .sort_values(
            ["CONTRATO", "_PRIORIDADE_VEL", "_VEL_MBPS"],
            ascending=[True, True, False],
        )
        .drop_duplicates(subset=["CONTRATO"])
        [["CONTRATO", "_VEL_LABEL"]]
        .rename(columns={"_VEL_LABEL": "VELOCIDADE_BANDA"})
    )

    df = df.drop(columns=["VELOCIDADE_BANDA"], errors="ignore")
    df = df.merge(vel_por_contrato, on="CONTRATO", how="left")
    df["VELOCIDADE_BANDA"] = df["VELOCIDADE_BANDA"].fillna("")
    return df.drop(
        columns=["_VEL_LABEL", "_VEL_MBPS", "_PRIORIDADE_VEL"], errors="ignore"
    )


# ====================================================
# 3. CSS LOCAL
# ====================================================
def _injetar_css_local() -> None:
    st.markdown(
        """
        <style>
        /* ── Resultado da Base ── */
        .resultado-base {
            background: linear-gradient(135deg, #012869 0%, #1E40AF 50%, #F37C04 100%);
            padding: 1rem 1.5rem; border-radius: 0.75rem;
            margin-bottom: 1.5rem; display: flex; align-items: center;
            flex-wrap: wrap; gap: 0.6rem;
            box-shadow: 0 6px 20px rgba(1,40,105,0.20);
            position: relative; overflow: hidden;
        }
        .resultado-base::before {
            content:''; position:absolute; top:-50%; right:-5%;
            width:250px; height:250px;
            background:rgba(255,255,255,0.05); border-radius:50%;
            pointer-events:none;
        }
        .resultado-base-label {
            color:#FFFFFF; font-size:0.8rem; font-weight:700;
            text-transform:uppercase; letter-spacing:0.08em;
            margin-right:0.3rem; position:relative; z-index:2;
        }
        .resultado-base-regiao {
            padding:0.35rem 0.95rem; border-radius:999px;
            font-size:0.82rem; font-weight:700; letter-spacing:0.04em;
            border:2px solid rgba(255,255,255,0.4);
            position:relative; z-index:2;
        }
        .resultado-base-count {
            color:#FFFFFF; font-size:0.78rem; margin-left:auto;
            font-weight:700; position:relative; z-index:2;
        }

        /* ── Tabelas Rota/Turno ── */
        .rota-wrapper {
            margin: 24px 0;
            box-shadow: 0 4px 20px rgba(0,0,0,0.05);
            border-radius: 12px; overflow: hidden; border: 1px solid #E2E8F0;
        }
        .rota-titulo {
            background: #0F172A; color: #F8FAFC;
            font-weight: 700; font-size: 14px; text-align: left;
            padding: 14px 20px; letter-spacing: 0.5px;
            text-transform: uppercase; border-bottom: 3px solid #312E81;
        }
        table.rota-tab {
            width: 100%; border-collapse: separate; border-spacing: 0;
            background: white; font-size: 13px;
        }
        table.rota-tab th {
            background: #1E293B; color: #F1F5F9;
            font-weight: 600; font-size: 11px;
            text-transform: uppercase; letter-spacing: 0.05em;
            padding: 12px 14px; text-align: center;
            border-bottom: 1px solid #334155;
        }
        table.rota-tab th.th-os     { background: #2E2514; color: #FBBF24; }
        table.rota-tab th.th-equipe { background: #064E3B; color: #6EE7B7; }
        table.rota-tab th.th-media  { background: #1E3A8A; color: #93C5FD; }
        table.rota-tab td {
            padding: 10px 14px; text-align: center;
            border-bottom: 1px solid #F1F5F9;
            color: #334155; font-variant-numeric: tabular-nums;
        }
        table.rota-tab td.col-monitor {
            text-align: left; font-weight: 600; padding-left: 20px; color: #0F172A;
        }
        table.rota-tab td.col-os {
            background: #FFFBEB; font-weight: 700; color: #92400E;
            border-left: 1px solid #FEF3C7; border-right: 1px solid #FEF3C7;
        }
        table.rota-tab td.col-equipe {
            background: #F0FDF4; font-weight: 700; color: #166534;
            border-left: 1px solid #D1FAE5; border-right: 1px solid #D1FAE5;
        }
        table.rota-tab td.col-media {
            background: #F8FAFC; font-weight: 700; color: #1E293B;
        }
        table.rota-tab tr.total-escalados td {
            background: #F1F5F9; color: #0F172A; font-weight: 700; font-size: 13px;
            border-top: 2px solid #CBD5E1; border-bottom: 1px solid #CBD5E1;
        }
        table.rota-tab tr.total-escalados td.col-os     { background: #FEF3C7; color: #92400E; }
        table.rota-tab tr.total-escalados td.col-equipe { background: #D1FAE5; color: #065F46; }
        table.rota-tab tr.total-escalados td.col-media  { background: #E2E8F0; color: #0F172A; }
        table.rota-tab tr.total-montados td {
            background: #E2E8F0; color: #0F172A; font-weight: 800; font-size: 13px;
            border-bottom: none;
        }
        table.rota-tab tr.total-montados td.col-os     { background: #FDE68A; color: #78350F; }
        table.rota-tab tr.total-montados td.col-equipe { background: #A7F3D0; color: #064E3B; }
        table.rota-tab tr.total-montados td.col-media  { background: #CBD5E1; color: #0F172A; }
        table.rota-tab tbody tr:nth-child(even):not(.total-escalados):not(.total-montados) td
            { background: #FAFAFA; }
        table.rota-tab tbody tr:nth-child(even):not(.total-escalados):not(.total-montados) td.col-os
            { background: #FFFDF5; }
        table.rota-tab tbody tr:nth-child(even):not(.total-escalados):not(.total-montados) td.col-equipe
            { background: #F4FDF7; }
        table.rota-tab tbody tr:hover:not(.total-escalados):not(.total-montados) td
            { background: #F1F5F9; }

        /* ── DataFrames estilizados ── */
        .styled-table-wrapper {
            background: #FFFFFF; border-radius: 0.75rem;
            padding: 1rem 1.2rem; box-shadow: 0 2px 8px rgba(0,0,0,0.04);
            margin-bottom: 0.5rem;
        }
        .styled-table-title {
            font-size: 1rem; font-weight: 700; color: #0F172A;
            margin-bottom: 0.4rem; display: flex; align-items: center; gap: 0.5rem;
        }
        .styled-table-badge {
            font-size: 0.68rem; background: #E0F2FE; color: #0369A1;
            padding: 0.15rem 0.5rem; border-radius: 999px; font-weight: 600;
        }
        </style>
        """,
        unsafe_allow_html=True,
    )


# ====================================================
# 4. COMPONENTES LOCAIS
# ====================================================
def render_resultado_base(regioes: List[str], total: int) -> None:
    badges = ""
    for reg in sorted(regioes):
        c = CORES_REGIAO.get(reg, CORES_REGIAO["OUTRAS"])
        badges += (
            f'<span class="resultado-base-regiao" '
            f'style="background:{c["bg"]};color:{c["text"]};'
            f'border-color:{c["border"]}">{reg}</span>'
        )
    st.markdown(
        f"""
        <div class="resultado-base">
            <span class="resultado-base-label">📋 Resultado da Base:</span>
            {badges}
            <span class="resultado-base-count">{total:,} registros</span>
        </div>
        """,
        unsafe_allow_html=True,
    )


def render_dataframe_local(
    df: pd.DataFrame,
    titulo: str = "",
    icone: str = "📊",
    badge: str = "",
    fmt: Optional[Dict[str, Any]] = None,
    color_col: Optional[str] = None,
    color_meta: Optional[float] = None,
    color_invertido: bool = False,
    height: int | Literal["auto", "stretch", "content"] = "auto",
) -> None:
    badge_text = badge or f"{len(df)} registros"
    st.markdown(
        f"""
        <div class="styled-table-wrapper">
            <div class="styled-table-title">
                <span>{icone}</span><span>{titulo}</span>
                <span class="styled-table-badge">{badge_text}</span>
            </div>
        </div>
        """,
        unsafe_allow_html=True,
    )

    df_display = df.copy()
    colunas_para_renomear: Dict[str, str] = {}
    nomes_existentes = set(df_display.columns)
    nomes_ja_usados: set[str] = set()

    for col_original in df_display.columns:
        if col_original in RENOMEAR_COLUNAS:
            novo_nome = RENOMEAR_COLUNAS[col_original]
            if novo_nome == col_original or (
                novo_nome not in nomes_existentes
                and novo_nome not in nomes_ja_usados
            ):
                colunas_para_renomear[col_original] = novo_nome
                nomes_ja_usados.add(novo_nome)

    df_display = df_display.rename(columns=colunas_para_renomear)

    if fmt:
        fmt = {colunas_para_renomear.get(k, k): v for k, v in fmt.items()}
    if color_col:
        color_col = colunas_para_renomear.get(color_col, color_col)

    colunas_int_originais = [
        "Executada", "Não Executada", "Pendente", "Baixadas",
        "Total Alocado", "Projeção", "Alocado", "Considerado",
        "OS", "GPON", "ND", "PME", "Migração",
        "Qtd_4K", "Ultra", "Soundbox", "Equipe", "TOTAL_TAREFAS",
    ]
    for c in [colunas_para_renomear.get(c, c) for c in colunas_int_originais]:
        if c in df_display.columns:
            df_display[c] = (
                pd.to_numeric(df_display[c], errors="coerce").fillna(0).astype(int)
            )

    styler = df_display.style
    if fmt:
        styler = styler.format(fmt)

    if color_col and color_col in df_display.columns and color_meta is not None:
        def _cor(val: Any) -> str:
            try:
                v = float(val)
            except (ValueError, TypeError):
                return ""
            if color_invertido:
                if v > color_meta:
                    return "background-color:#FEE2E2;color:#991B1B;font-weight:600;"
                if v > color_meta * 0.85:
                    return "background-color:#FEF9C3;color:#854D0E;font-weight:600;"
                return "background-color:#DCFCE7;color:#166534;font-weight:600;"
            else:
                if v >= color_meta:
                    return "background-color:#DCFCE7;color:#166534;font-weight:600;"
                if v >= color_meta * 0.85:
                    return "background-color:#FEF9C3;color:#854D0E;font-weight:600;"
                return "background-color:#FEE2E2;color:#991B1B;font-weight:600;"

        styler = styler.map(_cor, subset=pd.Index([color_col]))

    styler = styler.set_table_styles([
        {"selector": "th", "props": [
            ("background-color", "#0F172A"), ("color", "#FFFFFF"),
            ("font-size", "0.78rem"), ("font-weight", "700"),
            ("text-transform", "uppercase"), ("letter-spacing", "0.03em"),
            ("padding", "0.6rem 0.8rem"), ("border", "none"),
        ]},
        {"selector": "td", "props": [
            ("font-size", "0.82rem"), ("padding", "0.5rem 0.8rem"),
            ("border-bottom", "1px solid #F1F5F9"),
        ]},
        {"selector": "tr:hover td", "props": [("background-color", "#F8FAFC")]},
    ])

    st.dataframe(styler, use_container_width=True, hide_index=True, height="auto")


def gerar_excel(df: pd.DataFrame, aba: str = "Dados") -> bytes:
    out = BytesIO()
    with pd.ExcelWriter(out, engine="openpyxl") as w:
        df.to_excel(w, index=False, sheet_name=aba[:31])
        ws = w.sheets[aba[:31]]
        hf = PatternFill("solid", fgColor="0F172A")
        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(color="FFFFFF", bold=True)
        for i in range(1, len(df.columns) + 1):
            ws.column_dimensions[get_column_letter(i)].width = 20
    return out.getvalue()


# ====================================================
# 5. CARREGAMENTO DE DADOS
# ====================================================
@st.cache_data(ttl=600, show_spinner=False)
def buscar_google_sheets() -> pd.DataFrame:
    try:
        conn = st.connection("gsheets", type=GSheetsConnection)
        df = conn.read(
            spreadsheet=URL_GSHEETS,
            usecols=["Login", "Técnico", "Monitor", "Base"],
        )
        df = df.dropna(subset=["Login"])
        df["Login"] = (
            df["Login"].astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip().str.upper()
        )
        return df
    except Exception:
        return pd.DataFrame(columns=["Login", "Técnico", "Monitor", "Base"])


@st.cache_data(show_spinner=False)
def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
    bio = BytesIO(file_bytes)
    try:
        if filename.lower().endswith(".csv"):
            try:
                return pd.read_csv(
                    bio, sep=None, engine="python", encoding="utf-8-sig", dtype=str
                )
            except UnicodeDecodeError:
                bio.seek(0)
                return pd.read_csv(
                    bio, sep=None, engine="python", encoding="latin1", dtype=str
                )
        return pd.read_excel(bio, engine="openpyxl", dtype=str)
    except Exception as e:
        st.error(f"Erro ao ler arquivo: {e}")
        return pd.DataFrame()


@st.cache_data(show_spinner=False)
def processar_base(df_bruto: pd.DataFrame, df_ativos: pd.DataFrame) -> pd.DataFrame:
    if not isinstance(df_bruto, pd.DataFrame) or df_bruto.empty:
        return pd.DataFrame()

    _RE_REMOVER_FID = re.compile(r"\bFID\s*(12M|24M)\b", re.IGNORECASE)

    df = df_bruto.copy()
    df.columns = df.columns.astype(str).str.strip().str.upper()

    mapa = {
        "CONTRATO": ["CONTRATO"],
        "LOGIN_TECNICO": ["LOGIN DO TÉCNICO", "LOGIN DO TECNICO"],
        "STATUS_ATIVIDADE": ["STATUS DA ATIVIDADE"],
        "STATUS_PRODUTO": [
            "STATUS DO PRODUTO", "STATUS PRODUTO",
            "SITUACAO DO PRODUTO", "SITUAÇÃO DO PRODUTO", "STATUS_ITEM",
        ],
        "TOTAL_TAREFAS": ["TOTAL DE TAREFAS"],
        "TIPO_OS": ["TIPO O.S 1"],
        "HABILIDADE": ["HABILIDADE DE TRABALHO"],
        "PRODUTO": ["PRODUTO"],
        "INTERVALO": ["INTERVALO DE TEMPO", "INTERVALO"],
        "CIDADE": ["CIDADE"],
        "COORD_X": ["COORDENADA X", "LONGITUDE", "LON"],
        "COORD_Y": ["COORDENADA Y", "LATITUDE", "LAT"],
    }
    for padrao, variacoes in mapa.items():
        col = next((c for c in df.columns if c in variacoes), None)
        if col:
            df = df.rename(columns={col: padrao})
        else:
            df[padrao] = np.nan

    contrato = (
        df["CONTRATO"].astype("string")
        .str.replace("\u00a0", " ", regex=False).str.strip()
    )
    mask_vazio = (
        contrato.isna() | contrato.eq("")
        | contrato.str.upper().isin(CONTRATO_VALORES_VAZIOS)
    )
    removidos = int(mask_vazio.sum())
    df = df.loc[~mask_vazio].copy()
    df["CONTRATO"] = contrato.loc[df.index].str.upper()

    if df.empty:
        return pd.DataFrame()

    df["LOGIN_TECNICO"] = (
        df["LOGIN_TECNICO"].astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip().str.upper()
    )
    df["TOTAL_TAREFAS"] = (
        pd.to_numeric(
            df["TOTAL_TAREFAS"].astype(str).str.replace(",", "."), errors="coerce"
        ).fillna(1).astype(int)
    )
    df["STATUS_ATIVIDADE"] = (
        df["STATUS_ATIVIDADE"].astype(str).str.strip().str.upper()
    )
    if "STATUS_PRODUTO" in df.columns:
        df["STATUS_PRODUTO"] = (
            df["STATUS_PRODUTO"].astype(str).str.strip().str.upper()
        )

    if isinstance(df_ativos, pd.DataFrame) and not df_ativos.empty:
        df = df.drop(
            columns=[c for c in ["Técnico", "Monitor", "Base"] if c.upper() in df.columns],
            errors="ignore",
        )
        df = df.merge(df_ativos, left_on="LOGIN_TECNICO", right_on="Login", how="left")
        df = df.rename(columns={"Técnico": "NOME_OFICIAL"})

    df["NOME_OFICIAL"] = df.get("NOME_OFICIAL", df["LOGIN_TECNICO"]).fillna(
        df["LOGIN_TECNICO"]
    )
    df["Monitor"] = df.get("Monitor", pd.Series(dtype=str)).fillna("SEM MONITOR")

    hab = df["HABILIDADE"].astype(str).str.upper()
    tipo = df["TIPO_OS"].astype(str).str.upper()

    prod = df["PRODUTO"].astype(str).str.upper()
    prod = prod.str.replace(_RE_REMOVER_FID, "", regex=True)
    prod = prod.str.replace(r"\s{2,}", " ", regex=True).str.strip()
    df["PRODUTO"] = prod

    df["Check_GPON"] = hab.str.contains(r"PON\(1/100\)", regex=True, na=False)
    df["Check_ND"] = tipo.str.contains("ADESAO", na=False)
    df["Check_Migracao"] = (tipo.str.strip() == "24 - MUDANCA DE PACOTE") & df["Check_GPON"]
    df["Check_PME"] = df["Check_ND"] & hab.str.contains("PME", na=False)
    df["Check_Streaming"] = hab.str.contains("TV VAS(1/100)", na=False)
    df["Check_Ponto_Ultra"] = hab.str.contains("NETLAR", na=False)
    df["Check_4K"] = prod.str.contains("4K", na=False)
    df["Check_Soundbox"] = prod.str.contains("SOUND", na=False)

    df = atribuir_velocidade_por_contrato(df)

    df["PERIODO_TRATADO"] = (
        df["INTERVALO"].astype(str).str.strip()
        .map(MAPEAMENTO_PERIODOS).fillna("Outros/Sem Período")
    )

    cidade = (
        df["CIDADE"].fillna("").astype(str).str.strip().str.upper()
        .apply(
            lambda v: unicodedata.normalize("NFKD", v)
            .encode("ASCII", "ignore").decode()
        )
    )
    df["REGIÃO"] = np.select(
        [
            cidade.isin(["SAO PAULO"]),
            cidade.isin([
                "GUARULHOS", "ARUJA", "MOGI DAS CRUZES", "SUZANO",
                "ITAQUAQUECETUBA", "FERRAZ DE VASCONCELOS", "POA",
            ]),
            cidade.isin([
                "SANTO ANDRE", "SAO BERNARDO DO CAMPO", "SAO CAETANO DO SUL",
                "DIADEMA", "MAUA", "RIBEIRAO PIRES", "RIO GRANDE DA SERRA",
            ]),
        ],
        ["LESTE", "GRU", "ABCDM"],
        default="OUTRAS",
    )

    df.attrs["diagnostico"] = {"contrato_vazio": removidos}
    return df


# ====================================================
# 6. TABELAS ROTA/TURNO
# ====================================================
def _fmt_num_br(v: Any) -> str:
    try:
        return f"{int(v):,}".replace(",", ".")
    except (ValueError, TypeError):
        return "0"


def _fmt_media_br(v: Any) -> str:
    try:
        return f"{float(v):.2f}".replace(".", ",")
    except (ValueError, TypeError):
        return "0,00"


def calcular_tabela_rota_turno(
    df: pd.DataFrame,
    turno: Optional[str] = None,
    total_equipe_montada: Optional[int] = None,
) -> pd.DataFrame:
    if df.empty:
        return pd.DataFrame()

    df_work = df.copy()
    if turno:
        df_work = df_work[df_work["PERIODO_TRATADO"] == turno].copy()
    if df_work.empty:
        return pd.DataFrame()

    df_work = df_work[
        ~df_work["Monitor"].astype(str).str.upper()
        .isin({"NAN", "SEM MONITOR", "NÃO MAPEADO", ""})
    ].copy()
    if df_work.empty:
        return pd.DataFrame()

    df_work["TOTAL_TAREFAS"] = (
        pd.to_numeric(df_work["TOTAL_TAREFAS"], errors="coerce").fillna(1).astype(int)
    )

    linhas: List[Dict[str, Any]] = []
    for mon in sorted(df_work["Monitor"].unique()):
        df_mon = df_work[df_work["Monitor"] == mon]
        total_os = int(df_mon["TOTAL_TAREFAS"].sum())
        equipe = int(df_mon["LOGIN_TECNICO"].nunique())
        linhas.append({
            "Monitor": mon,
            "WO": int(df_mon["Check_ND"].sum()),
            "GPON": int(df_mon["Check_GPON"].sum()),
            "OS": total_os,
            "ND": int(df_mon["Check_ND"].sum()),
            "Migração": int(df_mon["Check_Migracao"].sum()),
            "Equipe": equipe,
            "Média": total_os / equipe if equipe > 0 else 0.0,
        })

    df_out = pd.DataFrame(linhas)
    total_os_g = int(df_out["OS"].sum())
    total_eq_escalados = int(df_out["Equipe"].sum())
    total_eq_montados = (
        total_equipe_montada if total_equipe_montada is not None
        else int(df_work["LOGIN_TECNICO"].nunique())
    )

    df_out = pd.concat([
        df_out,
        pd.DataFrame([
            {
                "Monitor": "Total Geral | Escalados",
                "WO": int(df_out["WO"].sum()), "GPON": int(df_out["GPON"].sum()),
                "OS": total_os_g, "ND": int(df_out["ND"].sum()),
                "Migração": int(df_out["Migração"].sum()),
                "Equipe": total_eq_escalados,
                "Média": total_os_g / total_eq_escalados if total_eq_escalados else 0.0,
            },
            {
                "Monitor": "Total Geral | Montados",
                "WO": int(df_out["WO"].sum()), "GPON": int(df_out["GPON"].sum()),
                "OS": total_os_g, "ND": int(df_out["ND"].sum()),
                "Migração": int(df_out["Migração"].sum()),
                "Equipe": total_eq_montados,
                "Média": total_os_g / total_eq_montados if total_eq_montados else 0.0,
            },
        ]),
    ], ignore_index=True)

    return df_out


def render_tabela_rota_turno(df: pd.DataFrame, titulo: str) -> str:
    if df.empty:
        return (
            f'<div class="rota-wrapper"><div class="rota-titulo">{titulo}</div>'
            f'<table class="rota-tab"><tr>'
            f'<td colspan="8" style="padding:20px;color:#64748B;text-align:center;">'
            f"Sem dados disponíveis</td></tr></table></div>"
        )

    linhas_html: List[str] = []
    for _, row in df.iterrows():
        monitor = str(row["Monitor"])
        classe = (
            "total-escalados" if "Escalados" in monitor
            else "total-montados" if "Montados" in monitor
            else ""
        )
        linhas_html.append(
            f'<tr class="{classe}">'
            f'<td class="col-monitor">{monitor}</td>'
            f"<td>{_fmt_num_br(row['WO'])}</td>"
            f"<td>{_fmt_num_br(row['GPON'])}</td>"
            f'<td class="col-os">{_fmt_num_br(row["OS"])}</td>'
            f"<td>{_fmt_num_br(row['ND'])}</td>"
            f"<td>{_fmt_num_br(row['Migração'])}</td>"
            f'<td class="col-equipe">{_fmt_num_br(row["Equipe"])}</td>'
            f'<td class="col-media">{_fmt_media_br(row["Média"])}</td>'
            f"</tr>"
        )

    return (
        f'<div class="rota-wrapper">'
        f'<div class="rota-titulo">{titulo}</div>'
        f'<table class="rota-tab"><thead><tr>'
        f'<th style="width:32%;">Monitor</th>'
        f"<th>WO</th><th>GPON</th>"
        f'<th class="th-os">OS</th>'
        f"<th>ND</th><th>Migração</th>"
        f'<th class="th-equipe">Equipe</th>'
        f'<th class="th-media">Média</th>'
        f"</tr></thead>"
        f'<tbody>{"".join(linhas_html)}</tbody>'
        f"</table></div>"
    )


def render_bloco_rota_turno(df_master: pd.DataFrame, total_montados: int) -> None:
    from datetime import datetime
    data_hoje = datetime.now().strftime("%d/%m/%Y")
    total_efetivo = (
        total_montados if total_montados > 0
        else int(df_master["LOGIN_TECNICO"].nunique())
    )
    for turno, label in [
        (None, f"Rota Inicial — {data_hoje}"),
        ("Manhã", f"Manhã — {data_hoje}"),
        ("Tarde I", f"Tarde I — {data_hoje}"),
        ("Tarde II", f"Tarde II — {data_hoje}"),
    ]:
        df_t = calcular_tabela_rota_turno(
            df_master, turno=turno, total_equipe_montada=total_efetivo
        )
        st.markdown(render_tabela_rota_turno(df_t, label), unsafe_allow_html=True)


# ====================================================
# 7. APLICAÇÃO PRINCIPAL
# ====================================================
def main() -> None:
    aplicar_estilo()
    _injetar_css_local()

    render_hero(
        titulo="🗺️ Gestão de Rota Inicial",
        subtitulo="Visão operacional de rotas, serviços premium e distribuição de equipes",
    )

    # ── Sidebar — configurações ──────────────────────
    with st.sidebar:
        st.header("⚙️ Configurações")
        if st.button("🔄 Reiniciar Painel", use_container_width=True):
            st.session_state["df_master"] = None
            if "sim_equipe" in st.session_state:
                del st.session_state["sim_equipe"]
            st.session_state["total_montados_manual"] = 0
            st.session_state["_reset_counter"] = (
                int(st.session_state.get("_reset_counter", 0)) + 1
            )
            st.rerun()
        st.divider()

    reset_key = int(st.session_state.get("_reset_counter", 0))

    # ── Upload ───────────────────────────────────────
    if st.session_state["df_master"] is None:
        render_section("📁 Importação de Dados")
        arq = st.file_uploader(
            "Selecione a base (Excel/CSV)",
            type=["xlsx", "csv"],
            key=f"upload_base_{reset_key}",
        )
        if arq:
            with st.spinner("Processando..."):
                raw = ler_arquivo(arq.getvalue(), arq.name)
                gs = buscar_google_sheets()
                df_proc = processar_base(raw, gs)
                st.session_state["df_master"] = df_proc

            diag = df_proc.attrs.get("diagnostico", {})
            if diag.get("contrato_vazio", 0) > 0:
                st.toast(
                    f"🗑️ {diag['contrato_vazio']} linha(s) sem contrato removida(s).",
                    icon="⚠️",
                )
            st.rerun()
        return

    df_master = st.session_state["df_master"].copy()

    # ── Sidebar — equipe + filtros ───────────────────
    with st.sidebar:
        st.header("👥 Equipe Montada")
        total_montados_input = st.number_input(
            "Total de Técnicos Montados",
            min_value=0, max_value=999,
            value=st.session_state["total_montados_manual"],
            step=1,
            help="Deixe em 0 para usar a contagem automática da base.",
            key=f"input_total_montados_{reset_key}",
        )
        st.session_state["total_montados_manual"] = total_montados_input

        if total_montados_input > 0:
            st.caption(f"✅ Usando **{total_montados_input}** técnicos montados (manual)")
        else:
            auto_count = int(df_master["LOGIN_TECNICO"].nunique())
            st.caption(f"🔄 Usando **{auto_count}** técnicos (automático da base)")

        st.divider()
        st.header("🎯 Filtros")
        periodos = st.selectbox(
            "⏰ Período",
            ["Todos"] + sorted(df_master["PERIODO_TRATADO"].unique()),
            key=f"filtro_periodo_{reset_key}",
        )
        if periodos != "Todos":
            df_master = df_master[df_master["PERIODO_TRATADO"] == periodos]

        st.divider()
        st.subheader("🎛️ Filtros Premium")

        if st.checkbox("🟢 Apenas Adesão (ND)", key=f"chk_nd_{reset_key}"):
            df_master = df_master[df_master["Check_ND"]]

        if st.checkbox(
            "🏢 Apenas PME (Empresarial)",
            help="Adesões (ND) cuja habilidade contém 'PME'.",
            key=f"chk_pme_{reset_key}",
        ):
            df_master = df_master[df_master["Check_PME"]]

        if st.checkbox("🔄 Apenas Migração (MP GPON)", key=f"chk_mig_{reset_key}"):
            df_master = df_master[df_master["Check_Migracao"]]

        if st.checkbox("📡 Requer GPON", key=f"chk_gpon_{reset_key}"):
            df_master = df_master[df_master["Check_GPON"]]

        if st.checkbox("📺 Requer Streaming", key=f"chk_stream_{reset_key}"):
            df_master = df_master[df_master["Check_Streaming"]]

        if st.checkbox("📺 Requer 4K", key=f"chk_4k_{reset_key}"):
            df_master = df_master[df_master["Check_4K"]]

        if st.checkbox("🔌 Requer Ponto Ultra", key=f"chk_ultra_{reset_key}"):
            df_master = df_master[df_master["Check_Ponto_Ultra"]]

        if st.checkbox("🔊 Requer Soundbox", key=f"chk_sound_{reset_key}"):
            df_master = df_master[df_master["Check_Soundbox"]]

        if st.checkbox(
            "📉 Baixa Velocidade",
            help="Exibe apenas contratos com velocidade identificada inferior a 400 Mbps. "
                 "Contratos sem velocidade são excluídos.",
            key=f"filtro_vel_menor_400_{reset_key}",
        ):
            def _vel_para_mbps(label: str) -> float:
                if not isinstance(label, str) or not label.strip():
                    return np.nan
                try:
                    if "Gbps" in label:
                        return float(label.replace("Gbps", "").replace(",", ".").strip()) * 1000
                    if "Mbps" in label:
                        return float(label.replace("Mbps", "").replace(",", ".").strip())
                except ValueError:
                    pass
                return np.nan

            mbps_serie = df_master["VELOCIDADE_BANDA"].apply(_vel_para_mbps)
            df_master = df_master[mbps_serie.notna() & (mbps_serie < 400)]

    # ── Guard ────────────────────────────────────────
    if df_master.empty:
        render_insight("Nenhum dado para os filtros selecionados.", tipo="alerta")
        return

    # ── Resultado da base ────────────────────────────
    regioes = (
        sorted(df_master["REGIÃO"].unique()) if "REGIÃO" in df_master.columns else []
    )
    render_resultado_base(regioes, len(df_master))

    # ── KPIs Principais ─────────────────────────────
    soma_os = int(df_master["TOTAL_TAREFAS"].sum())
    tecnicos = df_master["LOGIN_TECNICO"].nunique()
    monitores_qtd = df_master["Monitor"].nunique()

    render_section_header("📊", "Indicadores Operacionais")
    c1, c2, c3 = st.columns(3)
    render_kpi(c1, "Volume O.S.", f"{soma_os:,}", tema="azul")
    render_kpi(
        c2, "Técnicos Operando", f"{tecnicos}",
        sub=f"Média: {soma_os / tecnicos:.1f} O.S./Téc." if tecnicos else "",
        tema="cinza",
    )
    render_kpi(c3, "Monitores", f"{monitores_qtd}", tema="azul")

    st.markdown("")

    # ── KPIs Secundários ────────────────────────────
    qtd_gpon = int(df_master["Check_GPON"].sum())
    qtd_nd = int(df_master["Check_ND"].sum())
    qtd_pme = int(df_master["Check_PME"].sum())
    qtd_4k = int(df_master["Check_4K"].sum())
    qtd_sound = int(df_master["Check_Soundbox"].sum())

    s1, s2, s3, s4, s5 = st.columns(5)
    render_kpi(s1, "GPON", f"{qtd_gpon:,}", tema="verde")
    render_kpi(s2, "Adesão (ND)", f"{qtd_nd:,}", tema="azul")
    render_kpi(
        s3, "PME", f"{qtd_pme:,}",
        sub=f"{qtd_pme / qtd_nd:.1%} das ND" if qtd_nd > 0 else "sem ND",
        tema="azul",
    )
    render_kpi(s4, "4K", f"{qtd_4k:,}", tema="laranja")
    render_kpi(s5, "Soundbox", f"{qtd_sound:,}", tema="cinza")

    st.markdown("")

    # ── Tabelas Rota / Turnos ────────────────────────
    render_section_header("🗺️", "Distribuição por Monitor e Turno")
    render_bloco_rota_turno(df_master, st.session_state["total_montados_manual"])

    # ── Gráficos Executivos ──────────────────────────
    render_section_header("📈", "Visão Executiva")
    g1, g2 = st.columns([1, 1.2])

    with g1:
        df_per = (
            df_master.groupby("PERIODO_TRATADO")["TOTAL_TAREFAS"].sum().reset_index()
        )
        fig_per = px.bar(
            df_per, x="PERIODO_TRATADO", y="TOTAL_TAREFAS",
            text_auto=True, color="PERIODO_TRATADO",
            color_discrete_sequence=[COR_PRIMARIA, COR_SECUNDARIA, COR_SUCESSO, COR_NEUTRO],
        )
        fig_per.update_layout(
            showlegend=False, margin=dict(t=30, b=0, l=0, r=0),
            height=320, xaxis_title="", yaxis_title="",
            title=dict(text="Pico de Agendamento", font=dict(size=14)),
        )
        st.plotly_chart(fig_per, use_container_width=True)

    with g2:
        df_prem = pd.DataFrame([
            {"Serviço": "GPON", "Qtd": qtd_gpon},
            {"Serviço": "PME", "Qtd": qtd_pme},
            {"Serviço": "4K", "Qtd": qtd_4k},
            {"Serviço": "Soundbox", "Qtd": qtd_sound},
            {"Serviço": "Ponto Ultra", "Qtd": int(df_master["Check_Ponto_Ultra"].sum())},
        ])
        df_prem = df_prem[df_prem["Qtd"] > 0]
        if not df_prem.empty:
            fig_prem = px.bar(
                df_prem, x="Qtd", y="Serviço",
                orientation="h", text_auto=True,
                color_discrete_sequence=[COR_SUCESSO],
            )
            fig_prem.update_layout(
                showlegend=False, margin=dict(t=30, b=0, l=0, r=0),
                height=320, xaxis_title="", yaxis_title="",
                title=dict(text="Mix Premium", font=dict(size=14)),
            )
            st.plotly_chart(fig_prem, use_container_width=True)
        else:
            render_insight("Nenhum serviço premium identificado.", tipo="info")

    # ── Abas de Detalhamento ─────────────────────────
    aba_tec, aba_mapa, aba_base, aba_contratos, aba_equalizacao = st.tabs(
        ["🏆 Top Técnicos", "🗺️ Mapa", "🗃️ Base Completa",
         "📄 Resumo Contratos", "⚖️ Equalização de Rota"]
    )

    # ============================================================
    # ABA — TOP TÉCNICOS
    # ============================================================
    with aba_tec:
        prod_df = (
            df_master.groupby("NOME_OFICIAL")
            .agg({"TOTAL_TAREFAS": "sum"})
            .reset_index()
            .sort_values("TOTAL_TAREFAS", ascending=False)
            .head(15)
        )
        fig_tec = px.bar(
            prod_df, x="TOTAL_TAREFAS", y="NOME_OFICIAL",
            orientation="h", color="TOTAL_TAREFAS",
            color_continuous_scale="Blues", text_auto=True,
        )
        fig_tec.update_layout(
            yaxis={"categoryorder": "total ascending"},
            height=500,
            title=dict(text="Top 15 Técnicos por Volume", font=dict(size=15)),
            margin=dict(t=50, b=10, l=10, r=10),
        )
        st.plotly_chart(fig_tec, use_container_width=True)

    # ============================================================
    # ABA — MAPA
    # ============================================================
    with aba_mapa:
        df_mapa = df_master.dropna(subset=["COORD_X", "COORD_Y"]).copy()
        if (
            not df_mapa.empty
            and df_mapa["COORD_X"].astype(str).str.contains(r"\d").any()
        ):
            df_mapa["COORD_X"] = pd.to_numeric(
                df_mapa["COORD_X"].astype(str).str.replace(",", "."), errors="coerce"
            )
            df_mapa["COORD_Y"] = pd.to_numeric(
                df_mapa["COORD_Y"].astype(str).str.replace(",", "."), errors="coerce"
            )
            df_mapa = df_mapa.dropna(subset=["COORD_X", "COORD_Y"])
            if not df_mapa.empty:
                fig_mapa = px.scatter_mapbox(
                    df_mapa, lat="COORD_Y", lon="COORD_X",
                    color="STATUS_ATIVIDADE",
                    zoom=9, height=550, hover_name="NOME_OFICIAL",
                )
                fig_mapa.update_layout(
                    mapbox_style="open-street-map",
                    margin={"r": 0, "t": 0, "l": 0, "b": 0},
                )
                st.plotly_chart(fig_mapa, use_container_width=True)
            else:
                render_insight("Coordenadas GPS não encontradas após limpeza.", tipo="info")
        else:
            render_insight(
                "A planilha não possui coordenadas GPS válidas para o mapa.", tipo="info"
            )

    # ============================================================
    # ABA — BASE COMPLETA
    # ============================================================
    with aba_base:
        render_dataframe_local(
            df_master.head(500),
            titulo="Base de Dados (prévia — 500 linhas)",
            icone="🗃️",
            badge=f"{len(df_master)} total",
            height="auto",
        )
        st.download_button(
            "📥 Baixar Base Completa",
            gerar_excel(df_master, "Base"),
            "base_completa.xlsx",
        )

    # ============================================================
    # ABA — RESUMO CONTRATOS
    # ============================================================
    with aba_contratos:
        render_section_header("📄", "Resumo dos Contratos da Rota")

        COLUNAS_CONTRATOS: Dict[str, List[str]] = {
            "CONTRATO": ["CONTRATO"],
            "INTERVALO": ["PERIODO_TRATADO", "INTERVALO"],
            "CEP": ["CEP/CÓDIGO POSTAL", "CEP", "CODIGO POSTAL"],
            "ÁREA TRABALHO": ["ÁREA DE TRABALHO", "AREA DE TRABALHO"],
            "TIPO OS": ["TIPO_OS", "TIPO O.S 1", "TIPO OS 1"],
            "TÉCNICO": ["NOME_OFICIAL"],
            "MONITOR": ["Monitor"],
            "VELOCIDADE": ["VELOCIDADE_BANDA"],
        }

        colunas_encontradas: Dict[str, str] = {}
        for nome_amigavel, aliases in COLUNAS_CONTRATOS.items():
            for alias in aliases:
                if alias in df_master.columns:
                    colunas_encontradas[alias] = nome_amigavel
                    break

        if not colunas_encontradas:
            render_insight(
                "Nenhuma das colunas esperadas foi encontrada na base.", tipo="alerta"
            )
        else:
            df_contratos = (
                df_master[list(colunas_encontradas.keys())]
                .rename(columns=colunas_encontradas)
                .copy()
            )

            col_f1, col_f2, col_info = st.columns([2, 2, 3])

            with col_f1:
                mon_contratos = ["Todos"] + sorted(
                    str(x)
                    for x in df_contratos.get("MONITOR", pd.Series()).dropna().unique()
                    if str(x) not in {"nan", "SEM MONITOR", "NÃO MAPEADO"}
                )
                sel_mon = st.selectbox(
                    "👔 Monitor", mon_contratos,
                    key=f"filtro_mon_contratos_{reset_key}",
                )

            df_cf = (
                df_contratos[df_contratos["MONITOR"] == sel_mon].copy()
                if sel_mon != "Todos"
                else df_contratos.copy()
            )

            with col_f2:
                tec_contratos = ["Todos"] + sorted(
                    str(x)
                    for x in df_cf.get("TÉCNICO", pd.Series()).dropna().unique()
                    if str(x) not in {"nan", "NÃO MAPEADO"}
                )
                sel_tec = st.selectbox(
                    "👤 Técnico", tec_contratos,
                    key=f"filtro_tec_contratos_{reset_key}",
                )

            if sel_tec != "Todos":
                df_cf = df_cf[df_cf["TÉCNICO"] == sel_tec].copy()

            with col_info:
                st.markdown("")
                st.markdown(
                    f"**{len(df_cf):,}** contratos exibidos de **{len(df_contratos):,}** total"
                )

            render_dataframe_local(
                df_cf,
                titulo="Resumo Contratos",
                icone="📄",
                badge=f"{len(df_cf)} contratos",
                height="auto",
            )
            st.download_button(
                "📥 Baixar Resumo Contratos",
                gerar_excel(df_cf, "Resumo Contratos"),
                "resumo_contratos.xlsx",
            )

    # ============================================================
    # ABA — EQUALIZAÇÃO
    # ============================================================
        # ============================================================
    # ABA — EQUALIZAÇÃO (por distribuição de O.S.)
    # ============================================================
    with aba_equalizacao:
        render_section_header(
            "⚖️",
            "Diagnóstico de Equalização por Distribuição de O.S."
        )

        df_eq_work = df_master[
            ~df_master["Monitor"].astype(str).str.upper()
            .isin({"NAN", "SEM MONITOR", "NÃO MAPEADO", ""})
        ].copy()

        if df_eq_work.empty:
            render_insight(
                "Dados insuficientes para calcular a equalização.",
                tipo="alerta",
            )
        else:
            # ============================================================
            # 1. CONFIGURAÇÃO — TOLERÂNCIA
            # ============================================================
            with st.expander("⚙️ Configurações de Equalização", expanded=False):
                col_cfg1, col_cfg2 = st.columns(2)

                with col_cfg1:
                    tolerancia = st.slider(
                        "Tolerância aceita (± %)",
                        min_value=5,
                        max_value=30,
                        value=15,
                        step=5,
                        help=(
                            "Faixa de desvio considerada aceitável em "
                            "relação à carga ideal de O.S. por monitor."
                        ),
                        key=f"eq_tolerancia_{reset_key}",
                    )

                with col_cfg2:
                    st.markdown(" ")
                    st.markdown(" ")
                    st.markdown(
                        f"📏 **Faixa aceitável:** desvio entre "
                        f"**−{tolerancia}%** e **+{tolerancia}%** da carga ideal."
                    )

            # ============================================================
            # 2. AGRUPAMENTO POR MONITOR
            # ============================================================
            df_eq_mon = (
                df_eq_work.groupby("Monitor")
                .agg(
                    OS_Atual=("TOTAL_TAREFAS", "sum"),
                    Equipe=("LOGIN_TECNICO", "nunique"),
                )
                .reset_index()
            )

            # ============================================================
            # 3. MÉTRICAS BASE
            # ============================================================
            total_os_g = float(df_eq_mon["OS_Atual"].sum())
            total_eq_g = int(df_eq_mon["Equipe"].sum())

            # Meta de produtividade global (O.S. por técnico)
            os_por_tecnico_ideal = (
                total_os_g / total_eq_g if total_eq_g > 0 else 0.0
            )

            # KPIs de topo
            k1, k2, k3 = st.columns(3)
            with k1:
                render_kpi_sm(
                    st, "Total O.S.",
                    f"{int(total_os_g):,}".replace(",", "."),
                    "Volume a distribuir", "azul",
                )
            with k2:
                render_kpi_sm(
                    st, "Total Técnicos",
                    f"{total_eq_g}",
                    "Equipe fixa", "laranja",
                )
            with k3:
                render_kpi_sm(
                    st, "Meta O.S./Técnico",
                    f"{os_por_tecnico_ideal:.1f}".replace(".", ","),
                    "Produtividade alvo", "verde",
                )

            st.markdown(
                f"🎯 **Meta:** distribuir as O.S. de modo que **cada técnico "
                f"execute ~{os_por_tecnico_ideal:.1f} O.S.** "
                f"(tolerância ±{tolerancia}%). "
                f"A equipe atual de cada monitor é mantida fixa."
            )

            # ============================================================
            # 4. CÁLCULO DE EQUALIZAÇÃO POR O.S.
            # ============================================================
            # OS Ideal = meta por técnico × equipe atual do monitor
            df_eq_mon["OS Ideal"] = (
                df_eq_mon["Equipe"] * os_por_tecnico_ideal
            ).round(0).astype(int)

            # Diferença = quantas O.S. sobram (+) ou faltam (−) no monitor
            df_eq_mon["Balanço (O.S.)"] = (
                df_eq_mon["OS_Atual"] - df_eq_mon["OS Ideal"]
            ).astype(int)

            # Média atual de O.S. por técnico
            df_eq_mon["Média Atual"] = (
                df_eq_mon["OS_Atual"] / df_eq_mon["Equipe"].replace(0, np.nan)
            ).fillna(0)

            # Desvio % vs meta
            df_eq_mon["Desvio %"] = np.where(
                os_por_tecnico_ideal > 0,
                ((df_eq_mon["Média Atual"] - os_por_tecnico_ideal)
                 / os_por_tecnico_ideal) * 100,
                0.0,
            )

            def classificar_status(desvio: float) -> str:
                if desvio > tolerancia:
                    return "🔴 Excesso de O.S."
                if desvio < -tolerancia:
                    return "🟡 Faltam O.S."
                return "🟢 Equilibrado"

            df_eq_mon["Status"] = df_eq_mon["Desvio %"].apply(classificar_status)

            df_eq_mon = df_eq_mon.sort_values(
                "Balanço (O.S.)", ascending=False
            ).reset_index(drop=True)

            # ============================================================
            # 5. TABELA CONSOLIDADA
            # ============================================================
            df_display_eq = df_eq_mon[[
                "Monitor", "Equipe", "OS_Atual", "OS Ideal",
                "Balanço (O.S.)", "Média Atual", "Desvio %", "Status",
            ]].rename(columns={"OS_Atual": "OS Atual"})

            render_dataframe_local(
                df_display_eq,
                titulo="Balanço de Distribuição de O.S. por Monitor",
                icone="📦",
                badge="Equalização por Volume",
                fmt={
                    "Equipe": "{:d}",
                    "OS Atual": "{:d}",
                    "OS Ideal": "{:d}",
                    "Balanço (O.S.)": "{:+d}",
                    "Média Atual": "{:.1f}",
                    "Desvio %": "{:+.1f}%",
                },
                color_col="Média Atual",
                color_meta=os_por_tecnico_ideal,
                height="auto",
            )

            # ============================================================
            # 6. GRÁFICO — O.S. ATUAL vs O.S. IDEAL
            # ============================================================
            st.subheader("📊 Distribuição de O.S. — Atual vs Ideal")

            fig = go.Figure()

            # Barras: O.S. Atual
            fig.add_trace(go.Bar(
                name="O.S. Atual",
                x=df_eq_mon["Monitor"],
                y=df_eq_mon["OS_Atual"],
                marker_color=[
                    "#EF4444" if d > tolerancia
                    else "#F59E0B" if d < -tolerancia
                    else "#10B981"
                    for d in df_eq_mon["Desvio %"]
                ],
                text=[f"{v}" for v in df_eq_mon["OS_Atual"]],
                textposition="outside",
            ))

            # Barras: O.S. Ideal (referência)
            fig.add_trace(go.Bar(
                name="O.S. Ideal",
                x=df_eq_mon["Monitor"],
                y=df_eq_mon["OS Ideal"],
                marker_color="#CBD5E1",
                marker_line=dict(color="#64748B", width=1),
                text=[f"{v}" for v in df_eq_mon["OS Ideal"]],
                textposition="outside",
                opacity=0.7,
            ))

            fig.update_layout(
                title=dict(
                    text="<b>O.S. Atual × O.S. Ideal (mantendo equipe fixa)</b>",
                    font=dict(family="Manrope", size=16, color="#012869"),
                ),
                barmode="group",
                xaxis_title="",
                yaxis_title="Quantidade de O.S.",
                height=420,
                margin=dict(t=60, b=40, l=40, r=60),
                plot_bgcolor="white",
                paper_bgcolor="white",
                legend=dict(
                    orientation="h",
                    yanchor="bottom", y=1.02,
                    xanchor="right", x=1,
                ),
            )
            fig.update_xaxes(
                tickfont=dict(family="Inter", size=12, color="#374151"),
            )
            fig.update_yaxes(
                gridcolor="#F1F5F9",
                tickfont=dict(family="Inter", size=11, color="#6B7280"),
            )

            st.plotly_chart(fig, use_container_width=True)

            # ============================================================
            # 7. PLANO DE REDISTRIBUIÇÃO DE O.S.
            # ============================================================
            st.subheader("💡 Plano de Redistribuição de O.S.")

            com_excesso = df_eq_mon[df_eq_mon["Balanço (O.S.)"] > 0]
            com_falta = df_eq_mon[df_eq_mon["Balanço (O.S.)"] < 0]

            if com_excesso.empty and com_falta.empty:
                render_insight(
                    "🎉 A distribuição de O.S. está **perfeitamente equilibrada**! "
                    "Todos os monitores têm carga proporcional à sua equipe.",
                    tipo="ok",
                )
            else:
                col_exc, col_fal = st.columns(2)

                with col_exc:
                    st.markdown("#### 🔴 Monitores com excesso de O.S.")
                    if com_excesso.empty:
                        st.info("Nenhum monitor com excesso.")
                    else:
                        for _, row in com_excesso.iterrows():
                            excesso = int(row["Balanço (O.S.)"])
                            st.markdown(
                                f"**{row['Monitor']}**  \n"
                                f"📤 Redistribuir **{excesso} O.S.**  \n"
                                f"📊 Atual: `{int(row['OS_Atual'])}` "
                                f"| Ideal: `{int(row['OS Ideal'])}`  \n"
                                f"👥 Equipe: {int(row['Equipe'])} téc. "
                                f"({row['Média Atual']:.1f} O.S./téc.)"
                            )
                            st.divider()

                with col_fal:
                    st.markdown("#### 🟡 Monitores com folga de capacidade")
                    if com_falta.empty:
                        st.info("Nenhum monitor com folga.")
                    else:
                        for _, row in com_falta.iterrows():
                            falta = abs(int(row["Balanço (O.S.)"]))
                            st.markdown(
                                f"**{row['Monitor']}**  \n"
                                f"📥 Pode receber **{falta} O.S.**  \n"
                                f"📊 Atual: `{int(row['OS_Atual'])}` "
                                f"| Ideal: `{int(row['OS Ideal'])}`  \n"
                                f"👥 Equipe: {int(row['Equipe'])} téc. "
                                f"({row['Média Atual']:.1f} O.S./téc.)"
                            )
                            st.divider()

                # Saldo geral
                total_excesso = int(com_excesso["Balanço (O.S.)"].sum())
                total_falta = int(com_falta["Balanço (O.S.)"].abs().sum())

                if total_excesso > 0 and total_falta > 0:
                    realocavel = min(total_excesso, total_falta)
                    render_insight(
                        f"💱 **Redistribuição interna possível:** transferir até "
                        f"**{realocavel} O.S.** dos monitores com excesso para "
                        f"os com folga "
                        f"(excesso: {total_excesso}, capacidade ociosa: {total_falta}).",
                        tipo="info",
                    )

        # ============================================================
        # 8. SIMULAÇÃO — REDISTRIBUIR O.S. MANUALMENTE
        # ============================================================
        if not df_eq_work.empty:
            st.markdown("---")
            st.subheader("🧪 Simulação: e se eu redistribuísse as O.S.?")
            st.markdown(
                "Ajuste manualmente o volume de O.S. que cada monitor deve "
                "receber e veja como fica a **carga por técnico**. "
                "O total geral é mantido automaticamente."
            )

            if "sim_os" not in st.session_state:
                st.session_state["sim_os"] = {}

            monitores_atuais = df_eq_mon["Monitor"].tolist()
            for mon in monitores_atuais:
                if mon not in st.session_state["sim_os"]:
                    st.session_state["sim_os"][mon] = int(
                        df_eq_mon.loc[
                            df_eq_mon["Monitor"] == mon, "OS_Atual"
                        ].iloc[0]
                    )

            # Botões de controle
            col_ctrl1, col_ctrl2, _ = st.columns([1, 1, 3])
            with col_ctrl1:
                if st.button(
                    "🔄 Resetar simulação",
                    key=f"btn_reset_sim_{reset_key}",
                ):
                    for mon in monitores_atuais:
                        st.session_state["sim_os"][mon] = int(
                            df_eq_mon.loc[
                                df_eq_mon["Monitor"] == mon, "OS_Atual"
                            ].iloc[0]
                        )
                    st.rerun()

            with col_ctrl2:
                if st.button(
                    "🎯 Aplicar distribuição ideal",
                    key=f"btn_ideal_sim_{reset_key}",
                ):
                    for mon in monitores_atuais:
                        st.session_state["sim_os"][mon] = int(
                            df_eq_mon.loc[
                                df_eq_mon["Monitor"] == mon, "OS Ideal"
                            ].iloc[0]
                        )
                    st.rerun()

            # Inputs de simulação — um por monitor
            cols_sim = st.columns(min(len(monitores_atuais), 4))
            for i, mon in enumerate(monitores_atuais):
                equipe_mon = int(
                    df_eq_mon.loc[
                        df_eq_mon["Monitor"] == mon, "Equipe"
                    ].iloc[0]
                )
                with cols_sim[i % len(cols_sim)]:
                    valor_input = st.number_input(
                        f"📦 {mon} ({equipe_mon} téc.)",
                        min_value=0,
                        step=1,
                        value=int(st.session_state["sim_os"][mon]),
                        key=f"sim_input_{mon}_{reset_key}",
                        help=f"Equipe fixa: {equipe_mon} técnicos",
                    )
                    st.session_state["sim_os"][mon] = int(valor_input)

            # ── Recálculo com base na simulação ──
            df_sim = df_eq_mon.copy()
            df_sim["OS Simulada"] = df_sim["Monitor"].map(
                st.session_state["sim_os"]
            )
            df_sim["Média Simulada"] = (
                df_sim["OS Simulada"]
                / df_sim["Equipe"].replace(0, np.nan)
            ).fillna(0)

            df_sim["Desvio Sim %"] = np.where(
                os_por_tecnico_ideal > 0,
                ((df_sim["Média Simulada"] - os_por_tecnico_ideal)
                 / os_por_tecnico_ideal) * 100,
                0.0,
            )
            df_sim["Status Sim"] = df_sim["Desvio Sim %"].apply(
                classificar_status
            )

            # KPIs comparativos
            total_sim_os = int(sum(st.session_state["sim_os"].values()))
            diff_total_os = total_sim_os - int(total_os_g)

            equilibrados_sim = int(
                (df_sim["Desvio Sim %"].abs() <= tolerancia).sum()
            )
            equilibrados_atual = int(
                (df_eq_mon["Desvio %"].abs() <= tolerancia).sum()
            )

            ks1, ks2, ks3 = st.columns(3)
            with ks1:
                render_kpi_sm(
                    st, "O.S. simuladas (total)",
                    f"{total_sim_os:,}".replace(",", "."),
                    f"Δ {diff_total_os:+d} vs atual",
                    "azul" if diff_total_os == 0 else "laranja",
                )
            with ks2:
                render_kpi_sm(
                    st, "Monitores equilibrados",
                    f"{equilibrados_sim} / {len(df_sim)}",
                    f"Antes: {equilibrados_atual}",
                    "verde" if equilibrados_sim >= equilibrados_atual
                    else "vermelho",
                )
            with ks3:
                variacao_max = float(df_sim["Desvio Sim %"].abs().max())
                render_kpi_sm(
                    st, "Maior desvio",
                    f"{variacao_max:.1f}%",
                    "Idealmente ≤ tolerância",
                    "verde" if variacao_max <= tolerancia else "vermelho",
                )

            # Tabela comparativa
            df_comp = df_sim[[
                "Monitor", "Equipe", "OS_Atual", "OS Simulada",
                "Média Atual", "Média Simulada",
                "Desvio %", "Desvio Sim %",
                "Status", "Status Sim",
            ]].rename(columns={
                "OS_Atual": "OS Atual",
                "Desvio %": "Desvio Atual",
                "Desvio Sim %": "Desvio Simulado",
                "Status": "Status Atual",
                "Status Sim": "Status Simulado",
            })

            render_dataframe_local(
                df_comp,
                titulo="Comparativo: Distribuição Atual × Simulada",
                icone="🧪",
                badge="Simulação Interativa",
                fmt={
                    "Equipe": "{:d}",
                    "OS Atual": "{:d}",
                    "OS Simulada": "{:d}",
                    "Média Atual": "{:.1f}",
                    "Média Simulada": "{:.1f}",
                    "Desvio Atual": "{:+.1f}%",
                    "Desvio Simulado": "{:+.1f}%",
                },
                color_col="Média Simulada",
                color_meta=os_por_tecnico_ideal,
                height="auto",
            )

            # Gráfico comparativo
            fig_sim = go.Figure()

            fig_sim.add_trace(go.Bar(
                name="O.S. Atual",
                x=df_sim["Monitor"],
                y=df_sim["OS_Atual"],
                marker_color="#94A3B8",
                text=[f"{v}" for v in df_sim["OS_Atual"]],
                textposition="outside",
            ))
            fig_sim.add_trace(go.Bar(
                name="O.S. Simulada",
                x=df_sim["Monitor"],
                y=df_sim["OS Simulada"],
                marker_color=[
                    "#EF4444" if d is not None and float(d) > tolerancia
                    else "#F59E0B" if d is not None and float(d) < (tolerancia * -1)
                    else "#10B981"
                    for d in df_sim["Desvio Sim %"]
                ],
                text=[f"{v}" for v in df_sim["OS Simulada"]],
                textposition="outside",
            ))

            fig_sim.update_layout(
                title=dict(
                    text="<b>Distribuição de O.S. — Atual × Simulada</b>",
                    font=dict(family="Manrope", size=16, color="#012869"),
                ),
                barmode="group",
                xaxis_title="",
                yaxis_title="Quantidade de O.S.",
                height=420,
                margin=dict(t=60, b=40, l=40, r=60),
                plot_bgcolor="white",
                paper_bgcolor="white",
                legend=dict(
                    orientation="h",
                    yanchor="bottom", y=1.02,
                    xanchor="right", x=1,
                ),
            )
            fig_sim.update_xaxes(
                tickfont=dict(family="Inter", size=12, color="#374151"),
            )
            fig_sim.update_yaxes(
                gridcolor="#F1F5F9",
                tickfont=dict(family="Inter", size=11, color="#6B7280"),
            )

            st.plotly_chart(fig_sim, use_container_width=True)

            # Insight final
            if diff_total_os != 0:
                render_insight(
                    f"⚠️ Sua simulação tem **{abs(diff_total_os)} O.S. "
                    f"{'a mais' if diff_total_os > 0 else 'a menos'}** que o "
                    f"volume total atual ({total_sim_os} vs {int(total_os_g)}). "
                    f"Mantenha o mesmo total para uma **redistribuição real** "
                    f"entre monitores.",
                    tipo="alerta",
                )
            elif equilibrados_sim > equilibrados_atual:
                ganho = equilibrados_sim - equilibrados_atual
                render_insight(
                    f"✅ Excelente! Sua simulação equilibra "
                    f"**{ganho} monitor(es) a mais** mantendo o total de "
                    f"O.S. e a equipe fixa. Cenário viável para implementação.",
                    tipo="ok",
                )
            elif equilibrados_sim == equilibrados_atual:
                render_insight(
                    "ℹ️ Sua simulação mantém o mesmo número de monitores "
                    "equilibrados. Tente redistribuir de forma diferente.",
                    tipo="info",
                )
            else:
                render_insight(
                    f"❌ Cuidado: sua simulação **piora** o equilíbrio "
                    f"({equilibrados_sim} equilibrados vs "
                    f"{equilibrados_atual} no cenário atual).",
                    tipo="critico",
                )


if __name__ == "__main__":
    main()
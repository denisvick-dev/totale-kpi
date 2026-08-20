"""
quebra_geral.py
===============
Super Relatório Corporativo de Desempenho | Quebra Operacional
"""

from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Literal, Optional, Tuple

TipoInsight = Literal["ok", "info", "alerta", "critico", "acao"]

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from components.componentes import (
    aplicar_estilo,
    render_kpi as _render_kpi_global,
    render_insight as _render_insight_global,
    render_section_header,
    FONTE_TEXTO,
    FONTE_TITULO,
)

# ═══════════════════════════════════════════════════════
# ✅ IMPORT DO MÓDULO CENTRALIZADO DE CRITÉRIOS
# ═══════════════════════════════════════════════════════
from components.criterios import (
    VAZIOS_CONTRATO,
    detectar_col_status_atividade,
    detectar_col_contrato,
    classificar_tipo_servico,
)

# ═══════════════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA
# ═══════════════════════════════════════════════════════
st.set_page_config(
    page_title="Quebra Operacional | TOTALE",
    page_icon="📉",
    layout="wide",
    initial_sidebar_state="expanded",
)

aplicar_estilo()

if "df_memoria" not in st.session_state:
    st.session_state["df_memoria"] = None


# ═══════════════════════════════════════════════════════
# CONSTANTES DE DOMÍNIO
# ═══════════════════════════════════════════════════════
class Config:
    SLA_QUEBRA_MAXIMA = 0.20

    URL_LISTA_ATIVOS = (
        "https://docs.google.com/spreadsheets/d/"
        "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
    )
    SHEET_ID_ATIVOS = "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg"
    WORKSHEET_ATIVOS = "lista_ativos"

    CONTRATO_VALORES_VAZIOS = VAZIOS_CONTRATO
    STATUS_ORDEM = ["Executada", "Não Executada", "Pendente"]

    COL_REGIAO = "REGIÃO"

    CORES_TIPO = {
        "Novos Domicílios": "#1E40AF",
        "Migração": "#0284C7",
        "PME": "#1E3A8A",
        "Quebra Geral": "#78350F",
        "Outros": "#64748B",
    }
    ORDEM_TIPOS = ["Novos Domicílios", "Migração", "PME"]


CORES_REGIAO: Dict[str, Dict[str, str]] = {
    "LESTE": {"bg": "#DBEAFE", "text": "#1E40AF", "border": "#3B82F6"},
    "GRU": {"bg": "#D1FAE5", "text": "#065F46", "border": "#10B981"},
    "ABCDM": {"bg": "#EDE9FE", "text": "#5B21B6", "border": "#8B5CF6"},
    "OUTRAS": {"bg": "#F1F5F9", "text": "#475569", "border": "#94A3B8"},
}

_MAPA_TEMA_GLOBAL: Dict[str, str] = {
    "azul": "azul",
    "verde": "verde",
    "vermelho": "vermelho",
    "laranja": "laranja",
    "cinza": "cinza",
}


# ═══════════════════════════════════════════════════════
# WRAPPERS DE UI
# ═══════════════════════════════════════════════════════
def render_kpi(col, label: str, value: str, sub: str = "", tema: str = "azul") -> None:
    _render_kpi_global(col, label, value, sub, _MAPA_TEMA_GLOBAL.get(tema, "azul"))  # type: ignore


def render_insight(texto: str, tipo: TipoInsight = "info") -> None:
    _render_insight_global(texto, tipo)


def render_section(titulo: str) -> None:
    partes = titulo.strip().split(" ", 1)
    primeiro_char = partes[0][0] if partes[0] else ""
    if len(partes) == 2 and not primeiro_char.isascii():
        icon, title = partes[0], partes[1]
    else:
        icon, title = "📊", titulo
    render_section_header(icon, title)


def html_resultado_base(regioes: List[str], total: int) -> str:
    badges = "".join(
        [
            f'<span style="padding:0.3rem 0.9rem;border-radius:999px;'
            f"font-size:0.82rem;font-weight:700;border:2px solid;"
            f'background:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["bg"]};'
            f'color:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["text"]};'
            f'border-color:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["border"]};">'
            f"{r}</span>"
            for r in sorted(regioes)
        ]
    )
    total_fmt = f"{total:,}".replace(",", ".")
    return (
        '<div style="background:linear-gradient(135deg, #0F172A 0%, #1E3A5F 100%);'
        "padding:1rem 1.5rem;border-radius:0.75rem;margin-bottom:1.5rem;"
        "display:flex;align-items:center;flex-wrap:wrap;gap:0.6rem;"
        'box-shadow:0 4px 12px rgba(0,0,0,0.15);">'
        '<span style="color:#94A3B8;font-size:0.8rem;font-weight:700;'
        'text-transform:uppercase;letter-spacing:0.08em;">📋 Resultado da Base:</span>'
        f"{badges}"
        '<span style="color:#FFFFFF;font-size:0.78rem;margin-left:auto;'
        f'font-weight:700;">{total_fmt} registros</span>'
        "</div>"
    )


def render_hero_topo_fixo(
    titulo: str, subtitulo: str, regioes: List[str], total: int, badge: str = ""
) -> None:
    badge_html = ""
    if badge:
        badge_html = (
            f'<span style="display:inline-block;background:rgba(255,255,255,0.20);'
            f"padding:5px 16px;border-radius:20px;font-size:12px;font-weight:700;"
            f"margin-top:10px;letter-spacing:0.6px;text-transform:uppercase;"
            f'color:white;border:1px solid rgba(255,255,255,0.30);">'
            f"{badge}</span>"
        )
    resultado_html = html_resultado_base(regioes, total) if total > 0 else ""
    st.markdown(
        f'<div style="position:sticky;top:0.75rem;z-index:1000;'
        f"background:rgba(248,250,252,0.92);backdrop-filter:blur(10px);"
        f'-webkit-backdrop-filter:blur(10px);padding:0.5rem 0;border-radius:14px;">'
        f'<div style="background:linear-gradient(135deg, #012869 0%, #1E40AF 50%, #F37C04 100%);'
        f"padding:28px 40px;border-radius:14px;color:white;"
        f"box-shadow:0 10px 40px rgba(1,40,105,0.30);margin-bottom:12px;"
        f'position:relative;overflow:hidden;border:1px solid rgba(255,255,255,0.10);">'
        f'<div style="position:absolute;top:50%;right:-100px;transform:translateY(-50%);'
        f"width:420px;height:420px;background:radial-gradient(circle at center,"
        f"rgba(255,180,90,0.35) 0%, rgba(243,124,4,0.20) 35%,"
        f"rgba(232,93,4,0.08) 60%, transparent 78%);"
        f'border-radius:50%;pointer-events:none;filter:blur(2px);"></div>'
        f'<div style="position:relative;z-index:2;">'
        f'<h1 style="margin:0;font-size:30px;font-weight:800;color:white!important;'
        f'letter-spacing:-0.5px;text-shadow:0 2px 4px rgba(0,0,0,0.45);">{titulo}</h1>'
        f'<p style="margin:6px 0 0 0;font-size:14px;opacity:0.95;'
        f'color:#F8FAFC;text-shadow:0 1px 3px rgba(0,0,0,0.40);">{subtitulo}</p>'
        f"{badge_html}</div></div>"
        f"{resultado_html}</div>",
        unsafe_allow_html=True,
    )


def render_hero_upload() -> None:
    st.markdown(
        f'<div style="background:linear-gradient(135deg, #012869 0%, #1E40AF 50%, #F37C04 100%);'
        f"padding:32px 44px;border-radius:14px;color:white;"
        f"box-shadow:0 10px 40px rgba(1,40,105,0.30);margin-bottom:24px;"
        f'position:relative;overflow:hidden;border:1px solid rgba(255,255,255,0.10);">'
        f'<div style="position:absolute;top:50%;right:-100px;transform:translateY(-50%);'
        f"width:420px;height:420px;background:radial-gradient(circle at center,"
        f"rgba(255,180,90,0.35) 0%, rgba(243,124,4,0.20) 35%,"
        f"rgba(232,93,4,0.08) 60%, transparent 78%);"
        f'border-radius:50%;pointer-events:none;filter:blur(2px);"></div>'
        f'<div style="position:relative;z-index:2;">'
        f'<h1 style="margin:0;font-size:34px;font-weight:800;color:white!important;'
        f'letter-spacing:-0.8px;text-shadow:0 2px 4px rgba(0,0,0,0.45);">'
        f"📉 Gestão de Quebra de Agenda</h1>"
        f'<p style="margin:8px 0 0 0;font-size:15px;opacity:0.95;'
        f'color:#F8FAFC;text-shadow:0 1px 3px rgba(0,0,0,0.40);">'
        f"Importe a base para gerar o Super Relatório Consolidado</p>"
        f'<span style="display:inline-block;background:rgba(255,255,255,0.20);'
        f"padding:5px 16px;border-radius:20px;font-size:12px;font-weight:700;"
        f"margin-top:12px;letter-spacing:0.6px;text-transform:uppercase;"
        f'color:white;border:1px solid rgba(255,255,255,0.30);">SISTEMA TOTALE</span>'
        f"</div></div>",
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════════════════════
# UTILITÁRIOS GERAIS
# ═══════════════════════════════════════════════════════
class Utils:
    @staticmethod
    def buscar_coluna(df: pd.DataFrame, palavras: list) -> Optional[str]:
        if df is None or df.empty:
            return None
        cols = {
            str(c)
            .strip()
            .upper()
            .replace(".", "")
            .replace("_", "")
            .replace("  ", " "): c
            for c in df.columns
        }
        for p in palavras:
            pn = (
                str(p)
                .strip()
                .upper()
                .replace(".", "")
                .replace("_", "")
                .replace("  ", " ")
            )
            for cn, co in cols.items():
                if pn in cn:
                    return co
        return None

    @staticmethod
    def classificar_status(serie: pd.Series) -> pd.Series:
        s = serie.fillna("").astype(str).str.strip().str.upper()
        exe = s == "EXECUTADA"
        nex = s.isin(["NÃO EXECUTADA", "NAO EXECUTADA"])
        return pd.Series(
            np.select([exe, nex], ["Executada", "Não Executada"], default="Pendente"),
            index=serie.index,
        )

    @staticmethod
    def gerar_excel(df: pd.DataFrame, aba: str = "Dados") -> bytes:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name=aba[:31])
            ws = w.sheets[aba[:31]]
            hf = PatternFill("solid", fgColor="0F172A")
            for cell in ws[1]:
                cell.fill = hf
                cell.font = Font(color="FFFFFF", bold=True)
            for i, col in enumerate(df.columns, 1):
                try:
                    serie_str = df[col].fillna("").astype(str)
                    tamanhos = serie_str.str.len()
                    max_len_dados = int(tamanhos.max()) if len(tamanhos) > 0 else 0
                    max_len = max(max_len_dados, len(str(col)))
                    largura = min(max(max_len + 2, 12), 40)
                    ws.column_dimensions[get_column_letter(i)].width = largura
                except Exception:
                    ws.column_dimensions[get_column_letter(i)].width = 20
        return out.getvalue()


def _fmt_pct_br(v: Any) -> str:
    try:
        return (
            f"{float(v) * 100:,.2f}%".replace(",", "X")
            .replace(".", ",")
            .replace("X", ".")
        )
    except (ValueError, TypeError):
        return "0,00%"


def _fmt_int_br(v: Any) -> str:
    try:
        return f"{int(float(v)):,}".replace(",", ".")
    except (ValueError, TypeError):
        return "0"


# ═══════════════════════════════════════════════════════
# DATA LOADER
# ═══════════════════════════════════════════════════════
class DataLoader:
    @staticmethod
    @st.cache_data(show_spinner=False)
    def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
        bio = BytesIO(file_bytes)
        try:
            if filename.lower().endswith(".csv"):
                bio.seek(0)
                amostra = bio.read(5000).decode("utf-8", errors="ignore")
                bio.seek(0)
                try:
                    sep = csv.Sniffer().sniff(amostra).delimiter if amostra else ";"
                except Exception:
                    sep = ";"
                return pd.read_csv(
                    bio, sep=sep, encoding="utf-8", dtype=str, engine="python"
                )
            return pd.read_excel(bio, engine="openpyxl", dtype=str)
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(
        ttl=600, show_spinner="🔗 Conectando com Google Sheets (lista_ativos)..."
    )
    def buscar_gsheets() -> pd.DataFrame:
        try:
            from streamlit_gsheets import GSheetsConnection

            conn = st.connection("gsheets", type=GSheetsConnection)
            raw = conn.read(
                spreadsheet=Config.URL_LISTA_ATIVOS, worksheet=Config.WORKSHEET_ATIVOS
            )
            if raw is not None and not raw.empty:
                return DataLoader._processar_lista_ativos(raw)
        except Exception:
            pass
        try:
            url_csv = (
                f"https://docs.google.com/spreadsheets/d/"
                f"{Config.SHEET_ID_ATIVOS}/gviz/tq?tqx=out:csv&sheet={Config.WORKSHEET_ATIVOS}"
            )
            raw = pd.read_csv(url_csv)
            if raw is not None and not raw.empty:
                return DataLoader._processar_lista_ativos(raw)
        except Exception:
            pass
        try:
            url_csv = (
                f"https://docs.google.com/spreadsheets/d/"
                f"{Config.SHEET_ID_ATIVOS}/export?format=csv&gid=0"
            )
            raw = pd.read_csv(url_csv)
            if raw is not None and not raw.empty:
                return DataLoader._processar_lista_ativos(raw)
        except Exception as e:
            st.warning(f"⚠️ Não foi possível carregar lista_ativos: {e}")
        return pd.DataFrame()

    @staticmethod
    def _processar_lista_ativos(raw: pd.DataFrame) -> pd.DataFrame:
        if raw is None or raw.empty:
            return pd.DataFrame()
        raw.columns = raw.columns.astype(str).str.strip()
        rename_map = {}
        for col in raw.columns:
            col_upper = col.upper().strip()
            if col_upper in ("LOGIN", "MATRÍCULA", "MATRICULA", "ID"):
                rename_map[col] = "Login"
            elif col_upper in ("TÉCNICO", "TECNICO", "NOME", "NOME TÉCNICO"):
                rename_map[col] = "Técnico"
            elif col_upper in ("MONITOR", "GESTOR", "SUPERVISOR"):
                rename_map[col] = "Monitor"
            elif col_upper in ("BASE", "REGIÃO", "REGIAO"):
                rename_map[col] = "Base"
        raw = raw.rename(columns=rename_map)
        cols_uteis = [
            c for c in ["Login", "Técnico", "Monitor", "Base"] if c in raw.columns
        ]
        if "Login" not in cols_uteis:
            return pd.DataFrame()
        raw = raw[cols_uteis].copy()
        raw["Login"] = (
            raw["Login"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.upper()
        )
        raw = raw[raw["Login"].str.strip() != ""]
        raw = raw[~raw["Login"].isin(["NAN", "NONE", "NULL", "N/A"])]
        return raw.drop_duplicates(subset=["Login"], keep="last").reset_index(drop=True)

    @staticmethod
    @st.cache_data(show_spinner=False)
    def preparar_base(df: pd.DataFrame, df_gs: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()

        df = df.copy()
        df.columns = df.columns.astype(str).str.strip().str.upper()

        n_inicial = len(df)
        df.attrs["total_importado"] = n_inicial

        # ══════════════════════════════════════════════════════════
        # 🚫 1. REMOVER STATUS DA ATIVIDADE = "Suspenso"
        # ══════════════════════════════════════════════════════════
        col_atv = detectar_col_status_atividade(df)
        n_susp = 0
        if col_atv:
            serie_atv = df[col_atv].fillna("").astype(str).str.strip().str.upper()
            mask_susp = (
                serie_atv.str.contains("SUSP", na=False)
                | serie_atv.eq("SUSPENSO")
                | serie_atv.eq("SUSPENSA")
            )
            n_susp = int(mask_susp.sum())
            df = df[~mask_susp].copy()
        
        df.attrs["col_status_atividade"] = col_atv
        df.attrs["removidos_suspensos"] = n_susp

        # ══════════════════════════════════════════════════════════
        # 🚫 2. REMOVER CONTRATOS VAZIOS / INVÁLIDOS
        # ══════════════════════════════════════════════════════════
        col_con = detectar_col_contrato(df)
        n_invalidos = 0
        if col_con:
            serie_con = (
                df[col_con]
                .fillna("")
                .astype(str)
                .str.strip()
                .str.upper()
                .str.replace(r"\.0$", "", regex=True)
            )
            mask_invalido = serie_con.isin(VAZIOS_CONTRATO)
            n_invalidos = int(mask_invalido.sum())
            df = df[~mask_invalido].copy()

        df.attrs["col_contrato"] = col_con
        df.attrs["removidos_contrato"] = n_invalidos

        if df.empty:
            st.warning(
                "⚠️ Base ficou vazia após remoção de suspensos e contratos inválidos."
            )
            return pd.DataFrame()

        # ══════════════════════════════════════════════════════════
        # 3. TOTAL DE TAREFAS
        # ══════════════════════════════════════════════════════════
        col_tot = Utils.buscar_coluna(df, ["TOTAL DE TAREFAS", "QTD TAREFAS"])
        if col_tot:
            df["TOTAL DE TAREFAS"] = (
                pd.to_numeric(
                    df[col_tot].astype(str).str.replace(",", "."),
                    errors="coerce",
                )
                .fillna(1)
                .round()
                .astype("Int64")
            )
        else:
            df["TOTAL DE TAREFAS"] = pd.Series([1] * len(df), dtype="Int64")

        # ══════════════════════════════════════════════════════════
        # 4. MERGE COM LISTA_ATIVOS
        # ══════════════════════════════════════════════════════════
        col_login = Utils.buscar_coluna(
            df,
            ["LOGIN DO TÉCNICO", "LOGIN DO TECNICO", "LOGIN", "USUÁRIO", "MATRÍCULA"],
        )
        df.attrs["merge_aplicado"] = False
        df.attrs["merge_matches"] = 0
        df.attrs["merge_total"] = len(df)

        if col_login and not df_gs.empty and "Login" in df_gs.columns:
            df[col_login] = (
                df[col_login]
                .astype(str)
                .str.replace(r"\.0$", "", regex=True)
                .str.strip()
                .str.upper()
            )
            df = df.drop(
                columns=[c for c in ["TÉCNICO", "MONITOR", "Base"] if c in df.columns],
                errors="ignore",
            )
            df = df.merge(
                df_gs,
                left_on=col_login,
                right_on="Login",
                how="left",
                suffixes=("", "_gs"),
            )
            if "Login" in df.columns and col_login != "Login":
                df = df.drop(columns=["Login"], errors="ignore")
            if "Técnico" in df.columns:
                df.attrs["merge_matches"] = int(df["Técnico"].notna().sum())
                df.attrs["merge_aplicado"] = True

        if "Técnico" not in df.columns:
            col_tec_orig = Utils.buscar_coluna(
                df, ["TECNICO", "NOME TECNICO", "NOME DO TECNICO", "TÉCNICO"]
            )
            df["Técnico"] = (
                df[col_tec_orig]
                if col_tec_orig and col_tec_orig in df.columns
                else "NÃO MAPEADO"
            )

        if "Monitor" not in df.columns:
            col_mon_orig = Utils.buscar_coluna(
                df, ["MONITOR", "GESTOR", "SUPERVISOR", "NOME MONITOR"]
            )
            df["Monitor"] = (
                df[col_mon_orig]
                if col_mon_orig and col_mon_orig in df.columns
                else "SEM MONITOR"
            )

        df["TÉCNICO"] = (
            df["Técnico"].fillna("NÃO MAPEADO").astype(str).str.strip().str.upper()
        )
        df["MONITOR"] = (
            df["Monitor"].fillna("SEM MONITOR").astype(str).str.strip().str.upper()
        )
        df = df.drop(columns=["Técnico", "Monitor"], errors="ignore")
        df.loc[df["TÉCNICO"].isin(["", "NAN", "NONE", "NULL"]), "TÉCNICO"] = (
            "NÃO MAPEADO"
        )
        df.loc[df["MONITOR"].isin(["", "NAN", "NONE", "NULL"]), "MONITOR"] = (
            "SEM MONITOR"
        )

        # ══════════════════════════════════════════════════════════
        # 5. REGIÕES
        # ══════════════════════════════════════════════════════════
        col_cid = Utils.buscar_coluna(df, ["CIDADE", "LOCALIDADE"])
        cidade = (
            df[col_cid].fillna("").astype(str).str.strip().str.upper()
            if col_cid
            else pd.Series("", index=df.index)
        )
        df["REGIÃO"] = np.select(
            [
                cidade.isin(["SAO PAULO"]),
                cidade.isin(
                    [
                        "GUARULHOS",
                        "ARUJA",
                        "MOGI DAS CRUZES",
                        "SUZANO",
                        "ITAQUAQUECETUBA",
                        "FERRAZ DE VASCONCELOS",
                        "POA",
                    ]
                ),
                cidade.isin(
                    [
                        "SANTO ANDRE",
                        "SAO BERNARDO DO CAMPO",
                        "SAO CAETANO DO SUL",
                        "DIADEMA",
                        "MAUA",
                        "RIBEIRAO PIRES",
                        "RIO GRANDE DA SERRA",
                    ]
                ),
            ],
            ["LESTE", "GRU", "ABCDM"],
            default="OUTRAS",
        )

        # ══════════════════════════════════════════════════════════
        # 6. STATUS DO CONTRATO
        # ══════════════════════════════════════════════════════════
        col_status = Utils.buscar_coluna(
            df, ["STATUS DA O.S 1", "STATUS OS 1", "STATUS CONTRATO"]
        )
        df["Status Contrato"] = (
            Utils.classificar_status(df[col_status]) if col_status else "Pendente"
        )

        # ══════════════════════════════════════════════════════════
        # 7. ✅ CLASSIFICAÇÃO CENTRALIZADA
        # ══════════════════════════════════════════════════════════
        df, df["TIPO_SERVICO"] = classificar_tipo_servico(df)

        # ══════════════════════════════════════════════════════════
        # 8. MOTIVO DE BAIXA
        # ══════════════════════════════════════════════════════════
        col_cod = Utils.buscar_coluna(
            df, ["CÓD DE BAIXA 1", "COD DE BAIXA 1", "MOTIVO DE BAIXA"]
        )
        df["_COL_BAIXA"] = df[col_cod].astype(str) if col_cod else ""

        # ══════════════════════════════════════════════════════════
        # 9. DATA AGENDA
        # ══════════════════════════════════════════════════════════
        col_data = Utils.buscar_coluna(df, ["DATA", "DT AGENDA", "DATA AGENDA"])
        df["_DATA_AGENDA"] = (
            pd.to_datetime(df[col_data], errors="coerce", dayfirst=True)
            if col_data
            else pd.NaT
        )

        return df


# ═══════════════════════════════════════════════════════
# MOTOR ANALÍTICO
# ═══════════════════════════════════════════════════════
class Motor:
    @staticmethod
    def matriz_resumo(df: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()
        df_valid = df[df["TIPO_SERVICO"] != "Outros"].copy()
        if df_valid.empty:
            return pd.DataFrame()

        grp = (
            df_valid.groupby(["MONITOR", "TIPO_SERVICO"])
            .apply(
                lambda x: pd.Series(
                    {
                        "executados": x.loc[
                            x["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"
                        ].sum(),
                        "nao_executados": x.loc[
                            x["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"
                        ].sum(),
                        "total_tarefas": x["TOTAL DE TAREFAS"].sum(),
                    }
                )
            )
            .reset_index()
        )
        grp["denominador"] = grp["executados"] + grp["nao_executados"]
        grp["pct"] = np.where(
            grp["denominador"] > 0, grp["nao_executados"] / grp["denominador"], 0.0
        )

        pivot = grp.pivot_table(
            index="MONITOR", columns="TIPO_SERVICO", values="pct", fill_value=0.0
        )
        for t in Config.ORDEM_TIPOS:
            if t not in pivot.columns:
                pivot[t] = 0.0
        pivot = pivot[Config.ORDEM_TIPOS]

        exec_tot = (
            df_valid.loc[df_valid["Status Contrato"] == "Executada"]
            .groupby("MONITOR")["TOTAL DE TAREFAS"]
            .sum()
        )
        ne_tot = (
            df_valid.loc[df_valid["Status Contrato"] == "Não Executada"]
            .groupby("MONITOR")["TOTAL DE TAREFAS"]
            .sum()
        )
        tar_tot = df_valid.groupby("MONITOR")["TOTAL DE TAREFAS"].sum()
        df_tot = pd.DataFrame({"exec": exec_tot, "ne": ne_tot, "tar": tar_tot}).fillna(
            0
        )

        pivot["Quebra Geral"] = np.where(
            (df_tot["exec"] + df_tot["ne"]) > 0,
            df_tot["ne"] / (df_tot["exec"] + df_tot["ne"]),
            0.0,
        )
        pivot["Total Tarefas"] = df_tot["tar"].astype(int)
        pivot = pivot.reset_index().rename(columns={"MONITOR": "Monitor"})

        total_row: Dict[str, Any] = {"Monitor": "Total Geral"}
        for tipo in Config.ORDEM_TIPOS:
            sub = df_valid[df_valid["TIPO_SERVICO"] == tipo]
            ex = sub.loc[
                sub["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"
            ].sum()
            ne = sub.loc[
                sub["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"
            ].sum()
            total_row[tipo] = ne / (ex + ne) if (ex + ne) > 0 else 0.0

        ex_g = df_valid.loc[
            df_valid["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"
        ].sum()
        ne_g = df_valid.loc[
            df_valid["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"
        ].sum()
        total_row["Quebra Geral"] = ne_g / (ex_g + ne_g) if (ex_g + ne_g) > 0 else 0.0
        total_row["Total Tarefas"] = int(df_valid["TOTAL DE TAREFAS"].sum())

        return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)


# ═══════════════════════════════════════════════════════
# ESTILIZAÇÃO MATRIZ
# ═══════════════════════════════════════════════════════
def estilizar_matriz(df: pd.DataFrame, meta: float):
    cols_pct = [c for c in df.columns if c not in ("Monitor", "Total Tarefas")]

    def _cores(row):
        estilos = []
        is_total = str(row.get("Monitor", "")).upper() == "TOTAL GERAL"
        for col in df.columns:
            if col == "Monitor":
                estilos.append(
                    "background:linear-gradient(90deg,#012869 0%,#1E40AF 100%);color:white;font-weight:800;text-align:left;padding-left:16px;"
                    if is_total
                    else "background-color:#F8FAFC;font-weight:700;text-align:left;padding-left:16px;border-right:2px solid #E2E8F0;"
                )
            elif col == "Total Tarefas":
                estilos.append(
                    "background:#1E3A8A;color:white;font-weight:800;text-align:center;"
                    if is_total
                    else "background-color:#EFF6FF;color:#1E3A8A;font-weight:700;text-align:center;"
                )
            else:
                try:
                    val = float(row[col])
                except (ValueError, TypeError):
                    val = 0.0
                bg = "#FEE2E2" if val > meta else "#D1FAE5"
                tc = "#991B1B" if val > meta else "#065F46"
                if is_total:
                    bg = "#7F1D1D" if val > meta else "#064E3B"
                    tc = "white"
                estilos.append(
                    f"background-color:{bg};color:{tc};text-align:center;font-weight:800;"
                )
        return estilos

    styler = df.style.apply(_cores, axis=1)
    fmt: Dict[str, Any] = {c: _fmt_pct_br for c in cols_pct}
    if "Total Tarefas" in df.columns:
        fmt["Total Tarefas"] = _fmt_int_br
    styler = styler.format(fmt)
    return styler.set_table_styles(
        [
            {
                "selector": "th",
                "props": [
                    ("background", "#012869"),
                    ("color", "white"),
                    ("text-align", "center"),
                    ("padding", "10px"),
                    ("font-family", FONTE_TITULO),
                    ("font-weight", "700"),
                    ("text-transform", "uppercase"),
                    ("letter-spacing", "0.05em"),
                ],
            },
            {
                "selector": "td",
                "props": [
                    ("padding", "12px 10px"),
                    ("border-bottom", "1px solid #E2E8F0"),
                    ("font-variant-numeric", "tabular-nums"),
                ],
            },
        ]
    )


def render_visao_resumo(df: pd.DataFrame, meta_pct: float) -> None:
    if df.empty:
        render_insight("Sem dados para a Visão Resumo.", tipo="alerta")
        return

    with st.spinner("Gerando matriz corporativa..."):
        df_matriz = Motor.matriz_resumo(df)

    if df_matriz.empty:
        render_insight("Não foi possível gerar a matriz de resumo.", tipo="alerta")
        return

    total_row = df_matriz[df_matriz["Monitor"] == "Total Geral"].iloc[0]
    total_tar = int(total_row["Total Tarefas"])
    q_geral = float(total_row["Quebra Geral"])

    k1, k2, k3, k4 = st.columns(4)
    render_kpi(
        k1,
        "Total O.S.",
        f"{total_tar:,}".replace(",", "."),
        "Base válida analisada",
        "azul",
    )
    render_kpi(
        k2,
        "Quebra Consolidada",
        f"{q_geral:.2%}",
        "Todos os segmentos",
        "vermelho" if q_geral > meta_pct else "verde",
    )
    render_kpi(k3, "Meta Geral", f"{meta_pct:.0%}", "SLA Alvo", "cinza")

    pior_tipo = max(Config.ORDEM_TIPOS, key=lambda t: float(total_row.get(t, 0)))
    render_kpi(
        k4,
        "Segmento Crítico",
        pior_tipo,
        f"Quebra: {float(total_row[pior_tipo]):.2%}",
        "laranja",
    )

    st.markdown("<br>", unsafe_allow_html=True)
    render_section("📋 Matriz de Desempenho (Monitor × Segmento)")
    st.markdown(
        '<div style="background:#F1F5F9;padding:12px;border-radius:6px;'
        'font-size:13px;color:#334155;margin-bottom:16px;">'
        "🧮 <b>Fórmula:</b> Não Executadas ÷ (Executadas + Não Executadas). "
        "Pendentes não entram no cálculo.</div>",
        unsafe_allow_html=True,
    )
    styler = estilizar_matriz(df_matriz, meta_pct)
    st.markdown(
        f'<div style="background:white;padding:5px;border-radius:12px;'
        f'box-shadow:0 4px 12px rgba(0,0,0,0.08);">'
        f'{styler.hide(axis="index").to_html()}</div>',
        unsafe_allow_html=True,
    )

    st.markdown("<br>", unsafe_allow_html=True)
    c_dw1, _, _ = st.columns([1.2, 1, 3])
    with c_dw1:
        st.download_button(
            "📊 Baixar Matriz (Excel)",
            Utils.gerar_excel(df_matriz, "Matriz"),
            f"matriz_quebra_{datetime.now():%Y%m%d_%H%M}.xlsx",
            use_container_width=True,
        )

    st.markdown("<br>", unsafe_allow_html=True)
    render_section("📊 Distribuição Visual")
    df_plot = df_matriz[df_matriz["Monitor"] != "Total Geral"].copy()
    fig = go.Figure()
    for tipo in Config.ORDEM_TIPOS:
        fig.add_trace(
            go.Bar(
                name=tipo,
                x=df_plot["Monitor"],
                y=df_plot[tipo],
                marker_color=Config.CORES_TIPO.get(tipo, "#64748B"),
                text=[_fmt_pct_br(v) for v in df_plot[tipo]],
                textposition="outside",
            )
        )
    fig.add_hline(
        y=meta_pct,
        line_dash="dash",
        line_color="#DC2626",
        annotation_text=f"META: {meta_pct:.0%}",
    )
    fig.update_layout(
        barmode="group",
        height=500,
        yaxis_tickformat=".0%",
        legend=dict(orientation="h", yanchor="bottom", y=1.02),
    )
    st.plotly_chart(fig, use_container_width=True, config={"displayModeBar": False})


# ====================================================
# BARRA LATERAL (SIDEBAR)
# ====================================================
def render_sidebar(df_full: pd.DataFrame) -> Dict[str, Any]:
    with st.sidebar:
        st.markdown("### 🎯 Filtros Globais")

        monitores = ["Todos"] + sorted(
            str(x) for x in df_full["MONITOR"].dropna().unique()
            if str(x) not in {"nan", "SEM MONITOR", "NÃO MAPEADO"}
        )
        sel_mon = st.selectbox("👔 Monitor", monitores)
        df_filt = df_full if sel_mon == "Todos" else df_full[df_full["MONITOR"] == sel_mon]

        tecnicos = ["Todos"] + sorted(
            str(x) for x in df_filt["TÉCNICO"].dropna().unique()
            if str(x) not in {"nan", "NÃO MAPEADO"}
        )
        sel_tec = st.selectbox("👤 Técnico", tecnicos)
        df = df_filt if sel_tec == "Todos" else df_filt[df_filt["TÉCNICO"] == sel_tec]

        st.caption(f"📊 **{len(df):,}** registros após filtros".replace(",", "."))

        st.divider()
        meta = (
            st.number_input(
                "🎯 Meta Geral SLA (%)",
                0.0, 100.0,
                float(Config.SLA_QUEBRA_MAXIMA * 100),
                1.0,
            )
            / 100.0
        )

        st.divider()
        col_r1, col_r2 = st.columns(2)
        with col_r1:
            if st.button("🔄 Reiniciar", use_container_width=True):
                st.session_state["df_memoria"] = None
                st.rerun()
        with col_r2:
            if st.button("🗑️ Limpar Cache", use_container_width=True):
                st.cache_data.clear()
                st.session_state["df_memoria"] = None
                st.rerun()

    return {
        "df": df,
        "meta": meta,
    }


def garantir_colunas_criticas(df: pd.DataFrame) -> pd.DataFrame:
    """Garante colunas essenciais para a análise de quebra geral."""
    if df is None:
        return pd.DataFrame()

    df = df.copy()

    if "TOTAL DE TAREFAS" not in df.columns:
        df["TOTAL DE TAREFAS"] = 1
    if "TIPO_SERVICO" not in df.columns:
        df["TIPO_SERVICO"] = "Outros"
    if "REGIÃO" not in df.columns:
        df["REGIÃO"] = "OUTRAS"
    if "Status Contrato" not in df.columns:
        df["Status Contrato"] = "Pendente"
    if "MONITOR" not in df.columns:
        df["MONITOR"] = "SEM MONITOR"
    if "TÉCNICO" not in df.columns:
        df["TÉCNICO"] = "NÃO MAPEADO"

    for col in ["TOTAL DE TAREFAS", "MONITOR", "TÉCNICO", "REGIÃO", "TIPO_SERVICO"]:
        if col in df.columns:
            df[col] = df[col].fillna("" if col not in {"TOTAL DE TAREFAS"} else 0)

    df["TOTAL DE TAREFAS"] = pd.to_numeric(
        df["TOTAL DE TAREFAS"], errors="coerce"
    ).fillna(1).astype(int)

    return df


# ═══════════════════════════════════════════════════════
# PONTO DE ENTRADA PRINCIPAL
# ═══════════════════════════════════════════════════════
def main() -> None:
    if st.session_state["df_memoria"] is None:
        render_hero_upload()

        render_section("📁 Importação de Dados")
        arq = st.file_uploader("Selecione a base (Excel/CSV)", type=["xlsx", "csv"])

        if arq:
            with st.spinner("🔄 Limpando dados e classificando segmentos..."):
                raw = DataLoader.ler_arquivo(arq.getvalue(), arq.name)
                gs = DataLoader.buscar_gsheets()
                df_proc = DataLoader.preparar_base(raw, gs)
                df_proc = garantir_colunas_criticas(df_proc)
                st.session_state["df_memoria"] = df_proc

            # ── Relatório de Limpeza ───────────────────────────────────
            n_susp = df_proc.attrs.get("removidos_suspensos", 0)
            n_con = df_proc.attrs.get("removidos_contrato", 0)
            col_atv = df_proc.attrs.get("col_status_atividade", None)
            col_con = df_proc.attrs.get("col_contrato", None)
            total = len(raw)
            restou = len(df_proc)

            render_section("🧹 Relatório de Limpeza da Base")

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("📥 Total Importado", f"{total:,}".replace(",", "."))
            c2.metric(
                "🚫 Suspensos Removidos",
                f"{n_susp:,}".replace(",", "."),
                delta=f"-{n_susp}" if n_susp else None,
                delta_color="inverse",
            )
            c3.metric(
                "📄 Contratos Inválidos",
                f"{n_con:,}".replace(",", "."),
                delta=f"-{n_con}" if n_con else None,
                delta_color="inverse",
            )
            c4.metric("✅ Base Final", f"{restou:,}".replace(",", "."))

            if not col_atv:
                st.warning(
                    "⚠️ STATUS DA ATIVIDADE não detectada — suspensos não removidos."
                )
            else:
                st.success(f"✅ `{col_atv}` → **{n_susp}** suspensos removidos")

            if not col_con:
                st.warning("⚠️ CONTRATO não detectada — inválidos não removidos.")
            else:
                st.success(
                    f"✅ `{col_con}` → **{n_con}** contratos inválidos removidos"
                )

            if df_proc.attrs.get("merge_aplicado"):
                matches = df_proc.attrs.get("merge_matches", 0)
                total_m = df_proc.attrs.get("merge_total", len(df_proc))
                st.toast(
                    f"✅ Merge: {matches:,}/{total_m:,}".replace(",", "."), icon="🔗"
                )
            else:
                st.toast("⚠️ lista_ativos não carregada", icon="⚠️")

            st.rerun()
        return

    df_full = st.session_state["df_memoria"].copy()
    df_full = garantir_colunas_criticas(df_full)
    st.session_state["df_memoria"] = df_full

    config_user = render_sidebar(df_full)
    df = config_user["df"]
    meta = config_user["meta"]

    regioes_disp = (
        sorted(df[Config.COL_REGIAO].unique())
        if Config.COL_REGIAO in df.columns
        else ["OUTRAS"]
    )

    # ── Renderiza o hero fixo e a matriz executiva ──────────────────
    render_hero_topo_fixo(
        titulo="📉 Super Relatório de Quebra — Resumo Executivo",
        subtitulo="Matriz Monitor × Segmento · Novos Domicílios · Migração · PME",
        regioes=list(regioes_disp),
        total=len(df),
        badge="VISÃO CONSOLIDADA",
    )

    if df.empty:
        render_insight(
            "🔍 <b>Nenhum dado para os filtros selecionados.</b><br>"
            "Ajuste os filtros na barra lateral ou clique em <b>🔄 Reiniciar Painel</b>.",
            tipo="alerta",
        )
        return

    render_visao_resumo(df, meta)


if __name__ == "__main__":
    main()
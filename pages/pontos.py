import streamlit as st
import streamlit.components.v1 as components
import pandas as pd
import plotly.express as px
from plotly.graph_objects import Figure
import numpy as np
import datetime
import calendar
from io import BytesIO
from typing import Any, Optional
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from componentes import (
    aplicar_estilo,
    render_hero,
    render_kpi,
    render_insight,
    render_section_header,
    FONTE_TEXTO,
    COR_TEXTO_3,
)

# ⚠️ set_page_config DEVE ser a primeira chamada Streamlit
st.set_page_config(
    page_title="Central de Performance",
    page_icon="📈",
    layout="wide",
)

# Aplica CSS/tema/fontes ANTES de qualquer outro st.*
aplicar_estilo()


# ====================================================
# BLOCO 2: COMPONENTES VISUAIS
# ====================================================
class ComponenteVisual:
    @staticmethod
    def exibir_ticker(dados: list) -> None:
        """Renderiza ticker animado via components.html (isolado em iframe)."""
        if not dados:
            return

        html_itens = ""
        for item in dados:
            variacao = item.get("variacao", "")
            if variacao == "positiva":
                cor, simbolo = "#22c55e", "▲"
            elif variacao == "negativa":
                cor, simbolo = "#ef4444", "▼"
            elif variacao == "share":
                cor, simbolo = "#38bdf8", "◴"
            else:
                cor, simbolo = "#94a3b8", "■"

            html_itens += (
                f'<span class="ticker-item">'
                f'<span class="ticker-label">{item.get("label", "")}:</span>'
                f'<span class="ticker-valor">{item.get("valor", "")}</span>'
                f'<span class="ticker-delta" style="color:{cor};">'
                f'{simbolo} {item.get("delta", "")}</span>'
                f'</span>'
                f'<span class="ticker-sep">|</span>'
            )

        html = f"""
<!DOCTYPE html>
<html>
<head>
<link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;700&display=swap" rel="stylesheet">
<style>
  * {{ box-sizing: border-box; margin: 0; padding: 0; }}
  body {{
    background: transparent;
    font-family: 'Inter', -apple-system, sans-serif;
    overflow: hidden;
  }}
  .ticker-wrapper {{
    width: 100%;
    overflow: hidden;
    background: linear-gradient(90deg,#0f172a 0%,#1e293b 50%,#0f172a 100%);
    padding: 12px 0;
    border-radius: 8px;
    box-shadow: 0 4px 12px rgba(0,0,0,0.3);
    position: relative;
  }}
  .ticker-wrapper::before, .ticker-wrapper::after {{
    content: ''; position: absolute; top: 0; bottom: 0;
    width: 60px; z-index: 2; pointer-events: none;
  }}
  .ticker-wrapper::before {{
    left: 0; background: linear-gradient(90deg,#0f172a,transparent);
  }}
  .ticker-wrapper::after {{
    right: 0; background: linear-gradient(90deg,transparent,#0f172a);
  }}
  .ticker-content {{
    display: flex; width: max-content;
    animation: scroll 35s linear infinite;
  }}
  .ticker-wrapper:hover .ticker-content {{ animation-play-state: paused; }}
  @keyframes scroll {{
    0%   {{ transform: translate3d(0,0,0); }}
    100% {{ transform: translate3d(-50%,0,0); }}
  }}
  .ticker-item   {{ margin: 0 40px; font-size: 15px; white-space: nowrap; }}
  .ticker-label  {{ color: #94a3b8; font-weight: 500; }}
  .ticker-valor  {{ font-weight: 700; color: #FFFFFF; margin-left: 8px; }}
  .ticker-delta  {{ font-weight: 700; margin-left: 6px; font-size: 13px; }}
  .ticker-sep    {{ color: #334155; margin: 0 15px; }}
</style>
</head>
<body>
  <div class="ticker-wrapper">
    <div class="ticker-content">{html_itens}{html_itens}</div>
  </div>
</body>
</html>
"""
        components.html(html, height=60, scrolling=False)

    @staticmethod
    def gerar_podio(ranking_df: pd.DataFrame) -> None:
        if len(ranking_df) < 3:
            render_insight("Necessário ao menos 3 equipes para o pódio.", tipo="alerta")
            return

        top3 = ranking_df.head(3).reset_index(drop=True)
        c2, c1, c3 = st.columns([1, 1.2, 1])

        def _medalha_html(nome: str, pontos: float, fundo: str, borda: str, icone: str) -> str:
            return (
                f'<div style="background-color:{fundo};border:2px solid {borda};'
                f'border-radius:10px;padding:15px;text-align:center;'
                f'box-shadow:0 4px 8px rgba(0,0,0,0.1);">'
                f'<h1 style="margin:0;font-size:30px;">{icone}</h1>'
                f'<h4 style="margin:5px 0;color:#334155;">{nome}</h4>'
                f'<h3 style="margin:0;color:{borda};">{pontos:,.1f} pts</h3>'
                f'</div>'
            )

        medalhas = [
            (c1, 0, "#FEF9C3", "#EAB308", "🥇 1º Lugar"),
            (c2, 1, "#F1F5F9", "#94A3B8", "🥈 2º Lugar"),
            (c3, 2, "#FFEDD5", "#F97316", "🥉 3º Lugar"),
        ]
        for col, idx, fundo, borda, icone in medalhas:
            with col:
                st.markdown(
                    _medalha_html(
                        top3.iloc[idx]["Nome Equipe"],
                        top3.iloc[idx]["Pontos"],
                        fundo, borda, icone,
                    ),
                    unsafe_allow_html=True,
                )

    @staticmethod
    def gerar_insight_ia(media: float, dias_brutos: int) -> None:
        dias_txt = (
            f"Restam {dias_brutos} dia(s) útil(is) para o fechamento. "
            if dias_brutos > 0
            else "Último dia do período. "
        )
        if media >= 400:
            render_insight(
                f"**Visão da IA:** Operação em Alta Performance! "
                f"A média de {media:.1f} pts/equipe ultrapassa o teto máximo. "
                f"{dias_txt}Mantenha o ritmo para fechar o mês com recorde.",
                tipo="ok",
            )
        elif media >= 300:
            render_insight(
                f"**Visão da IA:** Operação Estável. "
                f"Média de {media:.1f} pts garante o atingimento da meta base. "
                f"{dias_txt}Foco em puxar os retardatários para garantir a projeção.",
                tipo="info",
            )
        else:
            render_insight(
                f"**Visão da IA:** Alerta de Performance! "
                f"A média de {media:.1f} pts está abaixo da linha de corte (300). "
                f"{dias_txt}Ação imediata necessária para reverter a tendência.",
                tipo="alerta",
            )

    @staticmethod
    def render_dataframe_corporativo(
        df: pd.DataFrame,
        titulo: str = "",
        icone: str = "📊",
        badge: str = "",
        modo_diario: bool = False,
    ) -> None:
        if df.empty:
            render_insight("Nenhum dado disponível.", tipo="info")
            return

        badge_txt = badge or f"{len(df)} equipes"
        modo_txt = "📅 Meta Diária" if modo_diario else "📆 Acumulado do Mês"

        st.markdown(
            f"""
            <div style="background:#FFFFFF;border-radius:12px 12px 0 0;
                 padding:14px 20px;border:1px solid #E2E8F0;border-bottom:none;
                 display:flex;align-items:center;gap:12px;">
                <span style="font-size:1.4rem;">{icone}</span>
                <div style="flex:1;">
                    <div style="font-size:0.95rem;font-weight:800;color:#0F172A;">{titulo}</div>
                    <div style="font-size:0.72rem;color:#64748B;
                         text-transform:uppercase;letter-spacing:0.05em;
                         font-weight:600;margin-top:2px;">{modo_txt}</div>
                </div>
                <span style="background:#EFF6FF;color:#1D4ED8;
                     padding:5px 12px;border-radius:999px;
                     font-size:0.75rem;font-weight:700;">{badge_txt}</span>
            </div>
            """,
            unsafe_allow_html=True,
        )

        df_display = df.copy()
        renomear = {
            "Posição":      "Rank",
            "CódAuxEquipe": "Cód. Equipe",
            "Nome Equipe":  "Equipe",
            "Supervisor":   "Supervisor",
            "Projeto":      "Projeto",
            "Pontos":       "Pontos",
            "Projeção":     "Proj. Fechamento",
        }
        df_display = df_display.rename(
            columns={k: v for k, v in renomear.items() if k in df_display.columns}
        )

        if "Rank" in df_display.columns:
            df_display["Rank"] = df_display["Rank"].apply(Utilitarios.formatar_posicao)

        cols_meta = [c for c in df_display.columns if "Meta" in c]
        formatos: dict[str, Any] = {c: "{:,.1f}" for c in cols_meta}
        if "Pontos" in df_display.columns:
            formatos["Pontos"] = "{:,.1f}"
        if "Proj. Fechamento" in df_display.columns:
            formatos["Proj. Fechamento"] = "{:,.1f}"

        styler = df_display.style.format(formatter=formatos)  # type: ignore[arg-type]

        if "Pontos" in df_display.columns:
            styler = styler.map(Utilitarios.colorir_metas, subset=["Pontos"])  # type: ignore[arg-type]
        if "Proj. Fechamento" in df_display.columns:
            styler = styler.map(Utilitarios.colorir_projecao, subset=["Proj. Fechamento"])  # type: ignore[arg-type]

        styler = styler.set_table_styles([
            {"selector": "thead th", "props": [
                ("background", "linear-gradient(180deg,#012869 0%,#1E3A8A 100%)"),
                ("color", "#FFFFFF"), ("font-weight", "700"),
                ("font-size", "0.72rem"), ("text-align", "center"),
                ("padding", "12px 10px"), ("letter-spacing", "0.04em"),
                ("text-transform", "uppercase"), ("border", "none"),
                ("border-right", "1px solid rgba(255,255,255,0.12)"),
            ]},
            {"selector": "tbody td", "props": [
                ("font-size", "0.82rem"), ("padding", "10px 12px"),
                ("border-bottom", "1px solid #F1F5F9"),
                ("font-variant-numeric", "tabular-nums"),
            ]},
            {"selector": "tbody tr:nth-child(even) td", "props": [("background-color", "#FAFBFC")]},
            {"selector": "tbody tr:hover td", "props": [("background-color", "#EFF6FF !important")]},
            {"selector": "table", "props": [
                ("border-collapse", "separate"), ("border-spacing", "0"),
                ("width", "100%"), ("font-family", FONTE_TEXTO),
            ]},
        ])

        st.markdown(
            '<div style="background:#FFFFFF;border-radius:0 0 12px 12px;'
            'border:1px solid #E2E8F0;border-top:none;overflow:hidden;'
            'box-shadow:0 4px 12px rgba(0,0,0,0.05);margin-bottom:16px;">',
            unsafe_allow_html=True,
        )
        st.dataframe(styler, use_container_width=True, hide_index=True, height=450)
        st.markdown("</div>", unsafe_allow_html=True)


# ====================================================
# BLOCO 3: UTILITÁRIOS
# ====================================================
class Utilitarios:
    @staticmethod
    def encontrar_coluna_data(df: pd.DataFrame) -> Optional[str]:
        for c in ["Data Agendamento", "Data Conclusão", "Data", "Date", "Data_Execucao"]:
            if c in df.columns:
                return c
        return None

    @staticmethod
    def calcular_variacao(vf: float, vg: float) -> tuple[str, str]:
        if vg == 0 or pd.isna(vg):
            return "neutra", "S/D"
        if abs(vf - vg) < 0.0001:
            return "neutra", "Visão Geral"
        p = ((vf - vg) / vg) * 100
        if p > 0:
            return "positiva", f"+{p:.1f}%"
        if p < 0:
            return "negativa", f"{p:.1f}%"
        return "neutra", "0%"

    @staticmethod
    def calcular_share(vf: float, vg: float) -> tuple[str, str]:
        if vg == 0 or pd.isna(vg):
            return "neutra", "0%"
        if abs(vf - vg) < 0.0001:
            return "neutra", "Visão Geral"
        return "share", f"{(vf / vg) * 100:.1f}% do Total"

    @staticmethod
    def formatar_numero(v: float) -> str:
        return f"{v:,.0f}".replace(",", ".")

    @staticmethod
    def formatar_posicao(valor: Any) -> str:
        try:
            v = int(valor)
            return {1: f"🥇 {v}º", 2: f"🥈 {v}º", 3: f"🥉 {v}º"}.get(v, f"{v}º")
        except (ValueError, TypeError):
            return str(valor)

    @staticmethod
    def colorir_metas(valor: Any) -> str:
        try:
            v = float(valor)
        except (ValueError, TypeError):
            return ""
        if v >= 400:
            return ("background-color:#1E3A8A;color:#FFFFFF;font-weight:800;"
                    "border-left:3px solid #0F172A;text-align:center;")
        if v >= 300:
            return ("background-color:#DCFCE7;color:#166534;font-weight:700;"
                    "border-left:3px solid #22C55E;text-align:center;")
        if v >= 275:
            return ("background-color:#FEF9C3;color:#854D0E;font-weight:700;"
                    "border-left:3px solid #EAB308;text-align:center;")
        return "font-weight:700;border-left:3px solid #EF4444;text-align:center;"

    @staticmethod
    def colorir_projecao(valor: Any) -> str:
        return ("background-color:#0F172A;color:#FFFFFF;font-weight:800;"
                "text-align:center;border-left:3px solid #64748B;")

    @staticmethod
    def calcular_dias_uteis(df: pd.DataFrame) -> tuple[int, int, Any, int]:
        col_data = Utilitarios.encontrar_coluna_data(df)
        data_max: datetime.date = (
            pd.to_datetime(df[col_data].max()).date()
            if col_data and pd.notna(df[col_data].max())
            else datetime.date.today()
        )
        ano, mes = data_max.year, data_max.month
        primeiro = datetime.date(ano, mes, 1)
        _, ult = calendar.monthrange(ano, mes)
        ultimo = datetime.date(ano, mes, ult)

        p_np = np.datetime64(primeiro)
        m_np = np.datetime64(data_max)
        u_np = np.datetime64(ultimo)

        total = int(np.busday_count(p_np, u_np + np.timedelta64(1, "D"), weekmask="1111110"))
        passados = int(np.busday_count(p_np, m_np + np.timedelta64(1, "D"), weekmask="1111110"))
        brutos = max(0, total - passados)
        seguros = max(1, brutos)
        return brutos, seguros, data_max, passados

    @staticmethod
    def exportar_excel(dataframe: pd.DataFrame) -> bytes:
        output = BytesIO()
        with pd.ExcelWriter(output, engine="openpyxl") as writer:
            dataframe.to_excel(writer, index=False, sheet_name="Ranking")
            ws = writer.sheets["Ranking"]

            cor_cab      = PatternFill("solid", fgColor="012869")
            cor_par      = PatternFill("solid", fgColor="F8FAFC")
            cor_impar    = PatternFill("solid", fgColor="FFFFFF")
            cor_proj     = PatternFill("solid", fgColor="303030")
            cor_alta     = PatternFill("solid", fgColor="1F497D")
            cor_ok       = PatternFill("solid", fgColor="C6EFCE")
            cor_proximo  = PatternFill("solid", fgColor="FFEB9C")

            f_cab     = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            f_branca  = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
            f_preta   = Font(name="Calibri", size=11, bold=True, color="000000")
            f_verde   = Font(name="Calibri", size=11, bold=True, color="006100")
            f_amarela = Font(name="Calibri", size=11, bold=True, color="9C5700")

            borda = Border(
                left=Side(style="thin", color="D9D9D9"),
                right=Side(style="thin", color="D9D9D9"),
                top=Side(style="thin", color="D9D9D9"),
                bottom=Side(style="thin", color="D9D9D9"),
            )
            centro = Alignment(horizontal="center", vertical="center")

            cols = list(dataframe.columns)
            col_int = [i + 1 for i, c in enumerate(cols) if c.lower() in ("posição", "posicao")]
            col_dec = [i + 1 for i, c in enumerate(cols)
                       if c.lower() not in ("posição", "posicao")
                       and pd.api.types.is_numeric_dtype(dataframe[c])]
            idx_proj   = cols.index("Projeção") + 1 if "Projeção" in cols else -1
            idx_pontos = cols.index("Pontos") + 1 if "Pontos" in cols else -1

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row,
                                    min_col=1, max_col=ws.max_column):
                for cel in row:
                    cel.border = borda
                    if cel.row == 1:
                        cel.fill = cor_cab
                        cel.font = f_cab
                        cel.alignment = centro
                        continue

                    cel.fill = cor_par if cel.row % 2 == 0 else cor_impar

                    if cel.column == idx_proj:
                        cel.fill = cor_proj
                        cel.font = f_branca
                    elif cel.column == idx_pontos:
                        try:
                            v = float(cel.value or 0)
                            if v >= 400:
                                cel.fill = cor_alta; cel.font = f_branca
                            elif v >= 300:
                                cel.fill = cor_ok; cel.font = f_verde
                            elif v >= 275:
                                cel.fill = cor_proximo; cel.font = f_amarela
                            else:
                                cel.font = f_preta
                        except (ValueError, TypeError):
                            pass

                    if cel.column in col_int:
                        cel.number_format = "#,##0"
                    elif cel.column in col_dec:
                        cel.number_format = "#,##0.0"

            for col in ws.columns:
                letra = get_column_letter(col[0].column)
                max_len = max((len(str(c.value)) for c in col if c.value), default=0)
                ws.column_dimensions[letra].width = max(max_len + 3, 12)

            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions

        return output.getvalue()


# ====================================================
# BLOCO 4: PROCESSAMENTO
# ====================================================
class ProcessamentoDados:
    @staticmethod
    def calcular_rankings(
        df: pd.DataFrame, dias_brutos: int, dias_seguros: int, dias_passados: int
    ) -> tuple[pd.DataFrame, pd.DataFrame]:
        base = (
            df.groupby(["CódAuxEquipe", "Nome Equipe", "Supervisor", "Projeto"])["Pontos"]
            .sum().reset_index().sort_values("Pontos", ascending=False)
        )

        if "Dias Trab Tecnico" in df.columns:
            dias_trab = (
                base["Nome Equipe"]
                .map(df.groupby("Nome Equipe")["Dias Trab Tecnico"].max())
                .fillna(dias_passados).astype(int)
            )
        else:
            dias_trab = pd.Series(dias_passados, index=base.index)

        dias_trab = dias_trab.replace(0, 1)
        media_pts = base["Pontos"] / dias_trab
        projecao = base["Pontos"] + (media_pts * dias_brutos)

        def _montar(modo_dia: bool) -> pd.DataFrame:
            r = base.copy()
            r.insert(0, "Posição", range(1, len(r) + 1))
            for m in [300, 350, 375, 400]:
                label = f"Meta Dia | {m}" if modo_dia else f"Meta | {m}"
                r[label] = (r["Pontos"] - m) / dias_seguros if modo_dia else r["Pontos"] - m
            r["Projeção"] = projecao.values
            return r

        return _montar(False), _montar(True)

    @staticmethod
    def calcular_saude_operacao(ranking: pd.DataFrame) -> pd.DataFrame:
        cond = [
            ranking["Pontos"] >= 400,
            (ranking["Pontos"] >= 300) & (ranking["Pontos"] < 400),
            ranking["Pontos"] < 300,
        ]
        esc = ["Alta (400+)", "Na Meta (300-399)", "Abaixo (< 300)"]
        r = ranking.copy()
        r["Status"] = np.select(cond, esc, default="Sem Dados")
        return r["Status"].value_counts().reset_index()

    @staticmethod
    def ranking_supervisores(ranking: pd.DataFrame) -> pd.DataFrame:
        return (
            ranking.groupby("Supervisor")
            .agg(Qtd_Equipes=("Nome Equipe", "count"), Total_Pontos=("Pontos", "sum"))
            .assign(Media_por_Equipe=lambda x: x["Total_Pontos"] / x["Qtd_Equipes"])
            .reset_index().sort_values("Media_por_Equipe", ascending=True)
        )


# ====================================================
# BLOCO 5: GRÁFICOS
# ====================================================
class Graficos:
    @staticmethod
    def _layout(fig: Figure) -> Figure:
        fig.update_layout(
            showlegend=False,
            plot_bgcolor="rgba(0,0,0,0)",
            paper_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0, r=0, t=30, b=0),
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.1)"),
        )
        return fig

    @staticmethod
    def barras_horizontal(df: pd.DataFrame, x: str, y: str) -> Figure:
        fig = px.bar(df, x=x, y=y, orientation="h", text_auto=True,
                     color=x, color_continuous_scale="Tealgrn")
        fig.update_traces(texttemplate="%{text:.1f}", textposition="outside", textfont_size=12)
        return Graficos._layout(fig)

    @staticmethod
    def rosca(df: pd.DataFrame, names: str, values: str) -> Figure:
        fig = px.pie(df, names=names, values=values, hole=0.6, color=names,
                     color_discrete_map={
                         "Alta (400+)": "#1E3A8A",
                         "Na Meta (300-399)": "#22C55E",
                         "Abaixo (< 300)": "#EF4444",
                     })
        fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            margin=dict(l=0, r=0, t=0, b=0),
            showlegend=True,
            legend=dict(orientation="h", yanchor="bottom", y=-0.2, xanchor="center", x=0.5),
        )
        return fig

    @staticmethod
    def linhas(df: pd.DataFrame, x: str, y: str, color: str) -> Figure:
        fig = px.line(df, x=x, y=y, color=color, markers=True)
        fig.update_layout(
            paper_bgcolor="rgba(0,0,0,0)",
            plot_bgcolor="rgba(0,0,0,0)",
            xaxis=dict(showgrid=False),
            yaxis=dict(showgrid=True, gridcolor="rgba(0,0,0,0.1)"),
            margin=dict(l=0, r=0, t=30, b=0),
        )
        return fig


# ====================================================
# BLOCO 6: HERO + DADOS
# ====================================================
render_hero(
    titulo="📈 Central de Performance | Produção por Técnico",
    subtitulo="Acompanhamento individual de execução, produtividade e performance em campo",
    badge="Atualização em tempo real",
)

if "dados_prod" not in st.session_state:
    render_insight("Carregue os dados na página principal primeiro.", tipo="alerta")
    st.stop()

prod = st.session_state["dados_prod"]["Prod"].copy()
gpon = st.session_state["dados_prod"]["Gpon"].copy()
prod["Pontos"] = pd.to_numeric(prod["Pontos"], errors="coerce").fillna(0)
gpon["Pontos"] = pd.to_numeric(gpon["Pontos"], errors="coerce").fillna(0)
df = pd.concat([prod, gpon], ignore_index=True)

TOTAL_GERAL_PONTOS  = df["Pontos"].sum()
TOTAL_GERAL_OS      = len(df)
TOTAL_GERAL_EQUIPES = df["Nome Equipe"].nunique() if "Nome Equipe" in df.columns else 0
MEDIA_GERAL_PONTOS  = TOTAL_GERAL_PONTOS / TOTAL_GERAL_EQUIPES if TOTAL_GERAL_EQUIPES > 0 else 0.0


# ====================================================
# BLOCO 7: FILTROS
# ====================================================
st.sidebar.header("🎯 Filtros Avançados")

for col_nome, label in [
    ("Projeto", "Projeto:"),
    ("Supervisor", "Supervisor:"),
    ("Nome Equipe", "Equipe:"),
]:
    if col_nome in df.columns:
        opcoes = ["Todos"] + sorted(df[col_nome].dropna().astype(str).unique())
        sel = st.sidebar.selectbox(label, opcoes, key=f"sel_{col_nome}")
        if sel != "Todos":
            df = df[df[col_nome] == sel]

if st.sidebar.button("🔄 Limpar Filtros"):
    st.rerun()


# ====================================================
# BLOCO 8: CÁLCULOS + KPIs
# ====================================================
dias_brutos, dias_seguros, ultima_atualizacao, dias_passados = (
    Utilitarios.calcular_dias_uteis(df)
)

total_equipes_filtro = df["Nome Equipe"].nunique() if "Nome Equipe" in df.columns else 0
total_os_filtro      = len(df)
total_pontos_filtro  = df["Pontos"].sum() if "Pontos" in df.columns else 0.0
media_pontos_filtro  = (
    total_pontos_filtro / total_equipes_filtro if total_equipes_filtro > 0 else 0.0
)

var_pontos  = Utilitarios.calcular_share(total_pontos_filtro, TOTAL_GERAL_PONTOS)
var_os      = Utilitarios.calcular_share(total_os_filtro, TOTAL_GERAL_OS)
var_equipes = Utilitarios.calcular_share(total_equipes_filtro, TOTAL_GERAL_EQUIPES)
var_media   = Utilitarios.calcular_variacao(media_pontos_filtro, MEDIA_GERAL_PONTOS)

ComponenteVisual.exibir_ticker([
    {"label": "Pontos",       "valor": Utilitarios.formatar_numero(total_pontos_filtro),  "variacao": var_pontos[0],  "delta": var_pontos[1]},
    {"label": "O.S.",         "valor": Utilitarios.formatar_numero(total_os_filtro),      "variacao": var_os[0],      "delta": var_os[1]},
    {"label": "Equipes",      "valor": str(total_equipes_filtro),                         "variacao": var_equipes[0], "delta": var_equipes[1]},
    {"label": "Média/Equipe", "valor": Utilitarios.formatar_numero(media_pontos_filtro),  "variacao": var_media[0],   "delta": var_media[1]},
])

ComponenteVisual.gerar_insight_ia(media_pontos_filtro, dias_brutos)

c1, c2, c3, c4 = st.columns(4)
render_kpi(c1, "Total Pontos · Geral",  Utilitarios.formatar_numero(TOTAL_GERAL_PONTOS),   tema="cinza")
render_kpi(c2, "Total Pontos · Filtro", Utilitarios.formatar_numero(total_pontos_filtro),  sub=var_pontos[1],  tema="azul")
render_kpi(c3, "Equipes · Filtro",      str(total_equipes_filtro),                         sub=var_equipes[1], tema="verde")
render_kpi(c4, "Total O.S. · Filtro",   Utilitarios.formatar_numero(total_os_filtro),      sub=var_os[1],      tema="laranja")

st.divider()


# ====================================================
# BLOCO 9: ABAS
# ====================================================
ranking = pd.DataFrame()
ranking_dia = pd.DataFrame()
if "Nome Equipe" in df.columns:
    ranking, ranking_dia = ProcessamentoDados.calcular_rankings(
        df, dias_brutos, dias_seguros, dias_passados
    )

aba_ranking, aba_executivo, aba_evolucao = st.tabs(
    ["🏆 Ranking & Metas", "👔 Visão Executiva", "📈 Evolução Temporal"]
)

# ── ABA 1 ──────────────────────────────────────────
with aba_ranking:
    render_section_header("🥇", "Pódio Corporativo — Top 3 Equipes")
    ComponenteVisual.gerar_podio(ranking)

    st.write("")

    col_t, col_tg = st.columns([3, 1])
    with col_t:
        render_section_header("📊", "Ranking Geral · Metas & Projeção")
    with col_tg:
        st.write("")
        por_dia = st.toggle("📅 **Modo Meta Diária**", key="tg_meta_dia")

    df_exibir = ranking_dia if por_dia else ranking
    modo_txt  = "Meta Diária" if por_dia else "Meta Mensal"

    ComponenteVisual.render_dataframe_corporativo(
        df_exibir,
        titulo=f"Performance por Equipe — {modo_txt}",
        icone="🏆",
        badge=f"{len(df_exibir)} equipes ativas",
        modo_diario=por_dia,
    )

    col_dl1, col_dl2, _ = st.columns([1.5, 1.5, 5])
    nome_arq = "ranking_diario.xlsx" if por_dia else "ranking_geral.xlsx"

    with col_dl1:
        st.download_button(
            "📥 **Exportar Excel**",
            data=Utilitarios.exportar_excel(df_exibir),
            file_name=nome_arq,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )
    with col_dl2:
        st.download_button(
            "📄 **Exportar CSV**",
            data=df_exibir.to_csv(index=False, decimal=",").encode("utf-8-sig"),
            file_name=nome_arq.replace(".xlsx", ".csv"),
            mime="text/csv",
            use_container_width=True,
        )

    st.markdown(
        f"""
        <div style="display:flex;gap:12px;flex-wrap:wrap;margin-top:16px;
             padding:12px 16px;background:#F8FAFC;border-radius:8px;
             border:1px solid #E2E8F0;font-size:0.78rem;
             font-family:{FONTE_TEXTO};">
            <span style="font-weight:700;color:{COR_TEXTO_3};
                 text-transform:uppercase;letter-spacing:0.05em;">🎨 Legenda:</span>
            <span style="background:#1E3A8A;color:white;padding:3px 10px;
                 border-radius:6px;font-weight:700;">🏆 400+ pts — Alta Performance</span>
            <span style="background:#DCFCE7;color:#166534;padding:3px 10px;
                 border-radius:6px;font-weight:700;">✅ 300-399 — Na Meta</span>
            <span style="background:#FEF9C3;color:#854D0E;padding:3px 10px;
                 border-radius:6px;font-weight:700;">⚠️ 275-299 — Próximo da Meta</span>
        </div>
        """,
        unsafe_allow_html=True,
    )

# ── ABA 2 ──────────────────────────────────────────
with aba_executivo:
    render_section_header("👔", "Performance de Gestão")

    c_e1, c_e2 = st.columns([1, 2])
    with c_e1:
        st.markdown("**Saúde da Operação (Faixa de Metas)**")
        df_saude = ProcessamentoDados.calcular_saude_operacao(ranking)
        st.plotly_chart(
            Graficos.rosca(df_saude, "Status", "count"),
            use_container_width=True, key="graf_saude",
        )
    with c_e2:
        st.markdown("**Ranking de Supervisores (Média pts por Equipe)**")
        df_sup = ProcessamentoDados.ranking_supervisores(ranking)
        st.plotly_chart(
            Graficos.barras_horizontal(df_sup, "Media_por_Equipe", "Supervisor"),
            use_container_width=True, key="graf_sup",
        )

# ── ABA 3 ──────────────────────────────────────────
with aba_evolucao:
    col_data = Utilitarios.encontrar_coluna_data(df)
    render_section_header("📈", "Curva de Tendência Diária")

    if col_data and not ranking.empty:
        top5 = ranking.head(5)["Nome Equipe"].tolist()
        df_ev = df[df["Nome Equipe"].isin(top5)].copy()
        df_ev[col_data] = pd.to_datetime(df_ev[col_data]).dt.date

        df_ag = (
            df_ev.groupby([col_data, "Nome Equipe"])["Pontos"].sum().reset_index()
        )
        df_ag["Pontos Acumulados"] = df_ag.groupby("Nome Equipe")["Pontos"].cumsum()

        st.plotly_chart(
            Graficos.linhas(df_ag, col_data, "Pontos Acumulados", "Nome Equipe"),
            use_container_width=True, key="graf_linha",
        )
        st.caption("Evolução de pontos acumulados das Top 5 equipes atuais.")
    else:
        render_insight(
            "Para ver a evolução, a planilha precisa de uma coluna de Data "
            "(ex: 'Data Agendamento' ou 'Data Conclusão').",
            tipo="info",
        )

# ── Rodapé ──────────────────────────────────────────
if pd.notna(ultima_atualizacao):
    st.sidebar.divider()
    st.sidebar.caption(
        f"🕒 **Última Atualização:**\n"
        f"{pd.to_datetime(ultima_atualizacao).strftime('%d/%m/%Y')}"
    )
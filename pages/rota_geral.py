# rota_geral.py

from __future__ import annotations

import unicodedata
from io import BytesIO, StringIO
from typing import (
    TYPE_CHECKING,
    Any,
    Optional,
    TypedDict,
    Union,
    cast,
    Literal,
)

import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.utils import get_column_letter

from components.componentes import (
    aplicar_estilo,
    render_hero,
    render_section_header,
    render_kpi,
    render_kpi_sm,
    render_insight,
    COR_PRIMARIA,
    COR_SECUNDARIA,
    COR_LARANJA_SUAVE,
)

if TYPE_CHECKING:
    from streamlit.runtime.uploaded_file_manager import UploadedFile


# ============================================================
# TIPOS
# ============================================================

class LinhaResultado(TypedDict, total=False):
    """Estrutura de uma linha do DataFrame consolidado."""
    BASE: str
    WO: int
    OS: int
    ND: int
    RC: int
    MESH: int
    MIGRAÇÃO: int
    GPON: int
    PME: int
    Rotas: int
    Montados: int
    Media_OS: float
    Media_Montados: float


TipoInsight = str  # Literal["ok", "info", "alerta", "critico", "acao"]
TipoTema = str     # Literal["azul", "verde", "vermelho", "laranja", "cinza"]


# ============================================================
# CONFIGURAÇÕES
# ============================================================

class Config:
    """Configurações globais da aplicação."""

    PAGE_TITLE: str = "Rota Geral | Totale"
    PAGE_ICON: str = "📊"
    LAYOUT: Literal["centered", "wide"] = "wide"

    BASES: list[str] = ["ABCDM", "GUARULHOS", "LESTE"]

    COLUNAS_INDICADORES: list[str] = [
        "WO", "OS", "ND", "RC", "MESH",
        "MIGRAÇÃO", "GPON", "PME",
    ]

    COLUNAS_METRICAS: list[str] = [
        "Rotas", "Média OS", "Montados", "Média Montados",
    ]

    ORDEM_COLUNAS: list[str] = ["BASE"] + COLUNAS_INDICADORES + COLUNAS_METRICAS

    BASES_CONFIG: dict[str, dict[str, str]] = {
        "ABCDM": {"cor": COR_PRIMARIA},
        "GUARULHOS": {"cor": COR_SECUNDARIA},
        "LESTE": {"cor": COR_LARANJA_SUAVE},
    }

    COLUNAS_ESPERADAS: list[str] = [
        "Contrato", "Total de tarefas", "Tipo O.S 1",
        "Tipo de Atividade.1", "Habilidade de Trabalho",
        "Login do Técnico",
    ]


# ============================================================
# UTILITÁRIOS
# ============================================================

class Utils:
    """Funções utilitárias para manipulação de dados."""

    @staticmethod
    def normalizar_texto(texto: object) -> str:
        if texto is None:
            return ""
        try:
            if pd.isna(texto):  # type: ignore[arg-type]
                return ""
        except (TypeError, ValueError):
            pass

        txt = str(texto).strip().upper()
        txt = unicodedata.normalize("NFKD", txt)
        return "".join(c for c in txt if not unicodedata.combining(c))

    @staticmethod
    def preparar_colunas(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        df.columns = pd.Index([str(c).strip() for c in df.columns])
        return df

    @staticmethod
    def localizar_coluna(df: pd.DataFrame, nome: str) -> Optional[str]:
        procurado = Utils.normalizar_texto(nome)
        for coluna in df.columns:
            if Utils.normalizar_texto(coluna) == procurado:
                return str(coluna)
        return None

    @staticmethod
    def localizar_colunas_tipo_os(df: pd.DataFrame) -> list[str]:
        encontradas: list[str] = []
        for i in range(1, 11):
            for nome in (
                f"Tipo O.S {i}", f"Tipo O.S.{i}",
                f"Tipo OS {i}", f"Tipo O.S{i}",
            ):
                col = Utils.localizar_coluna(df, nome)
                if col is not None:
                    encontradas.append(col)
                    break
        return encontradas

    @staticmethod
    def to_float(v: object, default: float = 0.0) -> float:
        if v is None:
            return default
        if isinstance(v, bool):
            return float(v)
        if isinstance(v, (int, float)):
            return float(v)
        if isinstance(v, str):
            try:
                return float(v.replace(",", "."))
            except ValueError:
                return default
        try:
            if hasattr(v, "__len__") and not isinstance(v, (bytes, str)):
                return default
            return float(v)  # type: ignore[arg-type]
        except (TypeError, ValueError):
            return default

    @staticmethod
    def to_int(v: object, default: int = 0) -> int:
        return int(Utils.to_float(v, float(default)))

    @staticmethod
    def fmt_int(v: object) -> str:
        return f"{Utils.to_int(v):,}".replace(",", ".")

    @staticmethod
    def fmt_float(v: object, casas: int = 2) -> str:
        return f"{Utils.to_float(v):.{casas}f}".replace(".", ",")


# ============================================================
# PROCESSAMENTO
# ============================================================

class DataProcessor:
    """Processamento de dados das bases."""

    # ---------- Flag GPON ----------

    @staticmethod
    def criar_flag_gpon(df: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        coluna = Utils.localizar_coluna(df, "Habilidade de Trabalho")

        if coluna is None:
            df["GPON_FLAG"] = 0
            return df

        serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
        df["GPON_FLAG"] = serie.str.contains(
            "PON", regex=False, na=False,
        ).astype(int)
        return df

    # ---------- Leitura ----------

    @staticmethod
    def carregar_arquivo(arquivo: Optional["UploadedFile"]) -> Optional[pd.DataFrame]:
        if arquivo is None:
            return None

        nome = arquivo.name.lower()
        try:
            if nome.endswith(".csv"):
                df = DataProcessor._ler_csv(arquivo)
            elif nome.endswith(".xlsx"):
                df = DataProcessor._ler_xlsx(arquivo)
            elif nome.endswith(".xls"):
                arquivo.seek(0)
                df = pd.read_excel(arquivo, engine="xlrd")
            else:
                raise ValueError("Formato não suportado.")

            df = Utils.preparar_colunas(df)
            df = DataProcessor.criar_flag_gpon(df)
            return df
        except Exception as erro:
            st.error(f"Erro ao carregar {arquivo.name}: {erro}")
            return None

    @staticmethod
    def _ler_csv(arquivo: "UploadedFile") -> pd.DataFrame:
        tentativas: list[dict[str, str]] = [
            {"sep": ";", "encoding": "utf-8"},
            {"sep": ";", "encoding": "latin-1"},
            {"sep": ",", "encoding": "utf-8"},
            {"sep": ",", "encoding": "latin-1"},
        ]
        for params in tentativas:
            try:
                arquivo.seek(0)
                df = pd.read_csv(
                    arquivo,
                    sep=params["sep"],
                    encoding=params["encoding"],
                    low_memory=False,
                )
                if len(df.columns) > 1:
                    return df
            except Exception:
                continue
        raise ValueError("Não foi possível identificar o formato do CSV.")

    @staticmethod
    def _ler_xlsx(arquivo: "UploadedFile") -> pd.DataFrame:
        arquivo.seek(0)
        conteudo_bytes: bytes = arquivo.read()
        conteudo = BytesIO(conteudo_bytes)

        try:
            return pd.read_excel(conteudo, engine="openpyxl")
        except Exception:
            try:
                texto = conteudo_bytes.decode("utf-8")
            except UnicodeDecodeError:
                texto = conteudo_bytes.decode("latin-1")
            return pd.read_html(StringIO(texto))[0]

    # ---------- Cálculos base ----------

    @staticmethod
    def _aplicar_mascara(
        df: pd.DataFrame,
        mascara: Optional[pd.Series],
    ) -> pd.DataFrame:
        if mascara is None:
            return df
        return df.loc[mascara]

    @staticmethod
    def _contratos_unicos(
        df: pd.DataFrame,
        mascara: Optional[pd.Series] = None,
    ) -> int:
        col = Utils.localizar_coluna(df, "Contrato")
        if col is None:
            return 0
        dados = DataProcessor._aplicar_mascara(df, mascara)
        s = dados[col].dropna().astype(str).str.strip()
        s = s[~s.isin(["", "nan", "None", "0"])]
        return int(s.nunique())

    @staticmethod
    def _total_tarefas(
        df: pd.DataFrame,
        mascara: Optional[pd.Series] = None,
    ) -> int:
        col = Utils.localizar_coluna(df, "Total de tarefas")
        if col is None:
            return 0
        dados = DataProcessor._aplicar_mascara(df, mascara)
        valores = pd.to_numeric(dados[col], errors="coerce").fillna(0)
        return int(valores.sum())

    @staticmethod
    def _coluna_contem(
        df: pd.DataFrame, nome_coluna: str, texto: str,
    ) -> pd.Series:
        coluna = Utils.localizar_coluna(df, nome_coluna)
        if coluna is None:
            return pd.Series(False, index=df.index)
        procurado = Utils.normalizar_texto(texto)
        serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
        return serie.str.contains(procurado, regex=False, na=False)

    @staticmethod
    def _coluna_igual(
        df: pd.DataFrame, nome_coluna: str, texto: str,
    ) -> pd.Series:
        coluna = Utils.localizar_coluna(df, nome_coluna)
        if coluna is None:
            return pd.Series(False, index=df.index)
        procurado = Utils.normalizar_texto(texto)
        serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
        return serie.eq(procurado)

    @staticmethod
    def _tipo_os_contem(
        df: pd.DataFrame, texto: str,
    ) -> pd.Series:
        procurado = Utils.normalizar_texto(texto)
        mascara: pd.Series = pd.Series(False, index=df.index)
        for coluna in Utils.localizar_colunas_tipo_os(df):
            serie = df[coluna].fillna("").astype(str).map(Utils.normalizar_texto)
            mascara = mascara | serie.str.contains(
                procurado, regex=False, na=False,
            )
        return mascara

    # ---------- Cálculos consolidados ----------

    @staticmethod
    def calcular_indicadores(df: pd.DataFrame) -> dict[str, int]:
        mask_adesao = DataProcessor._coluna_contem(
            df, "Tipo O.S 1", "ADESAO"
        )
        mask_pme = DataProcessor._coluna_contem(
            df, "Habilidade de Trabalho", "PME"
        )
        mask_gpon: pd.Series = df["GPON_FLAG"].eq(1)
        mask_pacote = DataProcessor._tipo_os_contem(df, "PACOTE")
        mask_rc = DataProcessor._coluna_igual(
            df, "Tipo de Atividade.1", "Retorno Credenciada"
        )
        mask_mesh = DataProcessor._tipo_os_contem(df, "MESH")

        return {
            "WO": DataProcessor._contratos_unicos(df),
            "OS": DataProcessor._total_tarefas(df),
            "ND": DataProcessor._contratos_unicos(df, mask_adesao),
            "RC": DataProcessor._contratos_unicos(df, mask_rc),
            "MESH": DataProcessor._total_tarefas(df, mask_mesh),
            "MIGRAÇÃO": DataProcessor._contratos_unicos(
                df, mask_pacote & mask_gpon,
            ),
            "GPON": DataProcessor._contratos_unicos(df, mask_gpon),
            "PME": DataProcessor._contratos_unicos(
                df, mask_adesao & mask_pme,
            ),
        }

    @staticmethod
    def calcular_rotas(df: pd.DataFrame) -> int:
        coluna = Utils.localizar_coluna(df, "Login do Técnico")
        if coluna is None:
            return 0
        logins = df[coluna].dropna().astype(str).str.strip()
        logins = logins[~logins.isin(["", "nan", "None"])]
        return int(logins.nunique())

    @staticmethod
    def processar_base(
        nome: str, df: pd.DataFrame, montados: int,
    ) -> dict[str, Union[str, int, float]]:
        ind = DataProcessor.calcular_indicadores(df)
        rotas = DataProcessor.calcular_rotas(df)
        os_qtd = int(ind["OS"])

        linha: dict[str, Union[str, int, float]] = {"BASE": nome}
        for chave, valor in ind.items():
            linha[chave] = valor
        linha["Rotas"] = rotas
        linha["Média OS"] = round(os_qtd / rotas, 2) if rotas > 0 else 0.0
        linha["Montados"] = int(montados)
        linha["Média Montados"] = round(os_qtd / montados, 2) if montados > 0 else 0.0
        return linha

    @staticmethod
    def adicionar_linha_total(df: pd.DataFrame) -> pd.DataFrame:
        soma_os = float(df["OS"].sum())
        soma_rotas = float(df["Rotas"].sum())
        soma_montados = float(df["Montados"].sum())

        total: dict[str, Union[str, int, float]] = {"BASE": "Total"}
        for col in Config.COLUNAS_INDICADORES + ["Rotas", "Montados"]:
            total[col] = int(df[col].sum())

        total["Média OS"] = round(soma_os / soma_rotas, 2) if soma_rotas > 0 else 0.0
        total["Média Montados"] = round(soma_os / soma_montados, 2) if soma_montados > 0 else 0.0

        return pd.concat([df, pd.DataFrame([total])], ignore_index=True)


# ============================================================
# VISUALIZAÇÃO
# ============================================================

class Visualization:
    """Geração de tabelas, gráficos, insights e exportações."""

    COLUNAS_DESTAQUE: dict[str, str] = {
        "WO":             "cel-destaque-wo",
        "OS":             "cel-destaque-os",
        "MIGRAÇÃO":       "cel-destaque-mig",
        "PME":            "cel-destaque-pme",
        "Rotas":          "cel-destaque-rotas",
        "Média OS":       "cel-destaque-mediaos",
        "Montados":       "cel-destaque-montados",
        "Média Montados": "cel-destaque-mediamont",
    }

    # ---------- Excel ----------

    @staticmethod
    def gerar_excel(df: pd.DataFrame) -> bytes:
        buffer = BytesIO()
        with pd.ExcelWriter(buffer, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Rota Geral", index=False)
            ws = writer.book["Rota Geral"]

            for cell in ws[1]:
                cell.font = cell.font.copy(bold=True, color="FFFFFF")
                cell.fill = cell.fill.copy(fill_type="solid", fgColor="012869")

            for idx, coluna in enumerate(ws.columns, start=1):
                letra = get_column_letter(idx)
                maior = 0
                for celula in coluna:
                    if celula.__class__.__name__ == "MergedCell":
                        continue
                    v = str(celula.value) if celula.value is not None else ""
                    if len(v) > maior:
                        maior = len(v)
                ws.column_dimensions[letra].width = maior + 3

        buffer.seek(0)
        return buffer.getvalue()

    # ---------- Tabela HTML ----------

    @staticmethod
    def _classe_celula(
        valor: float,
        coluna: str,
        df_bases: Optional[pd.DataFrame] = None,
    ) -> str:
        """
        Mapa de calor por RANKING (posição entre as bases):
        - 1º lugar (maior valor) → 🟢 Verde
        - 2º lugar (intermediário) → 🟡 Amarelo
        - 3º lugar (menor valor) → 🔴 Vermelho
        
        Para colunas negativas (ND, RC, MESH, PME), a lógica é invertida:
        menor valor = melhor (verde).
        """
        # Destaques especiais (mantém identidade Totale)
        if coluna == "Média OS":
            return "cel-media-os"
        if coluna == "Média Montados":
            return "cel-media-mont"
        
        if df_bases is None or coluna not in df_bases.columns:
            return "cel-cinza-corp"
        
        valores = pd.to_numeric(df_bases[coluna], errors="coerce").dropna()
        if len(valores) == 0:
            return "cel-cinza-corp"
        
        # Colunas onde MENOR valor é MELHOR
        COLS_NEGATIVAS = {"RC",}
        ascending = coluna in COLS_NEGATIVAS
        
        # Ranking: 1 = melhor, 2 = médio, 3 = pior
        ranking = valores.rank(method="min", ascending=ascending)
        
        # Descobre posição do valor atual
        try:
            idx_atual = valores[valores == valor].index[0]
            posicao = int(ranking.loc[idx_atual])
        except (IndexError, KeyError):
            return "cel-cinza-corp"
        
        # Classifica pela posição do ranking
        if posicao == 1:
            return "cel-heat-verde"
        elif posicao == 2:
            return "cel-heat-amarelo"
        else:
            return "cel-heat-vermelho"

    @staticmethod
    def renderizar_tabela_html(df: pd.DataFrame) -> None:
        thead = "".join(f"<th>{c}</th>" for c in df.columns)
        linhas_html: list[str] = []
        total_idx = len(df) - 1
        
        # Referência apenas com as bases (exclui linha Total)
        df_bases_ref = df.iloc[:-1].copy()

        for i, (_, row) in enumerate(df.iterrows()):
            eh_total = (i == total_idx)
            classe_tr = ' class="linha-total"' if eh_total else ""

            celulas: list[str] = []
            for c in df.columns:
                col_nome = str(c)
                val = row[c]

                if col_nome == "BASE":
                    celulas.append(f'<td class="col-base">{val}</td>')
                    continue

                v = Utils.to_float(val)
                if col_nome in ("Média OS", "Média Montados"):
                    texto = Utils.fmt_float(v)
                else:
                    texto = Utils.fmt_int(v)

                if eh_total:
                    celulas.append(f"<td>{texto}</td>")
                else:
                    classe = Visualization._classe_celula(v, col_nome, df_bases_ref)
                    celulas.append(f'<td class="{classe}">{texto}</td>')

            linhas_html.append(f"<tr{classe_tr}>{''.join(celulas)}</tr>")

        tabela = f"""
        <div class="tabela-rota-wrapper">
            <table class="tabela-rota">
                <thead><tr>{thead}</tr></thead>
                <tbody>{"".join(linhas_html)}</tbody>
            </table>
        </div>
        """
        st.markdown(tabela, unsafe_allow_html=True)
    
    
    # ---------- Gráficos ----------

    @staticmethod
    def grafico_comparativo_medias(df_bases: pd.DataFrame) -> go.Figure:
        fig = go.Figure()
        fig.add_trace(
            go.Bar(
                name="Média OS",
                x=df_bases["BASE"],
                y=df_bases["Média OS"],
                marker_color=COR_PRIMARIA,
                text=[Utils.fmt_float(v) for v in df_bases["Média OS"]],
                textposition="outside",
            )
        )
        fig.add_trace(
            go.Bar(
                name="Média Montados",
                x=df_bases["BASE"],
                y=df_bases["Média Montados"],
                marker_color=COR_SECUNDARIA,
                text=[Utils.fmt_float(v) for v in df_bases["Média Montados"]],
                textposition="outside",
            )
        )
        fig.update_layout(
            title="Média de OS por Rota × por Montado",
            barmode="group",
            height=380,
            margin=dict(t=60, b=40, l=20, r=20),
        )
        return fig

    @staticmethod
    def grafico_participacao(df_bases: pd.DataFrame) -> go.Figure:
        color_map: dict[str, str] = {
            k: v["cor"] for k, v in Config.BASES_CONFIG.items()
        }
        fig = px.pie(
            df_bases,
            values="OS",
            names="BASE",
            title="Participação de cada base no total de OS",
            hole=0.6,
            color="BASE",
            color_discrete_map=color_map,
        )
        fig.update_traces(
            textinfo="percent+label",
            textfont_size=13,
            marker=dict(line=dict(color="white", width=3)),
        )
        fig.update_layout(
            height=380,
            showlegend=False,
            margin=dict(t=60, b=20, l=20, r=20),
        )
        return fig

    @staticmethod
    def grafico_heatmap(df_bases: pd.DataFrame) -> go.Figure:
        matriz = df_bases.set_index("BASE")[Config.COLUNAS_INDICADORES]
        matriz_max = matriz.max().replace(0, 1)
        matriz_norm = matriz.div(matriz_max, axis=1)

        escala_totale: list[list[Union[float, str]]] = [
            [0.00, "#DBEAFE"], [0.20, "#BFDBFE"], [0.40, "#E0E7FF"],
            [0.55, "#FFF7ED"], [0.70, "#FED7AA"], [0.85, "#FDBA74"],
            [1.00, "#F37C04"],
        ]

        fig = go.Figure(
            data=go.Heatmap(
                z=matriz_norm.to_numpy(),
                x=list(matriz_norm.columns),
                y=list(matriz_norm.index),
                text=matriz.to_numpy(),
                texttemplate="<b>%{text}</b>",
                textfont=dict(
                    family="Inter, sans-serif",
                    size=13,
                    color="#1F2937",
                ),
                colorscale=escala_totale,  # type: ignore[arg-type]
                showscale=True,
                colorbar=dict(
                    title=dict(
                        text="Intensidade",
                        font=dict(
                            family="Inter, sans-serif",
                            size=11,
                            color="#4B5563",
                        ),
                    ),
                    thickness=12,
                    len=0.7,
                    tickfont=dict(
                        family="Inter, sans-serif",
                        size=10,
                        color="#6B7280",
                    ),
                    outlinewidth=0,
                ),
                hovertemplate=(
                    "<b>%{y}</b><br>"
                    "Indicador: <b>%{x}</b><br>"
                    "Valor: <b>%{text}</b><br>"
                    "Intensidade: %{z:.0%}"
                    "<extra></extra>"
                ),
                xgap=4,
                ygap=4,
            )
        )

        fig.update_layout(
            title=dict(
                text="<b>Intensidade relativa por indicador</b>",
                font=dict(
                    family="Manrope, sans-serif",
                    size=16,
                    color="#012869",
                ),
                x=0.01,
                xanchor="left",
            ),
            height=340,
            margin=dict(t=70, b=40, l=90, r=40),
            paper_bgcolor="white",
            plot_bgcolor="white",
            xaxis=dict(
                side="bottom",
                tickfont=dict(
                    family="Inter, sans-serif",
                    size=12,
                    color="#374151",
                ),
                showgrid=False,
                zeroline=False,
            ),
            yaxis=dict(
                tickfont=dict(
                    family="Inter, sans-serif",
                    size=12,
                    color="#012869",
                ),
                showgrid=False,
                zeroline=False,
                autorange="reversed",
            ),
        )
        return fig

    # ---------- Insights ----------

    @staticmethod
    def gerar_insights(
        df_bases: pd.DataFrame,
        total: dict[str, object],
    ) -> list[tuple[str, TipoInsight]]:
        insights: list[tuple[str, TipoInsight]] = []

        idx_top_os = int(cast(int, df_bases["Média OS"].idxmax()))
        top_os = df_bases.loc[idx_top_os]
        insights.append((
            f"🥇 **{top_os['BASE']}** possui a maior Média OS: "
            f"**{Utils.fmt_float(top_os['Média OS'])}** OS por rota.",
            "ok",
        ))

        idx_low_os = int(cast(int, df_bases["Média OS"].idxmin()))
        low_os = df_bases.loc[idx_low_os]
        if str(low_os["BASE"]) != str(top_os["BASE"]):
            insights.append((
                f"⬇️ **{low_os['BASE']}** apresenta a menor Média OS: "
                f"**{Utils.fmt_float(low_os['Média OS'])}** — "
                f"avaliar produtividade.",
                "alerta",
            ))

        idx_max_os = int(cast(int, df_bases["OS"].idxmax()))
        base_max = df_bases.loc[idx_max_os]
        total_os = Utils.to_float(total.get("OS"))
        valor_base = Utils.to_float(base_max["OS"])
        perc = (valor_base / total_os * 100) if total_os > 0 else 0.0
        insights.append((
            f"📊 **{base_max['BASE']}** concentra "
            f"**{Utils.fmt_float(perc, 1)}%** do total de OS "
            f"({Utils.fmt_int(valor_base)} de {Utils.fmt_int(total_os)}).",
            "info",
        ))

        total_mont = Utils.to_float(total.get("Montados"))
        if total_mont > 0:
            aprov = total_os / total_mont
            tipo: TipoInsight = (
                "ok" if aprov >= 8
                else "alerta" if aprov >= 5
                else "critico"
            )
            insights.append((
                f"🎯 Aproveitamento geral: **{Utils.fmt_float(aprov)}** "
                f"OS por montado "
                f"({Utils.fmt_int(total_mont)} montados no total).",
                tipo,
            ))

        for col in Config.COLUNAS_INDICADORES:
            zerados: list[str] = (
                df_bases[df_bases[col] == 0]["BASE"].astype(str).tolist()
            )
            if zerados and len(zerados) < len(df_bases):
                insights.append((
                    f"⚠️ Indicador **{col}** está zerado em: "
                    + ", ".join(f"**{b}**" for b in zerados),
                    "alerta",
                ))

        return insights


# ============================================================
# INTERFACE DO USUÁRIO
# ============================================================

class UI:
    """Interface do usuário."""

    @staticmethod
    def reiniciar_painel() -> None:
        """
        Incrementa o contador de reset e limpa chaves antigas
        do session_state relacionadas a uploads e inputs.
        """
        # Remove chaves antigas de uploads e montados
        chaves_para_limpar = [
            k for k in list(st.session_state.keys())
            if str(k).startswith(("up_", "mont_", "btn_"))
        ]
        for k in chaves_para_limpar:
            try:
                del st.session_state[k]
            except KeyError:
                pass

        # Incrementa contador de reset
        st.session_state["_reset_counter"] = (
            int(st.session_state.get("_reset_counter", 0)) + 1
        )
        st.rerun()

    @staticmethod
    def validar_colunas(nome_base: str, df: pd.DataFrame) -> list[str]:
        faltantes = [
            c for c in Config.COLUNAS_ESPERADAS
            if Utils.localizar_coluna(df, c) is None
        ]
        if faltantes:
            render_insight(
                f"**{nome_base}**: colunas não encontradas: "
                + ", ".join(f"`{c}`" for c in faltantes),
                tipo="alerta",  # type: ignore[arg-type]
            )
        return faltantes

    @staticmethod
    def mostrar_interface() -> None:
        st.set_page_config(
            page_title=Config.PAGE_TITLE,
            page_icon=Config.PAGE_ICON,
            layout=Config.LAYOUT,
        )
        aplicar_estilo()
        render_hero(
            titulo="📊 Rota Geral | TOTALE",
            subtitulo=("Ponto de partida da roteirização, unificando os dados das bases operacionais"),
            badge="Operacional",
        )
        UI._injetar_css()

        render_section_header(
            icon="📁",
            title="Importação das bases",
            badge="CSV ou XLSX",
        )
        arquivos = UI._mostrar_uploaders()

        render_section_header(
            icon="📝",
            title="Montados (digitação manual)",
            badge="Entrada",
        )
        montados_dict = UI._mostrar_inputs_montados()

        UI._mostrar_botoes_acao(arquivos, montados_dict)

    @staticmethod
    def _injetar_css() -> None:
        st.markdown("""
<style>

/* ============ CONTAINER GERAL ============ */
.block-container {
    padding-top: 1.5rem !important;
    padding-bottom: 3rem !important;
}

/* ============ FILE UPLOADERS ============ */
[data-testid="stFileUploader"] {
    background: linear-gradient(180deg, #FFFFFF 0%, #F9FAFB 100%);
    border: 2px dashed #CBD5E1;
    border-radius: 12px;
    padding: 8px;
    transition: all 0.25s ease;
}
[data-testid="stFileUploader"]:hover {
    border-color: #012869;
    background: linear-gradient(180deg, #FFFFFF 0%, #EFF6FF 100%);
    box-shadow: 0 4px 12px rgba(1, 40, 105, 0.08);
}
[data-testid="stFileUploader"] section {
    background: transparent;
    border: none;
}
[data-testid="stFileUploader"] button {
    background: linear-gradient(135deg, #012869 0%, #023A9E 100%) !important;
    color: white !important;
    border: none !important;
    border-radius: 6px !important;
    font-weight: 600 !important;
}

/* ============ NUMBER INPUTS ============ */
[data-testid="stNumberInput"] input {
    border-radius: 8px !important;
    border: 1.5px solid #E2E8F0 !important;
    font-weight: 600 !important;
    font-size: 15px !important;
    color: #012869 !important;
    transition: all 0.2s ease;
}
[data-testid="stNumberInput"] input:focus {
    border-color: #012869 !important;
    box-shadow: 0 0 0 3px rgba(1, 40, 105, 0.12) !important;
}

/* ============ BOTÃO PROCESSAR ============ */
.stButton > button[kind="primary"] {
    background: linear-gradient(135deg, #012869 0%, #023A9E 40%, #F37C04 100%) !important;
    color: white !important;
    border: none !important;
    padding: 14px 28px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    letter-spacing: 0.5px !important;
    box-shadow: 0 4px 14px rgba(1, 40, 105, 0.30) !important;
    transition: all 0.25s ease !important;
}
.stButton > button[kind="primary"]:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 22px rgba(1, 40, 105, 0.40) !important;
}

/* ============ DOWNLOAD BUTTON ============ */
.stDownloadButton > button {
    background: linear-gradient(135deg, #059669 0%, #10B981 100%) !important;
    color: white !important;
    border: none !important;
    padding: 12px 20px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    box-shadow: 0 4px 12px rgba(5, 150, 105, 0.28) !important;
    transition: all 0.25s ease !important;
}
.stDownloadButton > button:hover {
    transform: translateY(-2px);
    box-shadow: 0 8px 20px rgba(5, 150, 105, 0.40) !important;
}

/* ============ BOTÃO REINICIAR ============ */
.stButton > button[kind="secondary"] {
    background: linear-gradient(135deg, #F3F4F6 0%, #E5E7EB 100%) !important;
    color: #374151 !important;
    border: 1.5px solid #D1D5DB !important;
    padding: 14px 28px !important;
    border-radius: 10px !important;
    font-weight: 700 !important;
    font-size: 15px !important;
    letter-spacing: 0.5px !important;
    box-shadow: 0 2px 8px rgba(0,0,0,0.06) !important;
    transition: all 0.25s ease !important;
}
.stButton > button[kind="secondary"]:hover {
    background: linear-gradient(135deg, #FEE2E2 0%, #FECACA 100%) !important;
    border-color: #F87171 !important;
    color: #991B1B !important;
    transform: translateY(-2px);
    box-shadow: 0 6px 16px rgba(220, 38, 38, 0.20) !important;
}

/* ============ EXPANDER ============ */
[data-testid="stExpander"] {
    border: 1px solid #E2E8F0 !important;
    border-radius: 12px !important;
    background: #FFFFFF !important;
    box-shadow: 0 2px 6px rgba(0,0,0,0.04) !important;
}
[data-testid="stExpander"] summary {
    font-weight: 600 !important;
    color: #012869 !important;
}

/* ═══════════════════════════════════════════════════════════
   TABELA CORPORATIVA — Totale
   Layout fixo com colunas padronizadas
   ═══════════════════════════════════════════════════════════ */
.tabela-rota-wrapper {
    background: white;
    border-radius: 16px;
    padding: 0;
    box-shadow:
        0 10px 30px rgba(0,0,0,0.10),
        0 2px 6px rgba(0,0,0,0.06);
    margin: 20px 0 32px 0;
    overflow: hidden;
    border: 1px solid #E2E8F0;
}
.tabela-rota {
    width: 100%;
    border-collapse: separate;
    border-spacing: 0;
    font-family: 'Inter', sans-serif;
    font-size: 13.5px;
    table-layout: fixed; /* 🔑 Força larguras padronizadas */
}

/* ═════════ LARGURAS PADRONIZADAS ═════════ */
/* BASE = coluna maior; demais = mesma largura */
.tabela-rota th:first-child,
.tabela-rota td:first-child {
    width: 11%;
    min-width: 110px;
}
.tabela-rota th:not(:first-child),
.tabela-rota td:not(:first-child) {
    width: 7.4%; /* (100 - 11) / 12 colunas restantes ≈ 7.4% */
    min-width: 78px;
}

/* ═════════ CABEÇALHO — LARANJA TOTALE ═════════ */
.tabela-rota thead th {
    background: linear-gradient(135deg,
        #F37C04 0%,
        #E85D04 55%,
        #C44100 100%);
    color: white;
    font-weight: 700;
    text-transform: uppercase;
    letter-spacing: 0.9px;
    font-size: 11.5px;
    padding: 18px 8px;
    text-align: center;
    border: none;
    text-shadow: 0 1px 2px rgba(0,0,0,0.25);
    box-shadow: inset 0 -3px 0 rgba(0,0,0,0.12);
    white-space: nowrap;
    overflow: hidden;
    text-overflow: ellipsis;
}
.tabela-rota thead th:first-child { border-top-left-radius: 14px; }
.tabela-rota thead th:last-child  { border-top-right-radius: 14px; }

/* ═════════ CORPO ═════════ */
.tabela-rota tbody td {
    padding: 16px 8px;
    text-align: center;
    border-bottom: 1px solid #F1F5F9;
    font-variant-numeric: tabular-nums;
    font-weight: 600;
    font-size: 14px;
    letter-spacing: 0.2px;
    transition: filter 0.15s ease;
    overflow: hidden;
    text-overflow: ellipsis;
    white-space: nowrap;
}
.tabela-rota tbody tr:hover td {
    filter: brightness(1.05) saturate(1.08);
}

/* ═════════ COLUNA BASE — AZUL TOTALE ═════════ */
.tabela-rota td.col-base {
    background: linear-gradient(135deg,
        #012869 0%,
        #023A9E 55%,
        #1E5FCC 100%);
    color: white;
    font-weight: 800;
    text-align: center;
    letter-spacing: 0.8px;
    text-transform: uppercase;
    font-size: 13px;
    text-shadow: 0 1px 2px rgba(0,0,0,0.30);
    box-shadow:
        inset -3px 0 6px rgba(0,0,0,0.15),
        inset 0 1px 0 rgba(255,255,255,0.08);
    padding: 18px 10px;
}

/* ═════════ LINHA TOTAL — PRETO/CINZA/PRATA ═════════ */
.tabela-rota tbody tr.linha-total td {
    background: linear-gradient(135deg,
        #1F2937 0%,
        #374151 45%,
        #4B5563 80%,
        #6B7280 100%) !important;
    color: white !important;
    font-weight: 800 !important;
    font-size: 14.5px !important;
    letter-spacing: 0.5px !important;
    text-shadow: 0 1px 2px rgba(0,0,0,0.35);
    border-bottom: none !important;
    padding: 20px 8px !important;
    box-shadow: inset 0 2px 0 rgba(255,255,255,0.05);
}
.tabela-rota tbody tr.linha-total td.col-base {
    background: linear-gradient(135deg,
        #0F172A 0%,
        #1F2937 55%,
        #334155 100%) !important;
    letter-spacing: 1.2px;
}
.tabela-rota tbody tr.linha-total td:first-child {
    border-bottom-left-radius: 14px;
}
.tabela-rota tbody tr.linha-total td:last-child {
    border-bottom-right-radius: 14px;
}

/* ═══════════════════════════════════════════════════════════
   MAPA DE CALOR POR RANKING — Verde/Amarelo/Vermelho
   1º=Verde | 2º=Amarelo | 3º=Vermelho
   ═══════════════════════════════════════════════════════════ */

/* 🟢 1º LUGAR — Melhor desempenho */
.cel-heat-verde {
    background: linear-gradient(180deg, #D1FAE5 0%, #A7F3D0 100%);
    color: #065F46;
    font-weight: 700;
    border-left: 3px solid #10B981;
    position: relative;
}
.cel-heat-verde::after {
    content: "▲";
    position: absolute;
    top: 3px;
    right: 5px;
    font-size: 8px;
    color: #10B981;
    opacity: 0.75;
}

/* 🟡 2º LUGAR — Intermediário */
.cel-heat-amarelo {
    background: linear-gradient(180deg, #FEF3C7 0%, #FDE68A 100%);
    color: #92400E;
    font-weight: 700;
    border-left: 3px solid #F59E0B;
    position: relative;
}
.cel-heat-amarelo::after {
    content: "●";
    position: absolute;
    top: 3px;
    right: 5px;
    font-size: 8px;
    color: #F59E0B;
    opacity: 0.75;
}

/* 🔴 3º LUGAR — Menor valor */
.cel-heat-vermelho {
    background: linear-gradient(180deg, #FEE2E2 0%, #FECACA 100%);
    color: #991B1B;
    font-weight: 700;
    border-left: 3px solid #EF4444;
    position: relative;
}
.cel-heat-vermelho::after {
    content: "▼";
    position: absolute;
    top: 3px;
    right: 5px;
    font-size: 8px;
    color: #EF4444;
    opacity: 0.75;
}

/* CINZA — Fallback / Empate */
.cel-cinza-corp {
    background: linear-gradient(180deg, #F3F4F6 0%, #E5E7EB 100%);
    color: #374151;
    font-weight: 600;
}

/* ═════════ DESTAQUES ESPECIAIS (Médias) ═════════ */
.cel-media-os {
    background: linear-gradient(180deg, #1E5FCC 0%, #023A9E 100%);
    color: #FFFFFF;
    font-weight: 800;
    text-shadow: 0 1px 2px rgba(0,0,0,0.25);
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.10);
}
.cel-media-mont {
    background: linear-gradient(180deg, #F37C04 0%, #C44100 100%);
    color: #FFFFFF;
    font-weight: 800;
    text-shadow: 0 1px 2px rgba(0,0,0,0.25);
    box-shadow: inset 0 0 0 1px rgba(0,0,0,0.25);
}

/* Hover destacado no heatmap */
.tabela-rota tbody tr:hover .cel-heat-verde,
.tabela-rota tbody tr:hover .cel-heat-amarelo,
.tabela-rota tbody tr:hover .cel-heat-vermelho {
    filter: brightness(1.08) saturate(1.15);
    box-shadow: inset 0 0 0 1px rgba(0,0,0,0.05);
}

/* ═════════ DESTAQUES ESPECIAIS (Médias) ═════════ */
.cel-media-os {
    background: linear-gradient(180deg, #1E5FCC 0%, #023A9E 100%);
    color: #FFFFFF;
    font-weight: 800;
    text-shadow: 0 1px 2px rgba(0,0,0,0.25);
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.10);
}
.cel-media-mont {
    background: linear-gradient(180deg, #F37C04 0%, #C44100 100%);
    color: #FFFFFF;
    font-weight: 800;
    text-shadow: 0 1px 2px rgba(0,0,0,0.25);
    box-shadow: inset 0 0 0 1px rgba(255,255,255,0.10);
}

/* Hover destacado no heatmap */
.tabela-rota tbody tr:hover .cel-heat-verde,
.tabela-rota tbody tr:hover .cel-heat-amarelo,
.tabela-rota tbody tr:hover .cel-heat-vermelho {
    filter: brightness(1.08) saturate(1.15);
    transform: scale(1.01);
    z-index: 2;
    box-shadow: 0 2px 8px rgba(0,0,0,0.10);
}
</style>
""", unsafe_allow_html=True)

    @staticmethod
    def _mostrar_uploaders() -> dict[str, Optional["UploadedFile"]]:
        reset = int(st.session_state.get("_reset_counter", 0))
        col_up = st.columns(3)
        arquivos: dict[str, Optional["UploadedFile"]] = {}
        for i, base in enumerate(Config.BASES):
            with col_up[i]:
                arquivos[base] = cast(
                    "Optional[UploadedFile]",
                    st.file_uploader(
                        f"Base {base}",
                        type=["csv", "xlsx"],
                        key=f"up_{base}_{reset}",
                    ),
                )
        return arquivos

    @staticmethod
    def _mostrar_inputs_montados() -> dict[str, int]:
        reset = int(st.session_state.get("_reset_counter", 0))
        col_mont = st.columns(3)
        montados_dict: dict[str, int] = {}
        for i, base in enumerate(Config.BASES):
            with col_mont[i]:
                valor = st.number_input(
                    f"Montados {base}",
                    min_value=0,
                    step=1,
                    value=0,
                    key=f"mont_{base}_{reset}",
                )
                montados_dict[base] = int(valor)
        return montados_dict

    @staticmethod
    def _mostrar_botoes_acao(
        arquivos: dict[str, Optional["UploadedFile"]],
        montados_dict: dict[str, int],
    ) -> None:
        st.markdown("<br>", unsafe_allow_html=True)
        col_btn1, col_btn2 = st.columns([3, 1])

        with col_btn1:
            processar: bool = st.button(
                "🚀 Processar bases",
                type="primary",
                use_container_width=True,
                key="btn_processar",
            )

        with col_btn2:
            if st.button(
                "🔄 Reiniciar painel",
                type="secondary",
                use_container_width=True,
                key="btn_reiniciar",
            ):
                UI.reiniciar_painel()

        if processar:
            UI._processar_bases(arquivos, montados_dict)

    @staticmethod
    def _processar_bases(
        arquivos: dict[str, Optional["UploadedFile"]],
        montados_dict: dict[str, int],
    ) -> None:
        if not all(arquivos.values()):
            render_insight(
                "Selecione as **três bases** antes de processar.",
                tipo="critico",  # type: ignore[arg-type]
            )
            st.stop()

        with st.spinner("Processando as bases..."):
            bases_dfs: dict[str, Optional[pd.DataFrame]] = {
                nome: DataProcessor.carregar_arquivo(arq)
                for nome, arq in arquivos.items()
            }
            if any(df is None for df in bases_dfs.values()):
                st.stop()

            bases_validas: dict[str, pd.DataFrame] = {
                nome: df
                for nome, df in bases_dfs.items()
                if df is not None
            }

            for nome, df in bases_validas.items():
                UI.validar_colunas(nome, df)

            resultado: list[dict[str, Union[str, int, float]]] = [
                DataProcessor.processar_base(nome, df, montados_dict[nome])
                for nome, df in bases_validas.items()
            ]
            df_bases = pd.DataFrame(resultado)[Config.ORDEM_COLUNAS]
            df_final = DataProcessor.adicionar_linha_total(df_bases)

        render_insight("Bases processadas com sucesso.", tipo="ok")  # type: ignore[arg-type]

        total: dict[str, object] = {
            str(k): v
            for k, v in df_final.iloc[-1].to_dict().items()
        }

        UI._mostrar_kpis_gerais(total)
        UI._mostrar_kpis_por_base(df_bases)

        render_section_header(icon="📋", title="Resultado consolidado", badge="Detalhamento")
        Visualization.renderizar_tabela_html(df_final)
        
        # Legenda do mapa de calor
        st.markdown("""
        <div style="
            display: flex;
            gap: 16px;
            justify-content: flex-end;
            align-items: center;
            padding: 8px 4px;
            font-size: 12px;
            font-family: 'Inter', sans-serif;
            color: #4B5563;
            font-weight: 600;
        ">
            <span style="display:flex;align-items:center;gap:6px;">
                <span style="width:14px;height:14px;background:linear-gradient(180deg,#D1FAE5,#A7F3D0);border-left:3px solid #10B981;border-radius:3px;"></span>
                Melhor
            </span>
            <span style="display:flex;align-items:center;gap:6px;">
                <span style="width:14px;height:14px;background:linear-gradient(180deg,#FEF3C7,#FDE68A);border-left:3px solid #F59E0B;border-radius:3px;"></span>
                Intermediário
            </span>
            <span style="display:flex;align-items:center;gap:6px;">
                <span style="width:14px;height:14px;background:linear-gradient(180deg,#FEE2E2,#FECACA);border-left:3px solid #EF4444;border-radius:3px;"></span>
                Pior
            </span>
        </div>
        """, unsafe_allow_html=True)

        UI._mostrar_graficos(df_bases)
        UI._mostrar_insights(df_bases, total)
        UI._mostrar_diagnostico(bases_validas)

        render_section_header(icon="⬇️", title="Exportação", badge="Excel")
        st.download_button(
            "Baixar resultado em Excel",
            data=Visualization.gerar_excel(df_final),
            file_name="rota_geral.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
        )

    @staticmethod
    def _mostrar_kpis_gerais(total: dict[str, object]) -> None:
        render_section_header(icon="📈", title="Indicadores gerais", badge="Total consolidado")
        k = st.columns(4)
        render_kpi(k[0], "Total OS", Utils.fmt_int(total.get("OS")), "Soma de Total de tarefas", "azul")
        render_kpi(k[1], "Rotas", Utils.fmt_int(total.get("Rotas")), "Logins únicos", "laranja")
        render_kpi(k[2], "Média OS", Utils.fmt_float(total.get("Média OS")), "OS por rota", "verde")
        render_kpi(k[3], "Média Montados", Utils.fmt_float(total.get("Média Montados")), "OS por montado", "vermelho")

    @staticmethod
    def _mostrar_kpis_por_base(df_bases: pd.DataFrame) -> None:
        render_section_header(icon="🏢", title="Desempenho por base", badge="Comparativo rápido")
        cols_bases = st.columns(len(df_bases))
        for i, (_, row) in enumerate(df_bases.iterrows()):
            with cols_bases[i]:
                st.markdown(f"##### 🏷️ {row['BASE']}")
                render_kpi_sm(st, "OS", Utils.fmt_int(row["OS"]), f"{Utils.fmt_int(row['Rotas'])} rotas", "azul")
                render_kpi_sm(st, "Média OS", Utils.fmt_float(row["Média OS"]), "OS por rota", "verde")
                render_kpi_sm(st, "Média Montados", Utils.fmt_float(row["Média Montados"]), "OS por montado", "laranja")

    @staticmethod
    def _mostrar_graficos(df_bases: pd.DataFrame) -> None:
        render_section_header(icon="📊", title="Análise visual", badge="Gráficos comparativos")
        g1, g2 = st.columns(2)
        with g1:
            st.plotly_chart(Visualization.grafico_participacao(df_bases), use_container_width=True)
        with g2:
            st.plotly_chart(Visualization.grafico_comparativo_medias(df_bases), use_container_width=True)
        st.plotly_chart(Visualization.grafico_heatmap(df_bases), use_container_width=True)

    @staticmethod
    def _mostrar_insights(df_bases: pd.DataFrame, total: dict[str, object]) -> None:
        render_section_header(icon="💡", title="Insights automáticos", badge="Leitura dos dados")
        for msg, tipo in Visualization.gerar_insights(df_bases, total):
            render_insight(msg, tipo=tipo)  # type: ignore[arg-type]

    @staticmethod
    def _mostrar_diagnostico(bases_validas: dict[str, pd.DataFrame]) -> None:
        with st.expander("🔎 Conferência técnica dos cálculos"):
            for nome, df in bases_validas.items():
                st.markdown(f"**{nome}**")
                st.write(f"Linhas na base: {Utils.fmt_int(len(df))}")
                st.write("Colunas Tipo O.S encontradas:", Utils.localizar_colunas_tipo_os(df))
                st.write("Linhas com GPON_FLAG = 1:", int((df["GPON_FLAG"] == 1).sum()))
                st.divider()


# ============================================================
# EXECUÇÃO PRINCIPAL
# ============================================================

if __name__ == "__main__":
    UI.mostrar_interface()
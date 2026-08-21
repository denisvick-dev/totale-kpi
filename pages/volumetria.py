"""
Gestão de Volumetria
Arquivo: pages/volumetria.py
"""
from __future__ import annotations

import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Optional, cast

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from streamlit_gsheets import GSheetsConnection
import streamlit as st

# ====================================================
# IMPORTAÇÃO DOS COMPONENTES CORPORATIVOS
# ====================================================
from components.componentes import (
    aplicar_estilo,
    render_hero,
    render_insight,
    render_kpi,
    render_section_header,
    render_dataframe,
)

# ==========================================================
# CONFIGURAÇÃO E CONSTANTES
# ==========================================================
st.set_page_config(
    page_title="Volumetria",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)

class Config:
    URL_GSHEETS = "https://docs.google.com/spreadsheets/d/1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
    META_EXECUCAO = 0.80

    COL_STATUS = "STATUS CONTRATO"
    COL_TOTAL = "TOTAL DE TAREFAS"
    COL_TECNICO = "TÉCNICO"
    COL_MONITOR = "MONITOR"
    COL_REGIAO = "REGIÃO"

    STATUS_ORDEM = ["Executada", "Não Executada", "Pendente"]
    CONTRATO_VALORES_VAZIOS = {"", "NAN", "NONE", "N/A", "NA", "-", "0", "NULL"}


# ==========================================================
# UTILITÁRIOS
# ==========================================================
class Utils:
    @staticmethod
    def remover_acentos(valor: Any) -> str:
        if pd.isna(valor):
            return ""
        return unicodedata.normalize("NFKD", str(valor)).encode("ASCII", "ignore").decode("ASCII")

    @staticmethod
    def normalizar_chave(serie: pd.Series) -> pd.Series:
        s = serie.copy()
        return (
            s.where(s.notna(), "")
            .astype(str)
            .str.strip()
            .str.upper()
            .apply(Utils.remover_acentos)
        )

    @staticmethod
    def normalizar_login(serie: pd.Series) -> pd.Series:
        return Utils.normalizar_chave(serie).str.replace(r"\.0$", "", regex=True)

    @staticmethod
    def buscar_coluna(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
        cols_map = {Utils.normalizar_chave(pd.Series([c]))[0]: c for c in df.columns}
        for alias in aliases:
            chave = Utils.normalizar_chave(pd.Series([alias]))[0]
            if chave in cols_map:
                return cols_map[chave]
        return None

    @staticmethod
    def classificar_status(status_os: pd.Series) -> pd.Series:
        s = Utils.normalizar_chave(status_os)
        nao_exec = s.str.contains(r"NAO\s*EXECUT", regex=True, na=False)
        exec_ = s.str.contains(r"EXECUT", regex=True, na=False) & ~nao_exec
        return pd.Series(
            np.select(
                [exec_, nao_exec],
                ["Executada", "Não Executada"],
                default="Pendente",
            ),
            index=status_os.index,
            dtype="object",
        )

    @staticmethod
    def contrato_valido(serie: pd.Series) -> pd.Series:
        norm = serie.astype(str).str.strip().str.upper().apply(Utils.remover_acentos)
        return ~norm.isin(Config.CONTRATO_VALORES_VAZIOS)


# ==========================================================
# CARREGAMENTO DE DADOS
# ==========================================================
class DataLoader:
    @staticmethod
    @st.cache_data(show_spinner=False)
    def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
        if not file_bytes:
            raise ValueError("O arquivo enviado está vazio.")

        nome = filename.lower()
        if nome.endswith(".xlsx") or nome.endswith(".xls"):
            return pd.read_excel(BytesIO(file_bytes), engine="openpyxl" if nome.endswith(".xlsx") else None)
        if nome.endswith(".csv"):
            try:
                return pd.read_csv(BytesIO(file_bytes), sep=None, engine="python", encoding="utf-8-sig")
            except UnicodeDecodeError:
                return pd.read_csv(BytesIO(file_bytes), sep=None, engine="python", encoding="latin1")
        raise ValueError("Formato não suportado. Use .xlsx, .xls ou .csv.")

    @staticmethod
    @st.cache_data(ttl=600, show_spinner=False)
    def buscar_hierarquia_gsheets() -> pd.DataFrame:
        try:
            conn = st.connection("gsheets", type=GSheetsConnection)
            raw = conn.read(spreadsheet=Config.URL_GSHEETS)
            if raw is None or raw.empty:
                return pd.DataFrame()
            
            col_login = Utils.buscar_coluna(raw, ["LOGIN", "MATRÍCULA", "ID"])
            col_tec = Utils.buscar_coluna(raw, ["TÉCNICO", "NOME"])
            col_mon = Utils.buscar_coluna(raw, ["MONITOR", "GESTOR"])
            
            if not col_login:
                return pd.DataFrame()
                
            df_gs = pd.DataFrame({
                "__LOGIN_KEY": Utils.normalizar_login(raw[col_login]),
                "__TEC_GS": raw[col_tec] if col_tec else "",
                "__MON_GS": raw[col_mon] if col_mon else "",
            })
            return df_gs.drop_duplicates("__LOGIN_KEY")
        except Exception:
            return pd.DataFrame()

    @staticmethod
    def preparar_base(df: pd.DataFrame, df_gs: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        diag = {"Inicial": len(df)}

        col_con = Utils.buscar_coluna(df, ["CONTRATO", "Nº CONTRATO", "NUMERO CONTRATO", "NUM CONTRATO"])
        if col_con:
            valida = Utils.contrato_valido(df[col_con])
            rem = (~valida).sum()
            df = df[valida].copy()
            diag["Removidos por contrato vazio"] = int(rem)
        else:
            st.warning("⚠️ Coluna de contrato não encontrada.")

        col_atv = Utils.buscar_coluna(df, ["STATUS DA ATIVIDADE"])
        if col_atv:
            susp = Utils.normalizar_chave(df[col_atv]).str.contains("SUSP", na=False)
            df = df[~susp].copy()

        col_os1 = Utils.buscar_coluna(df, ["STATUS DA O.S 1", "STATUS OS 1"])
        if not col_os1:
            st.error("Coluna 'Status da O.S 1' não encontrada!")
            st.stop()
        df[Config.COL_STATUS] = Utils.classificar_status(df[col_os1])

        col_qtd = Utils.buscar_coluna(df, ["TOTAL DE TAREFAS", "QUANTIDADE"])
        if col_qtd:
            df[Config.COL_TOTAL] = pd.to_numeric(df[col_qtd], errors="coerce").fillna(1)
        else:
            df[Config.COL_TOTAL] = 1
        df[Config.COL_TOTAL] = df[Config.COL_TOTAL].clip(lower=0)

        col_log = Utils.buscar_coluna(df, ["LOGIN DO TÉCNICO", "LOGIN", "USUÁRIO"])
        if col_log and not df_gs.empty:
            df["__LOGIN_KEY"] = Utils.normalizar_login(df[col_log])
            df = df.merge(df_gs, on="__LOGIN_KEY", how="left")
            
        for c in ("__TEC_GS", "__MON_GS"):
            if c not in df.columns:
                df[c] = np.nan

        col_tec_b = Utils.buscar_coluna(df, ["TÉCNICO", "NOME"]) or col_log
        col_mon_b = Utils.buscar_coluna(df, ["MONITOR", "GESTOR"])

        base_tec = df[col_tec_b] if col_tec_b else pd.Series(np.nan, index=df.index)
        base_mon = df[col_mon_b] if col_mon_b else pd.Series(np.nan, index=df.index)

        tec_gs = df["__TEC_GS"].where(df["__TEC_GS"].notna(), "")
        mon_gs = df["__MON_GS"].where(df["__MON_GS"].notna(), "")
        
        df[Config.COL_TECNICO] = tec_gs.mask(tec_gs.astype(str).str.strip().eq(""), base_tec)
        df[Config.COL_MONITOR] = mon_gs.mask(mon_gs.astype(str).str.strip().eq(""), base_mon)
        
        df[Config.COL_TECNICO] = df[Config.COL_TECNICO].where(
            df[Config.COL_TECNICO].notna() & df[Config.COL_TECNICO].astype(str).str.strip().ne(""),
            "NÃO MAPEADO",
        )
        df[Config.COL_MONITOR] = df[Config.COL_MONITOR].where(
            df[Config.COL_MONITOR].notna() & df[Config.COL_MONITOR].astype(str).str.strip().ne(""),
            "SEM MONITOR",
        )

        col_cid = Utils.buscar_coluna(df, ["CIDADE", "LOCALIDADE"])
        cidade = Utils.normalizar_chave(df[col_cid]) if col_cid else pd.Series("", index=df.index)
        
        df[Config.COL_REGIAO] = np.select(
            [
                cidade.isin(["SAO PAULO"]),
                cidade.isin(["GUARULHOS", "ARUJA", "MOGI DAS CRUZES", "SUZANO", "ITAQUAQUECETUBA", "FERRAZ DE VASCONCELOS", "POA"]),
                cidade.isin(["SANTO ANDRE", "SAO BERNARDO DO CAMPO", "SAO CAETANO DO SUL", "DIADEMA", "MAUA", "RIBEIRAO PIRES", "RIO GRANDE DA SERRA"]),
            ],
            ["LESTE", "GRU", "ABCDM"],
            default="OUTRAS",
        )

        diag["Final"] = len(df)
        df.attrs["diagnostico"] = diag
        return df


# ==========================================================
# CÁLCULOS
# ==========================================================
def calcular_kpis(df: pd.DataFrame) -> Dict[str, Any]:
    if df.empty:
        return {
            "total": 0, "executadas": 0, "nao_executadas": 0,
            "pendentes": 0, "baixadas": 0, "taxa": 0.0,
            "quebra": 0.0, "projecao": 0,
        }
    k_tot = int(round(df[Config.COL_TOTAL].sum()))
    k_exe = int(df[df[Config.COL_STATUS] == "Executada"][Config.COL_TOTAL].sum())
    k_nex = int(df[df[Config.COL_STATUS] == "Não Executada"][Config.COL_TOTAL].sum())
    k_pen = int(df[df[Config.COL_STATUS] == "Pendente"][Config.COL_TOTAL].sum())
    k_bai = k_exe + k_nex
    k_tx = k_exe / k_bai if k_bai > 0 else 0.0
    k_proj = int(k_exe + k_tx * k_pen)
    
    return {
        "total": k_tot, "executadas": k_exe, "nao_executadas": k_nex,
        "pendentes": k_pen, "baixadas": k_bai, "taxa": k_tx,
        "quebra": 1.0 - k_tx, "projecao": k_proj,
    }


def calcular_volumetria(df: pd.DataFrame, grupos: List[str]) -> pd.DataFrame:
    tabela = (
        df.groupby(grupos + [Config.COL_STATUS], observed=True)[Config.COL_TOTAL]
        .sum().unstack(Config.COL_STATUS, fill_value=0).reset_index()
    )
    for s in Config.STATUS_ORDEM:
        if s not in tabela.columns:
            tabela[s] = 0
        tabela[s] = tabela[s].astype(int)
        
    tabela["Baixadas"] = tabela["Executada"] + tabela["Não Executada"]
    tabela["Total Alocado"] = tabela["Baixadas"] + tabela["Pendente"]
    tabela["Taxa Execução"] = np.where(tabela["Baixadas"] > 0, tabela["Executada"] / tabela["Baixadas"], 0.0)
    tabela["Taxa Quebra"] = 1.0 - tabela["Taxa Execução"]
    tabela["Projeção"] = (tabela["Executada"] + tabela["Taxa Execução"] * tabela["Pendente"]).astype(int)
    
    if Config.COL_TECNICO in df.columns:
        n_tec = df.groupby(grupos, observed=True)[Config.COL_TECNICO].nunique().reset_index(name="Técnicos")
        tabela = tabela.merge(n_tec, on=grupos, how="left")
        tabela["Técnicos"] = tabela["Técnicos"].fillna(0).astype(int)
        tabela["OS/Técnico"] = np.where(tabela["Técnicos"] > 0, tabela["Total Alocado"] / tabela["Técnicos"], 0.0)
        tabela["Exec/Técnico"] = np.where(tabela["Técnicos"] > 0, tabela["Executada"] / tabela["Técnicos"], 0.0)
        
    return tabela.sort_values("Total Alocado", ascending=False)


def gerar_excel(df: pd.DataFrame, nome_aba: str) -> bytes:
    output = BytesIO()
    export = df.copy()
    
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name=nome_aba[:31])
        ws = writer.sheets[nome_aba[:31]]
        hf = PatternFill("solid", fgColor="0F172A")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(color="FFFFFF", bold=True)

        for i, col in enumerate(export.columns, 1):
            valores = export[col].head(500).astype(str)
            maior = max([len(str(col))] + [len(v) for v in valores], default=10)
            ws.column_dimensions[get_column_letter(i)].width = min(max(maior + 2, 12), 40)

    output.seek(0)
    return output.getvalue()


# ==========================================================
# MAIN
# ==========================================================
def main():
    # ── APLICA ESTILO CORPORATIVO ─────────────────────────
    aplicar_estilo()

    # ── HERO ──────────────────────────────────────────────
    render_hero(
        titulo="Gestão de Volumetria",
        subtitulo="Análise executiva de performance e projeções operacionais",
        badge="Operações"
    )

    if "base_data" not in st.session_state:
        st.session_state.base_data = None

    # ── SIDEBAR CONFIGURAÇÕES ─────────────────────────────
    with st.sidebar:
        st.header("⚙️ Configurações")
        if st.button("🔄 Reiniciar Painel", use_container_width=True):
            st.session_state.base_data = None
            st.rerun()

        st.divider()

        if st.session_state.base_data is not None:
            diag = st.session_state.base_data.attrs.get("diagnostico", {})
            st.write(f"📥 **Entrada:** {diag.get('Inicial', 0)}")
            st.write(f"🗑️ **Vazios:** {diag.get('Removidos por contrato vazio', 0)}")
            st.write(f"📈 **Processadas:** {diag.get('Final', 0)}")

    # ── UPLOAD ─────────────────────────────────────────────
    if st.session_state.base_data is None:
        render_section_header("upload_file", "Importação de Dados")
        u = st.file_uploader("Selecione a base (Excel/CSV)", type=["xlsx", "xls", "csv"])
        
        if u:
            try:
                with st.spinner("Processando..."):
                    arquivo = DataLoader.ler_arquivo(u.getvalue(), u.name)
                    if arquivo.empty:
                        st.error("O arquivo não possui registros.")
                        return

                    st.session_state.base_data = DataLoader.preparar_base(
                        arquivo,
                        DataLoader.buscar_hierarquia_gsheets(),
                    )
                    st.rerun()
            except Exception as exc:
                st.error(f"Não foi possível processar o arquivo: {exc}")
                st.exception(exc)
        return

    df_full = st.session_state.base_data

    # ── FILTROS (SIDEBAR) ──────────────────────────────────
    with st.sidebar:
        st.header("🎯 Filtros")
        mons = sorted(df_full[Config.COL_MONITOR].dropna().astype(str).unique())
        sel_m = st.multiselect("Monitor", mons, default=mons)
        
        regs = sorted(df_full[Config.COL_REGIAO].dropna().astype(str).unique())
        sel_r = st.multiselect("Região", regs, default=regs)

        tecnicos = sorted(df_full[Config.COL_TECNICO].dropna().astype(str).unique())
        sel_t = st.multiselect("Técnico", tecnicos, default=tecnicos)

        statuses = Config.STATUS_ORDEM
        sel_s = st.multiselect("Status", statuses, default=statuses)

        if st.button("↩️ Limpar filtros", use_container_width=True):
            st.rerun()

    # Aplicação dos filtros
    df = df_full[
        df_full[Config.COL_MONITOR].isin(sel_m)
        & df_full[Config.COL_REGIAO].isin(sel_r)
        & df_full[Config.COL_TECNICO].isin(sel_t)
        & df_full[Config.COL_STATUS].isin(sel_s)
    ]
    
    if df.empty:
        st.warning("Nenhum dado selecionado.")
        return

    # ── RESULTADO DA BASE (INSIGHT) ────────────────────────
    regioes_ativas = ", ".join(sorted(df[Config.COL_REGIAO].unique()))
    render_insight(
        f"**Resultado da Base:** Foram encontrados **{len(df):,}** registros. <br>Regiões ativas: **{regioes_ativas}**", 
        tipo="info"
    )

    # ── KPIs DE VOLUME ─────────────────────────────────────
    kpis = calcular_kpis(df)
    
    render_section_header("query_stats", "Resumo de Volume")
    c1, c2, c3, c4 = st.columns(4)
    render_kpi(c1, "Total Alocado", f"{kpis['total']:,}", f"{kpis['pendentes']:,} pendentes", tema="azul")
    render_kpi(c2, "Executadas", f"{kpis['executadas']:,}", f"Taxa: {kpis['taxa']:.1%}", tema="verde")
    render_kpi(c3, "Não Executadas", f"{kpis['nao_executadas']:,}", f"Quebra: {kpis['quebra']:.1%}", tema="vermelho")
    render_kpi(c4, "Projeção Final", f"{kpis['projecao']:,}", "Baseado na taxa atual", tema="laranja")

    # ── ABAS PRINCIPAIS ────────────────────────────────────
    t1, t2 = st.tabs(["👥 Equipes", "📋 Base de Dados"])

    with t1:
        te = calcular_volumetria(df, [Config.COL_REGIAO, Config.COL_MONITOR])
        
        # Formatação das colunas para % e decimal via dicionário formatador
        formato = {
            "Taxa Execução": "{:.1%}",
            "Taxa Quebra": "{:.1%}",
            "OS/Técnico": "{:.1f}",
            "Exec/Técnico": "{:.1f}"
        }
        
        render_dataframe(
            te, 
            titulo="Volumetria por Equipe", 
            icone="groups", 
            fmt=cast(Any, formato), 
            height=500
        )
        
        st.download_button(
            "📥 Baixar Relatório (Excel)", 
            gerar_excel(te, "Equipes"), 
            "volumetria_equipes.xlsx",
            use_container_width=True
        )

    with t2:
        render_dataframe(
            df.head(500), 
            titulo="Visualização da Base (limite 500 linhas)", 
            icone="table_rows", 
            height=500
        )


if __name__ == "__main__":
    main()
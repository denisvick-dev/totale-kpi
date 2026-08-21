"""
quebra.py
=========
Super Relatório Corporativo de Desempenho | Quebra Operacional Unificada
Com aba de Base Completa (dados originais + Monitor/Técnico do Merge).
"""

from __future__ import annotations

import csv
from datetime import datetime
from io import BytesIO
from typing import Any, Dict, List, Optional, cast

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════
# ✅ IMPORT DOS COMPONENTES CORPORATIVOS
# ═══════════════════════════════════════════════════════
from components.componentes import (
    aplicar_estilo,
    render_dataframe,
    render_hero,
    render_insight,
    render_kpi,
    render_section_header,
)

# ═══════════════════════════════════════════════════════
# ✅ IMPORT DAS REGRAS DE NEGÓCIO (CRITÉRIOS)
# ═══════════════════════════════════════════════════════
from components.criterios import (
    VAZIOS_CONTRATO,
    classificar_tipo_servico,
    detectar_col_contrato,
    detectar_col_status_atividade,
)

# ═══════════════════════════════════════════════════════
# CONFIGURAÇÃO DA PÁGINA
# ═══════════════════════════════════════════════════════
st.set_page_config(
    page_title="Quebra Operacional | TOTALE",
    page_icon="📉",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ═══════════════════════════════════════════════════════
# CONSTANTES DE DOMÍNIO
# ═══════════════════════════════════════════════════════
class Config:
    SLA_GERAL = 0.20
    SLA_MIGRACAO = 0.25
    SLA_PME = 0.20

    URL_LISTA_ATIVOS = "https://docs.google.com/spreadsheets/d/1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
    SHEET_ID_ATIVOS = "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg"
    WORKSHEET_ATIVOS = "lista_ativos"

    COL_REGIAO = "REGIÃO"
    ORDEM_TIPOS = ["Novos Domicílios", "Migração", "PME"]

    # Colunas prioritárias exibidas na aba Base Completa
    COLS_BASE_PRIORITARIAS = [
        "CONTRATO",
        "LOGIN",
        "TÉCNICO",
        "MONITOR",
        "ORIGEM_MAPEAMENTO",
        "REGIÃO",
        "TIPO_SERVICO",
        "Status Contrato",
        "TOTAL DE TAREFAS",
        "_COL_BAIXA",
    ]


# ═══════════════════════════════════════════════════════
# UTILITÁRIOS GERAIS
# ═══════════════════════════════════════════════════════
class Utils:
    @staticmethod
    def buscar_coluna(df: pd.DataFrame, palavras: list) -> Optional[str]:
        if df is None or df.empty:
            return None
        cols = {
            str(c).strip().upper().replace(".", "").replace("_", "").replace("  ", " "): c
            for c in df.columns
        }
        for p in palavras:
            pn = str(p).strip().upper().replace(".", "").replace("_", "").replace("  ", " ")
            for cn, co in cols.items():
                if pn in cn:
                    return co
        return None

    @staticmethod
    def classificar_status(serie: pd.Series) -> pd.Series:
        s = serie.fillna("").astype(str).str.strip().str.upper()
        exe = s == "EXECUTADA"
        nex = s.isin(["NÃO EXECUTADA", "NAO EXECUTADA"])
        return pd.Series(
            np.select([exe, nex], ["Executada", "Não Executada"], default="Pendente"),
            index=serie.index,
        )

    @staticmethod
    def gerar_excel(df: pd.DataFrame, aba: str = "Dados") -> bytes:
        out = BytesIO()
        with pd.ExcelWriter(out, engine="openpyxl") as w:
            df.to_excel(w, index=False, sheet_name=aba[:31])
            ws = w.sheets[aba[:31]]
            hf = PatternFill("solid", fgColor="0F172A")
            ws.freeze_panes = "A2"
            ws.auto_filter.ref = ws.dimensions
            for row in ws.iter_rows(min_row=1, max_row=1):
                for cell in row:
                    cell.fill = hf
                    cell.font = Font(color="FFFFFF", bold=True)
            for i, col in enumerate(df.columns, 1):
                try:
                    serie_str = df[col].fillna("").astype(str)
                    tamanhos = serie_str.str.len()
                    max_len_dados = int(tamanhos.max()) if len(tamanhos) > 0 else 0
                    max_len = max(max_len_dados, len(str(col)))
                    ws.column_dimensions[get_column_letter(i)].width = min(max(max_len + 2, 12), 40)
                except Exception:
                    ws.column_dimensions[get_column_letter(i)].width = 20
        return out.getvalue()


# ═══════════════════════════════════════════════════════
# DATA LOADER E LIMPEZA DE BASE
# ═══════════════════════════════════════════════════════
class DataLoader:
    @staticmethod
    @st.cache_data(show_spinner=False)
    def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
        bio = BytesIO(file_bytes)
        try:
            if filename.lower().endswith(".csv"):
                bio.seek(0)
                amostra = bio.read(5000).decode("utf-8", errors="ignore")
                bio.seek(0)
                sep = csv.Sniffer().sniff(amostra).delimiter if amostra else ";"
                return pd.read_csv(bio, sep=sep, encoding="utf-8", dtype=str, engine="python")
            return pd.read_excel(bio, engine="openpyxl", dtype=str)
        except Exception as e:
            st.error(f"Erro ao ler arquivo: {e}")
            return pd.DataFrame()

    @staticmethod
    @st.cache_data(ttl=600, show_spinner="🔗 Conectando à Lista de Ativos...")
    def buscar_gsheets() -> pd.DataFrame:
        """
        Carrega a planilha Lista de Ativos.
        Tenta, nesta ordem:
        1) streamlit_gsheets (secrets.toml)
        2) export CSV público (gid=0)
        3) export CSV pela aba 'lista_ativos'
        """
        SHEET_ID = Config.SHEET_ID_ATIVOS
        raw = None

        # ── Método 1: conexão nativa Streamlit ──
        try:
            from streamlit_gsheets import GSheetsConnection
            conn = st.connection("gsheets", type=GSheetsConnection)
            raw = conn.read(
                spreadsheet=Config.URL_LISTA_ATIVOS,
                worksheet=Config.WORKSHEET_ATIVOS,
            )
            if raw is not None and not raw.empty:
                return DataLoader._normalizar_lista_ativos(raw, origem="gsheets_conn")
        except Exception:
            pass

        # ── Método 2: CSV público gid=0 ──
        try:
            url = f"https://docs.google.com/spreadsheets/d/{SHEET_ID}/export?format=csv&gid=0"
            raw = pd.read_csv(url)
            if raw is not None and not raw.empty:
                return DataLoader._normalizar_lista_ativos(raw, origem="csv_gid0")
        except Exception:
            pass

        # ── Método 3: CSV pela aba nomeada ──
        try:
            url = (
                f"https://docs.google.com/spreadsheets/d/{SHEET_ID}"
                f"/gviz/tq?tqx=out:csv&sheet={Config.WORKSHEET_ATIVOS}"
            )
            raw = pd.read_csv(url)
            if raw is not None and not raw.empty:
                return DataLoader._normalizar_lista_ativos(raw, origem="csv_sheet")
        except Exception:
            pass

        return pd.DataFrame()


    @staticmethod
    def _normalizar_lista_ativos(raw: pd.DataFrame, origem: str = "") -> pd.DataFrame:
        """Padroniza colunas da Lista de Ativos para Login / Técnico / Monitor / Base."""
        if raw is None or raw.empty:
            return pd.DataFrame()

        raw = raw.copy()
        raw.columns = raw.columns.astype(str).str.strip()

        rename_map: dict[str, str] = {}
        for col in raw.columns:
            cu = (
                col.upper()
                .strip()
                .replace(".", "")
                .replace("_", " ")
                .replace("  ", " ")
            )
            if cu in ("LOGIN", "MATRICULA", "MATRÍCULA", "ID", "USUARIO", "USUÁRIO"):
                rename_map[col] = "Login"
            elif cu in ("TECNICO", "TÉCNICO", "NOME", "NOME TECNICO", "NOME DO TECNICO"):
                rename_map[col] = "Técnico"
            elif cu in ("MONITOR", "GESTOR", "SUPERVISOR", "COORDENADOR", "NOME MONITOR"):
                rename_map[col] = "Monitor"
            elif cu in ("BASE", "REGIAO", "REGIÃO", "CIDADE"):
                rename_map[col] = "Base"

        raw = raw.rename(columns=rename_map)
        cols = [c for c in ["Login", "Técnico", "Monitor", "Base"] if c in raw.columns]

        if "Login" not in cols:
            return pd.DataFrame()

        raw = raw[cols].copy()
        raw["Login"] = (
            raw["Login"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.upper()
        )
        raw = raw[raw["Login"].str.strip() != ""]
        raw = raw[~raw["Login"].isin(["NAN", "NONE", "NULL", "N/A", "0", "-"])]
        raw = raw.drop_duplicates(subset=["Login"], keep="last").reset_index(drop=True)

        raw.attrs["origem_gs"] = origem
        raw.attrs["qtd_ativos"] = len(raw)
        return raw


    @staticmethod
    @st.cache_data(show_spinner=False)
    def preparar_base(df: pd.DataFrame, df_gs: pd.DataFrame) -> pd.DataFrame:
        if df.empty:
            return pd.DataFrame()

        df = df.copy()
        df.columns = df.columns.astype(str).str.strip().str.upper()
        total_importado = len(df)

        # 1. Detecta colunas originais ANTES de qualquer alteração
        col_mon_orig = Utils.buscar_coluna(df, ["MONITOR", "SUPERVISOR", "GESTOR", "COORDENADOR", "NOME MONITOR"])
        col_tec_orig = Utils.buscar_coluna(df, ["TECNICO", "NOME TECNICO", "NOME DO TECNICO", "TÉCNICO"])

        # 2. Remover Suspensos
        col_atv = detectar_col_status_atividade(df)
        n_susp = 0
        if col_atv:
            serie_atv = df[col_atv].fillna("").astype(str).str.strip().str.upper()
            mask_susp = serie_atv.str.contains("SUSP", na=False) | serie_atv.eq("SUSPENSO")
            n_susp = int(mask_susp.sum())
            df = df[~mask_susp].copy()

        # 3. Remover Contratos Vazios
        col_con = detectar_col_contrato(df)
        n_inv = 0
        if col_con:
            serie_con = (
                df[col_con].fillna("").astype(str)
                .str.strip().str.upper().str.replace(r"\.0$", "", regex=True)
            )
            mask_invalido = serie_con.isin(VAZIOS_CONTRATO)
            n_inv = int(mask_invalido.sum())
            df = df[~mask_invalido].copy()

        # 4. Total de Tarefas
        col_tot = Utils.buscar_coluna(df, ["TOTAL DE TAREFAS", "QTD TAREFAS"])
        if col_tot:
            df["TOTAL DE TAREFAS"] = (
                pd.to_numeric(df[col_tot].astype(str).str.replace(",", "."), errors="coerce")
                .fillna(1).astype(int)
            )
        else:
            df["TOTAL DE TAREFAS"] = 1

        # 5. Merge com Lista Ativos
        col_login = Utils.buscar_coluna(
            df, ["LOGIN DO TÉCNICO", "LOGIN DO TECNICO", "LOGIN", "USUÁRIO", "MATRÍCULA"]
        )
        merge_ok = False
        if col_login and not df_gs.empty and "Login" in df_gs.columns:
            df["LOGIN"] = (
                df[col_login].astype(str)
                .str.replace(r"\.0$", "", regex=True).str.strip().str.upper()
            )
            df = df.merge(df_gs, left_on="LOGIN", right_on="Login", how="left", suffixes=("", "_GS"))
            merge_ok = True
        elif col_login:
            df["LOGIN"] = df[col_login].astype(str).str.strip().str.upper()
        else:
            df["LOGIN"] = "SEM LOGIN"

        # 6. Fallback de Monitor (GS -> Arquivo -> SEM MONITOR)
        mon_gs = (
            df["Monitor_GS"] if "Monitor_GS" in df.columns
            else (df["Monitor"] if "Monitor" in df.columns else pd.Series(np.nan, index=df.index))
        )
        mon_file = (
            df[col_mon_orig] if col_mon_orig and col_mon_orig in df.columns
            else pd.Series(np.nan, index=df.index)
        )

        mon_gs_norm = (
            mon_gs.fillna(pd.NA).astype("string").str.strip()
            .replace(["", "nan", "None", "NAN", "NULL"], pd.NA)
        )
        mon_file = (
            df[col_mon_orig].fillna(pd.NA).astype("string").str.strip()
            .replace(["", "nan", "None", "NAN", "NULL"], pd.NA)
            if col_mon_orig and col_mon_orig in df.columns
            else pd.Series(pd.NA, index=df.index, dtype="string")
        )
        origem = np.where(mon_gs_norm.notna(), "Merge (Sheets)",
                 np.where(mon_file.notna(), "Arquivo Origem", "Não Mapeado"))
        df["ORIGEM_MAPEAMENTO"] = origem

        df["MONITOR"] = (
            mon_gs_norm.combine_first(mon_file)
            .fillna("SEM MONITOR")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        df.loc[df["MONITOR"].isin(["", "NAN", "NONE", "NULL", "0", "N/A"]), "MONITOR"] = "SEM MONITOR"

        # 7. Fallback de Técnico
        tec_gs = (
            df["Técnico_GS"] if "Técnico_GS" in df.columns
            else (df["Técnico"] if "Técnico" in df.columns else pd.Series(np.nan, index=df.index))
        )
        tec_file = (
            df[col_tec_orig].fillna(pd.NA).astype("string").str.strip()
            .replace(["", "nan", "None", "NAN", "NULL"], pd.NA)
            if col_tec_orig and col_tec_orig in df.columns
            else pd.Series(pd.NA, index=df.index, dtype="string")
        )
        tec_gs_norm = (
            tec_gs.fillna(pd.NA).astype("string").str.strip()
            .replace(["", "nan", "None", "NAN", "NULL"], pd.NA)
        )
        df["TÉCNICO"] = (
            tec_gs_norm.combine_first(tec_file)
            .fillna("NÃO MAPEADO")
            .astype(str)
            .str.strip()
            .str.upper()
        )
        df.loc[df["TÉCNICO"].isin(["", "NAN", "NONE", "NULL", "0", "N/A"]), "TÉCNICO"] = "NÃO MAPEADO"

        # 8. Contrato normalizado
        if col_con and col_con in df.columns:
            df["CONTRATO"] = df[col_con].astype(str).str.strip()

        # 9. Região
        col_cid = Utils.buscar_coluna(df, ["CIDADE", "LOCALIDADE"])
        cidade = (
            df[col_cid].fillna("").astype(str).str.strip().str.upper()
            if col_cid else pd.Series("", index=df.index)
        )
        df["REGIÃO"] = np.select(
            [
                cidade.isin(["SAO PAULO"]),
                cidade.isin(["GUARULHOS", "ARUJA", "MOGI", "SUZANO", "ITAQUAQUECETUBA", "POA"]),
                cidade.isin(["SANTO ANDRE", "SAO BERNARDO DO CAMPO", "SAO CAETANO DO SUL", "DIADEMA", "MAUA"]),
            ],
            ["LESTE", "GRU", "ABCDM"], default="OUTRAS",
        )

        # 10. Status e Classificação
        col_status = Utils.buscar_coluna(df, ["STATUS DA O.S 1", "STATUS OS 1", "STATUS CONTRATO"])
        df["Status Contrato"] = Utils.classificar_status(df[col_status]) if col_status else "Pendente"
        df, df["TIPO_SERVICO"] = classificar_tipo_servico(df)

        # 11. Motivo Baixa
        col_cod = Utils.buscar_coluna(df, ["CÓD DE BAIXA 1", "MOTIVO DE BAIXA"])
        df["_COL_BAIXA"] = df[col_cod].astype(str) if col_cod else ""

        # Diagnósticos
        df.attrs["diag_importado"] = total_importado
        df.attrs["diag_susp"] = n_susp
        df.attrs["diag_inv"] = n_inv
        df.attrs["diag_final"] = len(df)
        df.attrs["merge_ok"] = merge_ok

        return df


# ═══════════════════════════════════════════════════════
# MOTORES ANALÍTICOS
# ═══════════════════════════════════════════════════════
class Analise:
    @staticmethod
    def matriz_resumo(df: pd.DataFrame) -> pd.DataFrame:
        df_valid = df[df["TIPO_SERVICO"] != "Outros"].copy()
        if df_valid.empty:
            return pd.DataFrame()

        grp = df_valid.groupby(["MONITOR", "TIPO_SERVICO"]).apply(
            lambda x: pd.Series({
                "executados": x.loc[x["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"].sum(),
                "nao_executados": x.loc[x["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"].sum(),
            })
        ).reset_index()

        grp["pct"] = np.where(
            (grp["executados"] + grp["nao_executados"]) > 0,
            grp["nao_executados"] / (grp["executados"] + grp["nao_executados"]), 0.0,
        )

        pivot = grp.pivot_table(index="MONITOR", columns="TIPO_SERVICO", values="pct", fill_value=0.0)
        for t in Config.ORDEM_TIPOS:
            if t not in pivot.columns:
                pivot[t] = 0.0
        pivot = pivot[Config.ORDEM_TIPOS]

        tot_status = (
            df_valid.groupby(["MONITOR", "Status Contrato"])["TOTAL DE TAREFAS"]
            .sum().unstack(fill_value=0).reindex(pivot.index).fillna(0)
        )
        exec_col = tot_status["Executada"] if "Executada" in tot_status.columns else pd.Series(0, index=pivot.index)
        ne_col = tot_status["Não Executada"] if "Não Executada" in tot_status.columns else pd.Series(0, index=pivot.index)
        den = exec_col + ne_col

        pivot["Quebra Geral"] = np.where(den > 0, ne_col / den, 0.0)
        pivot["Total Tarefas"] = (
            df_valid.groupby("MONITOR")["TOTAL DE TAREFAS"].sum()
            .reindex(pivot.index).fillna(0).astype(int)
        )
        pivot = pivot.reset_index().rename(columns={"MONITOR": "Monitor"})

        total_row: Dict[str, Any] = {"Monitor": "Total Geral"}
        for tipo in Config.ORDEM_TIPOS:
            sub = df_valid[df_valid["TIPO_SERVICO"] == tipo]
            ex = sub.loc[sub["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"].sum()
            ne = sub.loc[sub["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"].sum()
            total_row[tipo] = ne / (ex + ne) if (ex + ne) > 0 else 0.0

        ex_g = df_valid.loc[df_valid["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"].sum()
        ne_g = df_valid.loc[df_valid["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"].sum()
        total_row["Quebra Geral"] = ne_g / (ex_g + ne_g) if (ex_g + ne_g) > 0 else 0.0
        total_row["Total Tarefas"] = int(df_valid["TOTAL DE TAREFAS"].sum())

        return pd.concat([pivot, pd.DataFrame([total_row])], ignore_index=True)

    @staticmethod
    def metricas_segmento(df_seg: pd.DataFrame) -> Dict[str, Any]:
        aloc = float(df_seg["TOTAL DE TAREFAS"].sum())
        exe = float(df_seg.loc[df_seg["Status Contrato"] == "Executada", "TOTAL DE TAREFAS"].sum())
        nex = float(df_seg.loc[df_seg["Status Contrato"] == "Não Executada", "TOTAL DE TAREFAS"].sum())
        pen = max(0.0, aloc - exe - nex)
        quebra = (nex / (exe + nex)) if (exe + nex) > 0 else 0.0
        return {"alocado": aloc, "exec": exe, "naoexec": nex, "pend": pen, "quebra": quebra}

    @staticmethod
    def ranking_tecnicos(df_seg: pd.DataFrame) -> pd.DataFrame:
        if df_seg.empty:
            return pd.DataFrame()
        df_w = df_seg.copy()
        df_w["_exec"] = np.where(df_w["Status Contrato"] == "Executada", df_w["TOTAL DE TAREFAS"], 0)
        df_w["_ne"] = np.where(df_w["Status Contrato"] == "Não Executada", df_w["TOTAL DE TAREFAS"], 0)

        df_t = df_w.groupby(["TÉCNICO", "MONITOR"]).agg(
            Alocado=("TOTAL DE TAREFAS", "sum"),
            Executadas=("_exec", "sum"),
            Nao_Executadas=("_ne", "sum"),
        ).reset_index()

        df_t["Considerado"] = df_t["Executadas"] + df_t["Nao_Executadas"]
        df_t["Quebra Atual"] = np.where(
            df_t["Considerado"] > 0, df_t["Nao_Executadas"] / df_t["Considerado"], 0.0
        )
        return df_t.sort_values("Nao_Executadas", ascending=False).reset_index(drop=True)


# ═══════════════════════════════════════════════════════
# VISÕES DE PÁGINA
# ═══════════════════════════════════════════════════════
def view_executiva(df: pd.DataFrame, meta_pct: float):
    if df.empty:
        render_insight("Sem dados para a Visão Executiva.", tipo="alerta")
        return

    df_matriz = Analise.matriz_resumo(df)
    if df_matriz.empty:
        render_insight("Não foi possível gerar a matriz de resumo.", tipo="alerta")
        return

    total_row = df_matriz[df_matriz["Monitor"] == "Total Geral"].iloc[0]

    render_section_header("analytics", "Panorama Corporativo")
    k1, k2, k3, k4 = st.columns(4)
    render_kpi(k1, "Total O.S. (Base)", f"{int(total_row['Total Tarefas']):,}", tema="azul")
    render_kpi(
        k2, "Quebra Consolidada", f"{float(total_row['Quebra Geral']):.2%}",
        tema="vermelho" if float(total_row["Quebra Geral"]) > meta_pct else "verde",
    )
    render_kpi(k3, "Meta Global", f"{meta_pct:.0%}", tema="cinza")
    pior = max(Config.ORDEM_TIPOS, key=lambda t: float(total_row.get(t, 0)))
    render_kpi(k4, "Segmento Crítico", pior, f"Quebra: {float(total_row[pior]):.2%}", tema="laranja")

    st.markdown("<br>", unsafe_allow_html=True)
    render_section_header("grid_on", "Matriz de Desempenho (Monitor × Segmento)")

    def _cor_fundo(v):
        if isinstance(v, str):
            return ""
        try:
            return ("background-color:#FEE2E2;color:#991B1B;font-weight:bold;"
                    if float(v) > meta_pct
                    else "background-color:#D1FAE5;color:#065F46;font-weight:bold;")
        except Exception:
            return ""

    fmt_dict: dict[str, Any] = {c: "{:.1%}" for c in Config.ORDEM_TIPOS + ["Quebra Geral"]}
    fmt_dict["Total Tarefas"] = "{:,}"

    sty = df_matriz.style.format(cast(Any, fmt_dict), decimal=",", thousands=".")
    sty = sty.map(_cor_fundo, subset=cast(Any, Config.ORDEM_TIPOS + ["Quebra Geral"]))
    st.dataframe(sty, use_container_width=True, hide_index=True)

    st.markdown("<br>", unsafe_allow_html=True)
    render_section_header("bar_chart", "Distribuição Visual")
    df_plot = df_matriz[df_matriz["Monitor"] != "Total Geral"]

    fig = go.Figure()
    cores = {"Novos Domicílios": "#1E40AF", "Migração": "#0284C7", "PME": "#1E3A8A"}
    for tipo in Config.ORDEM_TIPOS:
        fig.add_trace(go.Bar(
            name=tipo, x=df_plot["Monitor"], y=df_plot[tipo],
            marker_color=cores.get(tipo, "#64748B"),
            text=[f"{v:.1%}" for v in df_plot[tipo]], textposition="outside",
        ))
    fig.add_hline(y=meta_pct, line_dash="dash", line_color="#DC2626",
                  annotation_text=f"META: {meta_pct:.0%}")
    fig.update_layout(barmode="group", height=400, yaxis_tickformat=".0%", margin=dict(t=20, b=20))
    st.plotly_chart(fig, use_container_width=True)


def view_segmento(df: pd.DataFrame, segmento: str, meta: float):
    df_seg = df[df["TIPO_SERVICO"] == segmento]
    if df_seg.empty:
        render_insight(f"Nenhum registro encontrado para **{segmento}**.", tipo="alerta")
        return

    m = Analise.metricas_segmento(df_seg)

    if m["quebra"] <= meta:
        render_insight(f"**{segmento}** está DENTRO DA META de {meta:.0%} (Atual: **{m['quebra']:.2%}**).", tipo="ok")
    else:
        render_insight(f"**{segmento}** está FORA DA META de {meta:.0%} (Atual: **{m['quebra']:.2%}**). Ação requerida!", tipo="critico")

    t1, t2, t3, t4 = st.tabs(["📊 Visão Geral", "🔍 Causa Raiz", "👤 Técnicos", "📋 Pendentes"])

    with t1:
        c1, c2, c3, c4 = st.columns(4)
        render_kpi(c1, "Alocado", f"{int(m['alocado']):,}", tema="azul")
        render_kpi(c2, "Executadas", f"{int(m['exec']):,}", tema="verde")
        render_kpi(c3, "Não Executadas", f"{int(m['naoexec']):,}", tema="laranja")
        render_kpi(c4, "Pendentes", f"{int(m['pend']):,}", tema="cinza")

    with t2:
        col_baixa = "_COL_BAIXA" if "_COL_BAIXA" in df_seg.columns else Utils.buscar_coluna(
            df_seg, ["MOTIVO DE BAIXA", "CÓD DE BAIXA 1"]
        )
        df_ne = df_seg[df_seg["Status Contrato"] == "Não Executada"]

        if not df_ne.empty and col_baixa:
            df_c = df_ne.groupby(col_baixa)["TOTAL DE TAREFAS"].sum().reset_index()
            df_c.columns = ["Motivo de Baixa", "Volume"]
            df_c = df_c.sort_values("Volume", ascending=False).head(10)
            df_c["% do Total"] = df_c["Volume"] / df_c["Volume"].sum()
            df_c["Acumulado"] = df_c["% do Total"].cumsum()
            render_dataframe(df_c, titulo="Pareto de Motivos de Baixa", icone="search",
                             fmt={"% do Total": "{:.1%}", "Acumulado": "{:.1%}"})
        else:
            render_insight("Nenhum motivo de baixa registrado.", tipo="info")

    with t3:
        df_tec = Analise.ranking_tecnicos(df_seg)
        render_dataframe(df_tec, titulo="Ranking de Quebra por Técnico", icone="person",
                         fmt={"Quebra Atual": "{:.1%}"})

    with t4:
        df_pend = df_seg[df_seg["Status Contrato"].str.upper().isin(
            ["PENDENTE", "PENDING", "ABERTO", "EM ABERTO"]
        )]
        cols = [c for c in ["CONTRATO", "TÉCNICO", "MONITOR", "REGIÃO"] if c in df_pend.columns]
        render_dataframe(df_pend[cols], titulo="Fila de Pendentes", icone="list_alt")


# ═══════════════════════════════════════════════════════
# 🆕 ABA: BASE COMPLETA (com Monitor do Merge)
# ═══════════════════════════════════════════════════════
def view_base_completa(df: pd.DataFrame, df_full: pd.DataFrame):
    if df.empty:
        render_insight("Nenhum registro disponível na base filtrada.", tipo="alerta")
        return

    # ── KPIs de Mapeamento ──
    render_section_header("hub", "Qualidade do Mapeamento (Merge)")

    total_reg = len(df)
    via_merge = int((df["ORIGEM_MAPEAMENTO"] == "Merge (Sheets)").sum()) if "ORIGEM_MAPEAMENTO" in df.columns else 0
    via_arquivo = int((df["ORIGEM_MAPEAMENTO"] == "Arquivo Origem").sum()) if "ORIGEM_MAPEAMENTO" in df.columns else 0
    sem_map = int((df["MONITOR"] == "SEM MONITOR").sum())
    pct_map = (total_reg - sem_map) / total_reg if total_reg else 0.0

    c1, c2, c3, c4 = st.columns(4)
    render_kpi(c1, "Total de Registros", f"{total_reg:,}".replace(",", "."), tema="azul")
    render_kpi(c2, "Via Merge (Sheets)", f"{via_merge:,}".replace(",", "."),
               f"{(via_merge/total_reg if total_reg else 0):.1%} do total", tema="verde")
    render_kpi(c3, "Via Arquivo Origem", f"{via_arquivo:,}".replace(",", "."),
               "Fallback aplicado", tema="laranja")
    render_kpi(c4, "Taxa de Mapeamento", f"{pct_map:.1%}",
               f"{sem_map:,} sem monitor".replace(",", "."),
               tema="verde" if pct_map >= 0.9 else "vermelho")

    if sem_map > 0:
        render_insight(
            f"Existem **{sem_map:,}** registros sem monitor vinculado. "
            "Verifique se os Logins destes técnicos constam na planilha **lista_ativos**.".replace(",", "."),
            tipo="alerta",
        )
    else:
        render_insight("100% dos registros possuem Monitor vinculado corretamente.", tipo="ok")

    # ── Controles de Visualização ──
    st.markdown("<br>", unsafe_allow_html=True)
    render_section_header("table_view", "Base de Dados Completa")

    with st.expander("⚙️ Configurar Visualização", expanded=False):
        cfg1, cfg2 = st.columns([1, 2])

        with cfg1:
            modo = st.radio(
                "Colunas exibidas",
                ["Essenciais", "Todas as colunas", "Personalizado"],
                key="base_modo_cols",
            )
            n_linhas = st.number_input(
                "Máx. de linhas exibidas", 100, 50000, 1000, 100, key="base_n_linhas"
            )

        with cfg2:
            cols_disponiveis = list(df.columns)
            if modo == "Essenciais":
                cols_sel = [c for c in Config.COLS_BASE_PRIORITARIAS if c in df.columns]
            elif modo == "Todas as colunas":
                cols_sel = cols_disponiveis
            else:
                default_cols = [c for c in Config.COLS_BASE_PRIORITARIAS if c in df.columns]
                cols_sel = st.multiselect(
                    "Selecione as colunas", cols_disponiveis, default=default_cols,
                    key="base_cols_custom",
                )

            busca = st.text_input(
                "🔎 Buscar (contrato, técnico, monitor...)", "", key="base_busca"
            )

    if not cols_sel:
        render_insight("Selecione ao menos uma coluna para exibir.", tipo="info")
        return

    df_view = df[cols_sel].copy()

    # ── Filtro de busca livre ──
    if busca.strip():
        termo = busca.strip().upper()
        mask = pd.Series(False, index=df_view.index)
        for c in df_view.columns:
            mask |= df_view[c].astype(str).str.upper().str.contains(termo, na=False, regex=False)
        df_view = df_view[mask]

    total_filtrado = len(df_view)
    df_view = df_view.head(int(n_linhas))

    st.caption(
        f"Exibindo **{len(df_view):,}** de **{total_filtrado:,}** registros "
        f"(base filtrada: {total_reg:,})".replace(",", ".")
    )

    # ── Estilização por Status ──
    def _cor_status(v):
        s = str(v).strip().upper()
        if s == "EXECUTADA":
            return "background-color:#ECFDF5;color:#065F46;font-weight:600;"
        if s in ("NÃO EXECUTADA", "NAO EXECUTADA"):
            return "background-color:#FEF2F2;color:#991B1B;font-weight:600;"
        if s == "PENDENTE":
            return "background-color:#FFFBEB;color:#92400E;font-weight:600;"
        return ""

    def _cor_monitor(v):
        if str(v).strip().upper() == "SEM MONITOR":
            return "background-color:#FEF2F2;color:#991B1B;font-weight:600;"
        return "background-color:#EFF6FF;color:#1E40AF;font-weight:600;"

    def _cor_origem(v):
        s = str(v).strip()
        if s == "Merge (Sheets)":
            return "background-color:#ECFDF5;color:#065F46;font-weight:600;"
        if s == "Arquivo Origem":
            return "background-color:#FFF7ED;color:#C2410C;font-weight:600;"
        return "background-color:#F3F4F6;color:#6B7280;"

    sty = df_view.style
    if "Status Contrato" in df_view.columns:
        sty = sty.map(_cor_status, subset=cast(Any, ["Status Contrato"]))
    if "MONITOR" in df_view.columns:
        sty = sty.map(_cor_monitor, subset=cast(Any, ["MONITOR"]))
    if "ORIGEM_MAPEAMENTO" in df_view.columns:
        sty = sty.map(_cor_origem, subset=cast(Any, ["ORIGEM_MAPEAMENTO"]))

    st.dataframe(sty, use_container_width=True, hide_index=True, height=600)

    # ── Exportações ──
    st.markdown("<br>", unsafe_allow_html=True)
    render_section_header("download", "Exportar Dados")

    e1, e2, e3 = st.columns(3)
    stamp = datetime.now().strftime("%Y%m%d_%H%M")

    with e1:
        st.download_button(
            "📥 Base Filtrada (Excel)",
            Utils.gerar_excel(df[cols_sel], "Base_Filtrada"),
            f"base_filtrada_{stamp}.xlsx",
            use_container_width=True,
        )
    with e2:
        st.download_button(
            "📦 Base Completa (Excel)",
            Utils.gerar_excel(df_full, "Base_Completa"),
            f"base_completa_{stamp}.xlsx",
            use_container_width=True,
        )
    with e3:
        sem_monitor = df[df["MONITOR"] == "SEM MONITOR"]
        st.download_button(
            "⚠️ Sem Monitor (Excel)",
            Utils.gerar_excel(
                sem_monitor if not sem_monitor.empty else pd.DataFrame({"Info": ["Nenhum registro"]}),
                "Sem_Monitor",
            ),
            f"sem_monitor_{stamp}.xlsx",
            use_container_width=True,
            disabled=sem_monitor.empty,
        )


# ═══════════════════════════════════════════════════════
# MAIN LOOP
# ═══════════════════════════════════════════════════════
def main():
    aplicar_estilo()

    # ── TELA DE IMPORTAÇÃO ──
    if "df_memoria" not in st.session_state or st.session_state["df_memoria"] is None:
        render_hero(
            titulo="Gestão de Quebra de Agenda",
            subtitulo="Importe a base operacional para gerar o Super Relatório Consolidado",
            badge="SISTEMA TOTALE",
        )
        render_section_header("folder_open", "Importação de Dados")
        arq = st.file_uploader("Selecione a base (Excel/CSV)", type=["xlsx", "csv"])

        if arq:
            with st.spinner("🔄 Lendo arquivo e vinculando monitores..."):
                raw = DataLoader.ler_arquivo(arq.getvalue(), arq.name)
                gs = DataLoader.buscar_gsheets()
                df_proc = DataLoader.preparar_base(raw, gs)
                st.session_state["df_memoria"] = df_proc
                st.rerun()
        return

    df_full = st.session_state["df_memoria"]

    # ── Sidebar ──
    with st.sidebar:
        st.header("🎯 Filtros Globais")

        mons = ["Todos"] + sorted(
            str(x) for x in df_full["MONITOR"].dropna().unique() if str(x).strip() != ""
        )
        sel_mon = st.selectbox("👔 Monitor", mons)
        df_filt = df_full if sel_mon == "Todos" else df_full[df_full["MONITOR"] == sel_mon]

        tecs = ["Todos"] + sorted(
            str(x) for x in df_filt["TÉCNICO"].dropna().unique() if str(x).strip() != ""
        )
        sel_tec = st.selectbox("👤 Técnico", tecs)
        df = df_filt if sel_tec == "Todos" else df_filt[df_filt["TÉCNICO"] == sel_tec]

        if Config.COL_REGIAO in df_full.columns:
            regs = ["Todas"] + sorted(str(x) for x in df_full[Config.COL_REGIAO].dropna().unique())
            sel_reg = st.selectbox("📍 Região", regs)
            if sel_reg != "Todas":
                df = df[df[Config.COL_REGIAO] == sel_reg]

        st.divider()
        if st.button("🔄 Trocar / Recarregar Base", use_container_width=True):
            st.session_state["df_memoria"] = None
            st.rerun()

    # ── Hero ──
    render_hero(
        "Super Relatório de Quebra Operacional",
        "Matriz Corporativa, Segmentos e Base Analítica Completa",
        "TOTALE",
    )

    d_imp = df_full.attrs.get("diag_importado", len(df_full))
    d_susp = df_full.attrs.get("diag_susp", 0)
    d_inv = df_full.attrs.get("diag_inv", 0)
    d_fin = df_full.attrs.get("diag_final", len(df_full))
    n_mon = df_full[df_full["MONITOR"] != "SEM MONITOR"]["MONITOR"].nunique()

    render_insight(
        f"**Base:** {d_imp:,} importados · {d_susp:,} suspensos e {d_inv:,} inválidos removidos · "
        f"**{d_fin:,} ativos** · **{n_mon} monitores mapeados**".replace(",", "."),
        tipo="info" if n_mon > 0 else "alerta",
    )

    # ── Abas ──
    tab1, tab2, tab3, tab4 = st.tabs([
        "🏢 Visão Executiva",
        "🔄 Migração",
        "🏢 PME",
        "📋 Base Completa",
    ])

    with tab1:
        view_executiva(df, Config.SLA_GERAL)
    with tab2:
        view_segmento(df, "Migração", Config.SLA_MIGRACAO)
    with tab3:
        view_segmento(df, "PME", Config.SLA_PME)
    with tab4:
        view_base_completa(df, df_full)


if __name__ == "__main__":
    main()
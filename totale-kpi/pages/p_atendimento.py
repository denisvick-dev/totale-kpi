"""
p_atendimento.py
================
App standalone para Primeiro Atendimento de OS.

Funcionalidades:
    - Upload central de Excel/CSV
    - Merge com Google Sheets (lista_ativos) — Técnico + Monitor + Base
    - Limpeza automática (contratos vazios + suspensos)
    - 🚙 PRIMEIRO HORÁRIO por técnico (coluna "Início" + status iniciado)
    - 🏆 Ranking de técnicos (mais madrugadores)
    - 👔 Ranking de monitores (média do primeiro atendimento)
    - KPIs executivos
    - Exportação Excel/CSV

Uso:
    streamlit run p_atendimento.py
"""

from __future__ import annotations

import unicodedata
from datetime import datetime, time
from io import BytesIO
from typing import Any, List, Optional, Tuple

import pandas as pd
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

# ═══════════════════════════════════════════════════════
# CONFIGURAÇÃO
# ═══════════════════════════════════════════════════════
st.set_page_config(
    page_title="Primeiro Atendimento | TOTALE",
    page_icon="🚙",
    layout="wide",
    initial_sidebar_state="expanded",
)


# ═══════════════════════════════════════════════════════
# CSS CORPORATIVO
# ═══════════════════════════════════════════════════════
def aplicar_estilo() -> None:
    st.markdown(
        """
    <link rel="preconnect" href="https://fonts.googleapis.com">
    <link href="https://fonts.googleapis.com/css2?family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet">

    <style>
    /* Tipografia e Fundo */
    html, body, [class*="css"], .stApp {
        font-family: 'Inter', -apple-system, sans-serif !important;
        background-color: #F8FAFC;
    }
    h1, h2, h3, h4 {
        font-family: 'Inter', sans-serif !important;
        letter-spacing: -0.02em;
    }
    .main .block-container {
        padding-top: 1.5rem;
        max-width: 1400px;
    }

    /* Sidebar Clean */
    [data-testid="stSidebar"] {
        background-color: #FFFFFF;
        border-right: 1px solid #E2E8F0;
    }
    [data-testid="stSidebar"] h3 {
        color: #0F172A !important;
        font-size: 14px !important;
        font-weight: 600 !important;
        text-transform: uppercase;
        letter-spacing: 0.05em;
    }

    /* ═════════ HERO TOTALE (Azul e Laranja) ═════════ */
    .hero-corp {
        background: linear-gradient(135deg, #012869 0%, #173b7d 50%, #F37C04 100%);
        padding: 32px 40px;
        border-radius: 8px;
        color: white;
        margin-bottom: 24px;
        border: none;
        box-shadow: 0 4px 15px rgba(1, 40, 105, 0.15);
    }
    .hero-title {
        font-size: 28px;
        font-weight: 700;
        margin: 0;
        color: #FFFFFF !important;
        letter-spacing: -0.02em;
    }
    .hero-subtitle {
        font-size: 14px;
        color: #F8FAFC;
        opacity: 0.9;
        margin: 8px 0 0 0;
        font-weight: 400;
    }
    .hero-badge {
        display: inline-block;
        background: rgba(255, 255, 255, 0.15);
        padding: 4px 12px;
        border-radius: 4px;
        font-size: 11px;
        font-weight: 600;
        margin-top: 16px;
        letter-spacing: 0.05em;
        text-transform: uppercase;
        color: #FFFFFF;
        border: 1px solid rgba(255, 255, 255, 0.3);
    }

    /* ═════════ KPI CARDS MINIMALISTAS ═════════ */
    .kpi-card {
        background: #FFFFFF;
        border-radius: 8px;
        padding: 20px;
        border: 1px solid #E2E8F0;
        box-shadow: 0 1px 2px rgba(0,0,0,0.02);
        height: 100%;
        position: relative;
    }
    .kpi-label {
        font-size: 12px;
        font-weight: 600;
        color: #64748B;
        text-transform: uppercase;
        letter-spacing: 0.02em;
        margin-bottom: 8px;
    }
    .kpi-value {
        font-size: 28px;
        font-weight: 700;
        color: #0F172A;
        line-height: 1;
        font-variant-numeric: tabular-nums;
    }
    .kpi-sub {
        font-size: 12px;
        color: #94A3B8;
        margin-top: 8px;
        font-weight: 500;
    }
    .kpi-icon {
        position: absolute;
        top: 20px;
        right: 20px;
        font-size: 20px;
        opacity: 0.8;
    }

    /* ═════════ UPLOAD ZONE ═════════ */
    .upload-card {
        background: #FFFFFF;
        border: 1px dashed #CBD5E1;
        border-radius: 8px;
        padding: 32px;
        text-align: center;
        margin-bottom: 16px;
        transition: all 0.2s;
    }
    .upload-icon {
        font-size: 32px;
        color: #0F172A;
        margin-bottom: 12px;
    }
    .upload-title {
        font-size: 16px;
        font-weight: 600;
        color: #0F172A;
        margin-bottom: 4px;
    }
    .upload-subtitle {
        font-size: 13px;
        color: #64748B;
    }

    /* ═════════ CABEÇALHOS DE SEÇÃO ═════════ */
    .section-header {
        display: flex;
        align-items: center;
        gap: 12px;
        margin: 40px 0 16px 0;
        padding-bottom: 8px;
        border-bottom: 1px solid #E2E8F0;
    }
    .section-title {
        font-size: 18px;
        font-weight: 600;
        color: #0F172A;
        margin: 0;
    }
    .section-badge {
        background: #F1F5F9;
        color: #475569;
        padding: 2px 8px;
        border-radius: 4px;
        font-size: 11px;
        font-weight: 600;
        text-transform: uppercase;
    }

    /* ═════════ DIAGNÓSTICO E INSIGHTS ═════════ */
    .diag-card {
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        border-radius: 8px;
        padding: 16px 20px;
    }
    .diag-item {
        display: flex;
        justify-content: space-between;
        align-items: center;
        padding: 8px 0;
        border-bottom: 1px dashed #E2E8F0;
        font-size: 13px;
    }
    .diag-item:last-child { border-bottom: none; }
    .diag-label { color: #475569; font-weight: 500; }
    .diag-value { font-weight: 600; color: #0F172A; }
    
    /* Alertas B2B */
    .insight-info { background: #EFF6FF; border: 1px solid #BFDBFE; color: #1E3A8A; padding: 12px 16px; border-radius: 6px; font-size: 13px; }
    .insight-alerta { background: #FFFBEB; border: 1px solid #FDE68A; color: #92400E; padding: 12px 16px; border-radius: 6px; font-size: 13px; }
    .insight-critico { background: #FEF2F2; border: 1px solid #FECACA; color: #991B1B; padding: 12px 16px; border-radius: 6px; font-size: 13px; }

    /* Barra Inferior (Resultado) */
    .resultado-base {
        background: #FFFFFF;
        border: 1px solid #E2E8F0;
        padding: 12px 16px;
        border-radius: 6px;
        display: flex;
        align-items: center;
        font-size: 13px;
    }
    </style>
    """,
        unsafe_allow_html=True,
    )


# ═══════════════════════════════════════════════════════
# CONSTANTES
# ═══════════════════════════════════════════════════════
CONTRATO_VALORES_VAZIOS = {
    "",
    "NAN",
    "NONE",
    "N/A",
    "NA",
    "-",
    "0",
    "NULL",
    "<NA>",
}


# ═══════════════════════════════════════════════════════
# COMPONENTES
# ═══════════════════════════════════════════════════════
def render_hero(titulo: str, subtitulo: str = "", badge: str = "") -> None:
    subtitulo_html = f'<p class="hero-subtitle">{subtitulo}</p>' if subtitulo else ""
    badge_html = f'<span class="hero-badge">{badge}</span>' if badge else ""
    st.markdown(
        f'<div class="hero-corp">'
        f'<h1 class="hero-title">{titulo}</h1>'
        f"{subtitulo_html}{badge_html}"
        f"</div>",
        unsafe_allow_html=True,
    )


def render_kpi(
    col,
    label: str,
    valor: str,
    sub: str = "",
    cor: str = "#012869",
    icone: str = "",
) -> None:
    icone_html = f'<div class="kpi-icon">{icone}</div>' if icone else ""
    col.markdown(
        f'<div class="kpi-card" style="border-left-color:{cor};">'
        f"{icone_html}"
        f'<div class="kpi-label">{label}</div>'
        f'<div class="kpi-value" style="color:{cor};">{valor}</div>'
        f'<div class="kpi-sub">{sub}</div>'
        f"</div>",
        unsafe_allow_html=True,
    )


def render_sidebar_kpi(label: str, valor: str, cor: str = "#012869") -> None:
    st.sidebar.markdown(
        f'<div class="sidebar-kpi" style="border-left-color:{cor};">'
        f'<div class="sidebar-kpi-label">{label}</div>'
        f'<div class="sidebar-kpi-value" style="color:{cor};">{valor}</div>'
        f"</div>",
        unsafe_allow_html=True,
    )


def render_section(icon: str, titulo: str, badge: str = "") -> None:
    badge_html = f'<span class="section-badge">{badge}</span>' if badge else ""
    st.markdown(
        f'<div class="section-header">'
        f'<span style="font-size:24px;">{icon}</span>'
        f'<h2 class="section-title">{titulo}</h2>'
        f"{badge_html}</div>",
        unsafe_allow_html=True,
    )


def render_insight(msg: str, tipo: str = "info") -> None:
    st.markdown(f'<div class="insight-{tipo}">{msg}</div>', unsafe_allow_html=True)


def render_upload_card() -> str:
    return (
        '<div class="upload-card">'
        '<div class="upload-icon">📊</div>'
        '<div class="upload-title">Faça upload da base de OSs</div>'
        '<div class="upload-subtitle">'
        "Formatos aceitos: <b>Excel (.xlsx, .xls)</b> ou <b>CSV</b> · "
        "A base será enriquecida com dados do <b>Google Sheets</b> (Técnico + Monitor + Base)"
        "</div>"
        "</div>"
    )


# ═══════════════════════════════════════════════════════
# UTILITÁRIOS
# ═══════════════════════════════════════════════════════
def normalizar_texto(texto: str) -> str:
    t = str(texto).strip().upper()
    t = unicodedata.normalize("NFKD", t).encode("ASCII", "ignore").decode()
    return t


def buscar_coluna(df: pd.DataFrame, palavras_chave: List[str]) -> Optional[str]:
    if df is None or df.empty:
        return None
    cols_norm = {normalizar_texto(c): c for c in df.columns}
    for palavra in palavras_chave:
        pn = normalizar_texto(palavra)
        for cn, co in cols_norm.items():
            if pn in cn:
                return co
    return None


def buscar_coluna_exata(df: pd.DataFrame, nome: str) -> Optional[str]:
    if df is None or df.empty:
        return None
    nome_norm = normalizar_texto(nome)
    for c in df.columns:
        if normalizar_texto(c) == nome_norm:
            return c
    return None


def gerar_excel(df: pd.DataFrame, aba: str = "Dados") -> bytes:
    """Exporta DataFrame para Excel com formatação corporativa e safe para NaN/float."""
    output = BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name=aba[:31])
        ws = writer.sheets[aba[:31]]

        # ─── Cabeçalho estilizado ───
        header_fill = PatternFill("solid", fgColor="012869")
        header_font = Font(color="FFFFFF", bold=True, size=11)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font

        # ─── Largura das colunas (safe para NaN/float) ───
        for i, col in enumerate(df.columns, 1):
            try:
                if df[col].empty:
                    max_len_dados = 0
                else:
                    # ✅ Converte TUDO para string primeiro, incluindo NaN
                    serie_str = df[col].fillna("").astype(str)
                    tamanhos = serie_str.str.len()
                    max_len_dados = int(tamanhos.max()) if len(tamanhos) > 0 else 0

                max_len = max(max_len_dados, len(str(col)))
                largura = min(max(max_len + 2, 12), 40)

                ws.column_dimensions[get_column_letter(i)].width = largura

            except Exception:
                # Se algo der errado nessa coluna, usa largura padrão
                ws.column_dimensions[get_column_letter(i)].width = 20

    return output.getvalue()


def fmt_int(v: int) -> str:
    return f"{v:,}".replace(",", ".")


def fmt_hora(t: time) -> str:
    """Formata datetime.time como HH:MM:SS."""
    if t is None:
        return "-"
    if isinstance(t, str):
        return t
    try:
        import math
        import numpy as np

        if t is getattr(pd, "NaT", None):
            return "-"
        if isinstance(t, float) and math.isnan(t):
            return "-"
        if isinstance(t, (np.floating,)) and np.isnan(t):
            return "-"
    except Exception:
        pass

    return t.strftime("%H:%M:%S")


# ═══════════════════════════════════════════════════════
# GOOGLE SHEETS — MERGE COM LISTA_ATIVOS
# ═══════════════════════════════════════════════════════
@st.cache_data(ttl=600, show_spinner="🔗 Conectando com Google Sheets...")
def carregar_lista_ativos() -> pd.DataFrame:
    """
    Lê a aba de técnicos usando a exportação nativa de CSV do Google Sheets.
    Retorna DataFrame com colunas: Login | Técnico | Monitor | Base
    """
    try:
        # 1. Define o ID padrão
        sheet_id = "1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg"

        # 2. Tenta ler do secrets de forma totalmente segura
        try:
            if "GSHEETS_ID" in st.secrets:
                sheet_id = st.secrets["GSHEETS_ID"]
        except Exception:
            # Se o arquivo secrets.toml não existir, ele ignora o erro silenciosamente
            pass

        # URL de exportação direta (mais rápido e sem erro 400)
        url_csv = (
            f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid=0"
        )

        df = pd.read_csv(url_csv)

        if df is None or df.empty:
            return pd.DataFrame()

        df.columns = df.columns.astype(str).str.strip()

        cols_esperadas = ["Login", "Técnico", "Monitor", "Base"]
        cols_encontradas = [c for c in cols_esperadas if c in df.columns]

        if "Login" not in cols_encontradas:
            st.warning(
                "⚠️ A coluna 'Login' não foi encontrada na planilha do Google Sheets."
            )
            return pd.DataFrame()

        df = df[cols_encontradas].copy()

        df["Login"] = (
            df["Login"]
            .astype(str)
            .str.replace(r"\.0$", "", regex=True)
            .str.strip()
            .str.upper()
        )

        df = df.drop_duplicates(subset=["Login"], keep="last")
        df = df[df["Login"].str.strip() != ""].reset_index(drop=True)

        return df

    except Exception as e:
        st.warning(f"⚠️ Não foi possível conectar com Google Sheets: {e}")
        return pd.DataFrame()


def aplicar_merge_ativos(
    df: pd.DataFrame,
    df_ativos: pd.DataFrame,
    col_login: Optional[str],
) -> Tuple[pd.DataFrame, dict]:
    """Faz merge da base carregada com a lista_ativos."""
    diag = {
        "aplicado": False,
        "matches": 0,
        "sem_match": 0,
        "total": len(df),
    }

    if df_ativos.empty or not col_login or col_login not in df.columns:
        return df, diag

    df = df.copy()

    df["_LOGIN_KEY"] = (
        df[col_login]
        .astype(str)
        .str.replace(r"\.0$", "", regex=True)
        .str.strip()
        .str.upper()
    )

    for c in ["Técnico", "Monitor", "Base"]:
        if c in df.columns:
            df = df.rename(columns={c: f"{c}_original"})

    df = df.merge(
        df_ativos,
        left_on="_LOGIN_KEY",
        right_on="Login",
        how="left",
        suffixes=("", "_gs"),
    )

    df = df.drop(columns=["_LOGIN_KEY", "Login"], errors="ignore")

    if "Técnico" in df.columns:
        matches = int(df["Técnico"].notna().sum())
        diag["matches"] = matches
        diag["sem_match"] = len(df) - matches
        diag["aplicado"] = True

    return df, diag


# ═══════════════════════════════════════════════════════
# CARREGAMENTO E LIMPEZA
# ═══════════════════════════════════════════════════════
@st.cache_data(show_spinner="🚗 Processando arquivo...")
def carregar_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
    bio = BytesIO(file_bytes)

    if filename.lower().endswith((".xlsx", ".xls")):
        return pd.read_excel(bio, engine="openpyxl", dtype=str)

    for sep in [";", ",", "\t", "|"]:
        for enc in ["utf-8-sig", "latin-1", "cp1252"]:
            try:
                bio.seek(0)
                df = pd.read_csv(
                    bio,
                    sep=sep,
                    encoding=enc,
                    dtype=str,
                    on_bad_lines="skip",
                    low_memory=False,
                )
                if len(df.columns) > 1:
                    return df
            except Exception:
                continue

    raise ValueError("Não foi possível ler o arquivo. Verifique o formato.")


def classificar_pendentes(df: pd.DataFrame, col_status: str) -> pd.Series:
    if col_status not in df.columns:
        return pd.Series([True] * len(df), index=df.index)

    s = df[col_status].fillna("").astype(str).str.upper()

    termos_pendentes = [
        "PENDENTE",
        "PENDING",
        "ABERTO",
        "EM ABERTO",
        "AGUARDAND",
        "AGENDAD",
        "NOVO",
        "NAO INICIAD",
        "NAO INICIA",
        "AGUARDA",
    ]
    termos_concluidos = [
        "EXECUTAD",
        "CONCLUID",
        "FINALIZAD",
        "REALIZAD",
        "ATENDID",
        "CANCELAD",
        "SUCESSO",
        "IMPRODUTIV",
    ]

    mask_pendente = s.apply(lambda v: any(t in v for t in termos_pendentes))
    mask_concluido = s.apply(lambda v: any(t in v for t in termos_concluidos))

    return mask_pendente | (~mask_concluido & (s == ""))


def limpar_base(
    df: pd.DataFrame,
    col_contrato: Optional[str],
    col_status_atividade: Optional[str],
) -> Tuple[pd.DataFrame, dict]:
    df_clean = df.copy()
    diag = {
        "total_inicial": len(df),
        "contratos_vazios": 0,
        "suspensos": 0,
        "total_final": 0,
    }

    if col_contrato and col_contrato in df_clean.columns:
        contrato_norm = (
            df_clean[col_contrato]
            .astype("string")
            .str.replace("\u00a0", " ", regex=False)
            .str.strip()
            .str.upper()
        )
        mask_vazio = (
            contrato_norm.isna()
            | contrato_norm.eq("")
            | contrato_norm.isin(CONTRATO_VALORES_VAZIOS)
        )
        diag["contratos_vazios"] = int(mask_vazio.sum())
        df_clean = df_clean.loc[~mask_vazio].copy()

    if col_status_atividade and col_status_atividade in df_clean.columns:
        status_norm = df_clean[col_status_atividade].fillna("").astype(str).str.upper()
        mask_suspenso = status_norm.str.contains(
            r"SUSPEN|BLOQUEAD|INATIV", regex=True, na=False
        )
        diag["suspensos"] = int(mask_suspenso.sum())
        df_clean = df_clean.loc[~mask_suspenso].copy()

    diag["total_final"] = len(df_clean)
    return df_clean.reset_index(drop=True), diag


# ═══════════════════════════════════════════════════════
# 🚗 PRIMEIRO HORÁRIO POR TÉCNICO
# ═══════════════════════════════════════════════════════
def calcular_primeiro_horario_tecnico(
    df: pd.DataFrame,
    col_tecnico: Optional[str],
    col_monitor: Optional[str],
    col_status_atividade: Optional[str],
    col_inicio: Optional[str],
) -> pd.DataFrame:
    if not col_tecnico or col_tecnico not in df.columns:
        return pd.DataFrame()
    if not col_status_atividade or col_status_atividade not in df.columns:
        return pd.DataFrame()
    if not col_inicio or col_inicio not in df.columns:
        return pd.DataFrame()

    status_norm = df[col_status_atividade].fillna("").astype(str).str.upper()
    termos_iniciado = [
        "INICIAD",
        "INICIA",
        "EM ANDAMENTO",
        "ANDAMENTO",
        "EXECUTANDO",
        "EM EXECUCAO",
        "EM EXECUÇÃO",
        "TRABALHANDO",
        "ATIVO",
    ]
    mask_iniciado = status_norm.apply(lambda v: any(t in v for t in termos_iniciado))

    df_iniciados = df.loc[mask_iniciado].copy()
    if df_iniciados.empty:
        return pd.DataFrame()

    df_iniciados["_INICIO"] = pd.to_datetime(
        df_iniciados[col_inicio],
        errors="coerce",
        dayfirst=True,
    )

    df_iniciados = df_iniciados.dropna(subset=["_INICIO"])
    if df_iniciados.empty:
        return pd.DataFrame()

    agg_dict = {
        "Primeiro_Inicio": ("_INICIO", "min"),
        "Status_Original": (col_status_atividade, "first"),
        "Qtd_Atividades": (col_status_atividade, "count"),
    }

    if col_monitor and col_monitor in df_iniciados.columns:
        agg_dict["Monitor"] = (col_monitor, "first")

    df_primeiro = (
        df_iniciados.sort_values(["_INICIO"])
        .groupby(col_tecnico)
        .agg(**agg_dict)
        .reset_index()
        .rename(columns={col_tecnico: "Técnico"})
    )

    df_primeiro["Data"] = df_primeiro["Primeiro_Inicio"].dt.strftime("%d/%m/%Y")
    df_primeiro["Hora Início"] = df_primeiro["Primeiro_Inicio"].dt.strftime("%H:%M:%S")

    df_primeiro = df_primeiro.sort_values("Primeiro_Inicio").reset_index(drop=True)
    df_primeiro["Rank"] = range(1, len(df_primeiro) + 1)

    cols_finais = ["Rank", "Técnico"]
    if "Monitor" in df_primeiro.columns:
        cols_finais.append("Monitor")
    cols_finais.extend(["Data", "Hora Início", "Status_Original", "Qtd_Atividades"])

    df_final = df_primeiro[cols_finais].rename(
        columns={
            "Status_Original": "Status",
            "Qtd_Atividades": "Qtd. Atividades",
        }
    )

    df_final["_datetime"] = df_primeiro["Primeiro_Inicio"].values

    return df_final


# ═══════════════════════════════════════════════════════
# 👔 RANKING DE MONITORES (média do primeiro atendimento)
# ═══════════════════════════════════════════════════════
def calcular_ranking_monitores(
    df_primeiro_horario: pd.DataFrame,
) -> pd.DataFrame:
    if df_primeiro_horario.empty:
        return pd.DataFrame()
    if "Monitor" not in df_primeiro_horario.columns:
        return pd.DataFrame()
    if "_datetime" not in df_primeiro_horario.columns:
        return pd.DataFrame()

    df_work = df_primeiro_horario.copy()

    df_work["_segundos"] = df_work["_datetime"].apply(
        lambda dt: (
            dt.hour * 3600 + dt.minute * 60 + dt.second if pd.notna(dt) else None
        )
    )

    df_work = df_work[
        df_work["Monitor"].notna() & (df_work["Monitor"].astype(str).str.strip() != "")
    ].copy()

    if df_work.empty:
        return pd.DataFrame()

    df_mon = (
        df_work.groupby("Monitor")
        .agg(
            Media_Segundos=("_segundos", "mean"),
            Primeiro=("_segundos", "min"),
            Ultimo=("_segundos", "max"),
            Qtd_Tecnicos=("Técnico", "nunique"),
            Total_Atividades=("Qtd. Atividades", "sum"),
        )
        .reset_index()
    )

    def segundos_para_hora(seg: float) -> str:
        if pd.isna(seg):
            return "-"
        seg = int(seg)
        h = seg // 3600
        m = (seg % 3600) // 60
        s = seg % 60
        return f"{h:02d}:{m:02d}:{s:02d}"

    df_mon["Média do Primeiro Atend."] = df_mon["Media_Segundos"].apply(
        segundos_para_hora
    )
    df_mon["Mais Cedo"] = df_mon["Primeiro"].apply(segundos_para_hora)
    df_mon["Mais Tarde"] = df_mon["Ultimo"].apply(segundos_para_hora)

    df_mon = df_mon.sort_values("Media_Segundos").reset_index(drop=True)
    df_mon["Rank"] = range(1, len(df_mon) + 1)

    return df_mon[
        [
            "Rank",
            "Monitor",
            "Média do Primeiro Atend.",
            "Mais Cedo",
            "Mais Tarde",
            "Qtd_Tecnicos",
            "Total_Atividades",
        ]
    ].rename(
        columns={
            "Qtd_Tecnicos": "Qtd. Técnicos",
            "Total_Atividades": "Total Atividades",
        }
    )


# ═══════════════════════════════════════════════════════
# BADGES E CORES
# ═══════════════════════════════════════════════════════
def rank_medal(rank: int) -> str:
    if rank == 1:
        return "🥇"
    if rank == 2:
        return "🥈"
    if rank == 3:
        return "🥉"
    return f"#{rank}"


def formatar_ranking_para_exibicao(
    df: pd.DataFrame, col_rank: str = "Rank"
) -> pd.DataFrame:
    if df.empty:
        return df
    df = df.copy()
    if "_datetime" in df.columns:
        df = df.drop(columns=["_datetime"])
    df[col_rank] = df[col_rank].apply(rank_medal)
    return df


def aplicar_cor_horario(valor: Any) -> str:
    """Aplica cores corporativas baseado no horário exato (HH:MM:SS)"""
    if pd.isna(valor) or valor == "-" or not isinstance(valor, str):
        return ""

    try:
        # Extrai a string e remove espaços
        horario = valor[:8].strip()

        # Proteção: se o horário vier faltando o zero à esquerda (ex: "8:30:00")
        if len(horario) == 7 and horario[1] == ":":
            horario = "0" + horario

        if horario <= "08:30:00":
            return "background-color: #D1FAE5; color: #065F46; font-weight: 600;"
        elif horario < "08:40:00":
            return "background-color: #FEF3C7; color: #92400E; font-weight: 600;"
        else:
            return "background-color: #FEE2E2; color: #991B1B; font-weight: 600;"
    except:
        return ""


# ═══════════════════════════════════════════════════════
# APP PRINCIPAL
# ═══════════════════════════════════════════════════════
def main() -> None:
    aplicar_estilo()

    render_hero(
        titulo="Painel de Primeiro Atendimento",
        subtitulo="Gestão Operacional de Ordens de Serviço · Merge automático via Google Sheets",
        badge="SISTEMA TOTALE",
    )

    with st.sidebar:
        st.markdown("### ⚙️ Painel de Controle")

        if st.button("🔄 Reiniciar Aplicação", use_container_width=True):
            for key in list(st.session_state.keys()):
                del st.session_state[key]
            st.cache_data.clear()
            st.rerun()

    # ─────────────────────────────────────────────────
    # UPLOAD
    # ─────────────────────────────────────────────────
    st.markdown(render_upload_card(), unsafe_allow_html=True)

    arq = st.file_uploader(
        "Selecione o arquivo",
        type=["xlsx", "xls", "csv"],
        label_visibility="collapsed",
    )

    if not arq:
        with st.sidebar:
            st.divider()
            st.caption("💡 Aguardando upload de arquivo...")

        st.markdown("### 📋 Como funciona")
        col1, col2, col3 = st.columns(3)

        with col1:
            st.markdown(
                '<div style="background:white;padding:20px;border-radius:10px;'
                "border-left:4px solid #012869;box-shadow:0 2px 8px rgba(0,0,0,0.06);"
                'min-height:170px;">'
                '<div style="font-size:32px;margin-bottom:8px;">🚙</div>'
                '<h4 style="margin:0 0 8px 0;color:#012869;">Upload da Base</h4>'
                '<p style="margin:0;color:#64748B;font-size:13px;">'
                "Envie o arquivo Excel/CSV com as OSs. O sistema fará o merge "
                "automático com a lista de técnicos ativos.</p>"
                "</div>",
                unsafe_allow_html=True,
            )

        with col2:
            st.markdown(
                '<div style="background:white;padding:20px;border-radius:10px;'
                "border-left:4px solid #F37C04;box-shadow:0 2px 8px rgba(0,0,0,0.06);"
                'min-height:170px;">'
                '<div style="font-size:32px;margin-bottom:8px;">🔗</div>'
                '<h4 style="margin:0 0 8px 0;color:#F37C04;">Enriquecimento</h4>'
                '<p style="margin:0;color:#64748B;font-size:13px;">'
                "Merge automático com <b>Google Sheets</b> traz Técnico, "
                "Monitor e Base pelo Login.</p>"
                "</div>",
                unsafe_allow_html=True,
            )

        with col3:
            st.markdown(
                '<div style="background:white;padding:20px;border-radius:10px;'
                "border-left:4px solid #059669;box-shadow:0 2px 8px rgba(0,0,0,0.06);"
                'min-height:170px;">'
                '<div style="font-size:32px;margin-bottom:8px;">🏆</div>'
                '<h4 style="margin:0 0 8px 0;color:#059669;">Rankings</h4>'
                '<p style="margin:0;color:#64748B;font-size:13px;">'
                "Ranking de SLAs operacionais e média de horário "
                "de saída por monitor.</p>"
                "</div>",
                unsafe_allow_html=True,
            )
        return

    # ─────────────────────────────────────────────────
    # PROCESSAMENTO
    # ─────────────────────────────────────────────────
    try:
        df_raw = carregar_arquivo(arq.getvalue(), arq.name)
    except Exception as e:
        render_insight(
            f"❌ <b>Erro ao ler arquivo:</b> <code>{e}</code>", tipo="critico"
        )
        return

    if df_raw.empty:
        render_insight("⚠️ <b>Arquivo vazio</b>.", tipo="alerta")
        return

    # Detecção de colunas
    col_contrato = buscar_coluna(
        df_raw, ["CONTRATO", "NUM CONTRATO", "NUMERO", "N CONTRATO"]
    )
    col_status = buscar_coluna(
        df_raw, ["STATUS DA O.S", "STATUS OS", "STATUS", "SITUACAO", "SITUAÇÃO"]
    )
    col_status_atividade = buscar_coluna(
        df_raw, ["STATUS DA ATIVIDADE", "STATUS ATIVIDADE"]
    )
    col_cliente = buscar_coluna(
        df_raw, ["CLIENTE", "NOME", "RAZAO SOCIAL", "RAZÃO SOCIAL"]
    )
    col_tipo = buscar_coluna(df_raw, ["TIPO", "SERVICO", "SERVIÇO"])
    col_regiao = buscar_coluna(df_raw, ["REGIAO", "REGIÃO", "CIDADE", "UF", "ESTADO"])
    col_login = buscar_coluna(
        df_raw,
        ["LOGIN DO TÉCNICO", "LOGIN DO TECNICO", "LOGIN", "USUÁRIO", "MATRÍCULA"],
    )

    col_inicio = buscar_coluna_exata(df_raw, "Início")
    if not col_inicio:
        col_inicio = buscar_coluna(df_raw, ["INICIO", "INÍCIO"])

    # Limpeza
    df, diag_limpeza = limpar_base(
        df_raw,
        col_contrato=col_contrato,
        col_status_atividade=col_status_atividade,
    )

    # Merge com Google Sheets
    df_ativos = carregar_lista_ativos()
    df, diag_merge = aplicar_merge_ativos(df, df_ativos, col_login)

    col_tecnico = (
        "Técnico"
        if "Técnico" in df.columns
        else buscar_coluna(
            df_raw, ["TECNICO", "TÉCNICO", "PROFISSIONAL", "NOME EQUIPE"]
        )
    )
    col_monitor = "Monitor" if "Monitor" in df.columns else None
    col_base_final = "Base" if "Base" in df.columns else col_regiao

    with st.sidebar:
        st.divider()
        st.markdown("### 📊 Base Carregada")
        st.caption(f"📄 **Arquivo:** `{arq.name}`")

    if df.empty:
        render_insight(
            "⚠️ Após limpeza, não restou nenhum registro válido.", tipo="critico"
        )
        return

    if col_status:
        mask_pend = classificar_pendentes(df, col_status)
        df_pend = df[mask_pend].copy().reset_index(drop=True)
    else:
        df_pend = df.copy()

    # ─────────────────────────────────────────────────
    # KPIs PRINCIPAIS
    # ─────────────────────────────────────────────────
    total = len(df)
    pendentes = len(df_pend)
    executados = total - pendentes
    pct_pend = (pendentes / total * 100) if total > 0 else 0

    render_section("📈", "Panorama Geral", "KPIs")

    c1, c2, c3, c4 = st.columns(4)
    render_kpi(
        c1, "Base Válida", fmt_int(total), sub="após limpeza", cor="#059669", icone="🚗"
    )
    render_kpi(
        c2,
        "Pendentes",
        fmt_int(pendentes),
        sub=f"{pct_pend:.1f}% do total",
        cor="#F37C04",
        icone="⏳",
    )
    render_kpi(
        c3,
        "Executadas/Fechadas",
        fmt_int(executados),
        sub=f"{100-pct_pend:.1f}% do total",
        cor="#012869",
        icone="✅",
    )

    if col_monitor and col_monitor in df.columns:
        monitores_qtd = df[col_monitor].dropna().nunique()
        render_kpi(
            c4,
            "Monitores",
            str(monitores_qtd),
            sub="na base",
            cor="#7C3AED",
            icone="👔",
        )
    elif col_base_final:
        regioes = df[col_base_final].dropna().nunique()
        render_kpi(
            c4,
            "Regiões/Bases",
            str(regioes),
            sub="com pendências",
            cor="#7C3AED",
            icone="🌎",
        )
    else:
        render_kpi(
            c4,
            "Colunas",
            str(len(df.columns)),
            sub="detectadas",
            cor="#7C3AED",
            icone="📋",
        )

    # ─────────────────────────────────────────────────
    # 🏆 RANKING DE TÉCNICOS (primeiro horário)
    # ─────────────────────────────────────────────────
    df_primeiro_horario = calcular_primeiro_horario_tecnico(
        df,
        col_tecnico=col_tecnico,
        col_monitor=col_monitor,
        col_status_atividade=col_status_atividade,
        col_inicio=col_inicio,
    )

    render_section(
        "🏆",
        "Ranking de Técnicos — Primeiro Horário de Saída",
        (
            f"{len(df_primeiro_horario)} técnicos"
            if not df_primeiro_horario.empty
            else "sem dados"
        ),
    )

    if df_primeiro_horario.empty:
        motivos = []
        if not col_tecnico:
            motivos.append("coluna de <b>Técnico</b> não detectada")
        if not col_status_atividade:
            motivos.append("coluna de <b>Status da Atividade</b> não detectada")
        if not col_inicio:
            motivos.append(
                "coluna <b>Início</b> não encontrada (nome esperado: <code>Início</code>)"
            )
        if not motivos:
            motivos.append(
                "nenhum registro com status <code>INICIADO</code>, <code>EM ANDAMENTO</code> ou similar"
            )

        render_insight(
            "⚠️ <b>Não foi possível calcular o ranking:</b><br>"
            + "<br>".join(f"• {m}" for m in motivos),
            tipo="alerta",
        )
    else:
        kh1, kh2, kh3 = st.columns(3)

        tec_iniciaram = len(df_primeiro_horario)
        primeiro = df_primeiro_horario.iloc[0]
        ultimo = df_primeiro_horario.iloc[-1]

        render_kpi(
            kh1,
            "Primeiro Início (Técnico)",
            primeiro["Hora Início"],
            sub=str(primeiro["Técnico"])[:30],
            cor="#0284C7",
            icone="⏱️",
        )
        render_kpi(
            kh2,
            "Último Início (Técnico)",
            ultimo["Hora Início"],
            sub=str(ultimo["Técnico"])[:30],
            cor="#475569",
            icone="📉",
        )
        render_kpi(
            kh3,
            "Volume em Rota",
            str(tec_iniciaram),
            sub="técnicos com status iniciado",
            cor="#059669",
            icone="📈",
        )

        st.markdown("")

        df_exibir_tec = formatar_ranking_para_exibicao(df_primeiro_horario)
        df_exibir_tec = df_exibir_tec.rename(
            columns={"Hora Início": "Primeiro Horário"}
        )

        colunas_desejadas = [
            "Rank",
            "Técnico",
            "Monitor",
            "Data",
            "Primeiro Horário",
            "Status",
        ]
        colunas_exibir = [c for c in colunas_desejadas if c in df_exibir_tec.columns]

        df_exibir_tec = df_exibir_tec[colunas_exibir]

        styled_tec = df_exibir_tec.style
        style_func = getattr(styled_tec, "map", None) or getattr(
            styled_tec, "applymap", None
        )
        if style_func is not None:
            styled_tec = style_func(aplicar_cor_horario, subset=["Primeiro Horário"])

        st.dataframe(
            styled_tec,
            use_container_width=True,
            hide_index=True,
            height=400,
            column_config={
                "Rank": st.column_config.TextColumn("Rank", width="small"),
                "Técnico": st.column_config.TextColumn("Técnico", width="medium"),
                "Monitor": (
                    st.column_config.TextColumn("Monitor", width="medium")
                    if "Monitor" in colunas_exibir
                    else None
                ),
                "Data": st.column_config.TextColumn("Data", width="small"),
                "Status": st.column_config.TextColumn("Status", width="medium"),
            },
        )

        col_dl1, col_dl2, _ = st.columns([1, 1, 2])
        timestamp = datetime.now().strftime("%Y%m%d_%H%M")

        with col_dl1:
            st.download_button(
                "📥 **Baixar Excel**",
                data=gerar_excel(df_exibir_tec, "Ranking_Tecnicos"),
                file_name=f"ranking_tecnicos_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )
        with col_dl2:
            csv_t = df_exibir_tec.to_csv(
                index=False, sep=";", encoding="utf-8-sig"
            ).encode("utf-8-sig")
            st.download_button(
                "📄 **Baixar CSV**",
                data=csv_t,
                file_name=f"ranking_tecnicos_{timestamp}.csv",
                mime="text/csv",
                use_container_width=True,
            )

        # ─────────────────────────────────────────────────
        # 👔 RANKING DE MONITORES
        # ─────────────────────────────────────────────────
        df_ranking_monitores = calcular_ranking_monitores(df_primeiro_horario)

        render_section(
            "👔",
            "Ranking de Monitores — Média de Primeiro Atendimento",
            (
                f"{len(df_ranking_monitores)} monitores"
                if not df_ranking_monitores.empty
                else "sem dados"
            ),
        )

        if df_ranking_monitores.empty:
            render_insight(
                "⚠️ Não foi possível calcular o ranking de monitores. Verifique a coluna <b>Monitor</b>.",
                tipo="alerta",
            )
        else:
            km1, km2, km3 = st.columns(3)
            monitor_madrugador = df_ranking_monitores.iloc[0]

            render_kpi(
                km1,
                "Melhor SLA (Monitor)",
                monitor_madrugador["Média do Primeiro Atend."],
                sub=str(monitor_madrugador["Monitor"])[:30],
                cor="#0284C7",
                icone="👤",
            )
            render_kpi(
                km2,
                "Cobertura",
                f"{len(df_ranking_monitores)} monitores",
                sub="com equipes ativas",
                cor="#475569",
                icone="📊",
            )
            render_kpi(
                km3,
                "Atividades Processadas",
                fmt_int(int(df_ranking_monitores["Total Atividades"].sum())),
                sub="O.S. iniciadas no total",
                cor="#059669",
                icone="📋",
            )

            st.markdown("")

            df_exibir_mon = formatar_ranking_para_exibicao(df_ranking_monitores)

            styled_mon = df_exibir_mon.style
            style_func_mon = getattr(styled_mon, "map", None) or getattr(
                styled_mon, "applymap", None
            )
            if style_func_mon is not None:
                styled_mon = style_func_mon(
                    aplicar_cor_horario, subset=["Média do Primeiro Atend."]
                )

            st.dataframe(
                styled_mon,
                use_container_width=True,
                hide_index=True,
                height="auto",
                column_config={
                    "Rank": st.column_config.TextColumn("Rank", width="small"),
                    "Monitor": st.column_config.TextColumn("Monitor", width="medium"),
                    "Média do Primeiro Atend.": st.column_config.TextColumn(
                        "Média de Saída",
                        width="small",
                        help="Média do horário do primeiro atendimento dos técnicos",
                    ),
                    "Mais Cedo": st.column_config.TextColumn(
                        "Mais Cedo", width="small"
                    ),
                    "Mais Tarde": st.column_config.TextColumn(
                        "Mais Tarde", width="small"
                    ),
                    "Qtd. Técnicos": st.column_config.NumberColumn(
                        "Técnicos", width="small", format="%d"
                    ),
                    "Total Atividades": st.column_config.NumberColumn(
                        "Atividades", width="small", format="%d"
                    ),
                },
            )

            col_dm1, col_dm2, _ = st.columns([1, 1, 2])

            with col_dm1:
                st.download_button(
                    "📥 **Baixar Excel**",
                    data=gerar_excel(df_exibir_mon, "Ranking_Monitores"),
                    file_name=f"ranking_monitores_{timestamp}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True,
                )
            with col_dm2:
                csv_m = df_exibir_mon.to_csv(
                    index=False, sep=";", encoding="utf-8-sig"
                ).encode("utf-8-sig")
                st.download_button(
                    "📄 **Baixar CSV**",
                    data=csv_m,
                    file_name=f"ranking_monitores_{timestamp}.csv",
                    mime="text/csv",
                    use_container_width=True,
                )

    # ─────────────────────────────────────────────────
    # FILTROS + TABELA DE PENDENTES
    # ─────────────────────────────────────────────────
    render_section(
        "🚗", "Contratos Pendentes para Rota", f"{fmt_int(pendentes)} pendentes"
    )

    df_filtrado = df_pend.copy()

    col_f1, col_f2, col_f3 = st.columns(3)

    with col_f1:
        if col_base_final and col_base_final in df_filtrado.columns:
            # Dropna e Strip evitam categorias vazias/duplicadas
            regioes = ["Todas"] + sorted(
                [
                    str(r).strip()
                    for r in df_filtrado[col_base_final].dropna().unique()
                    if str(r).strip()
                ]
            )
            sel_regiao = st.selectbox("🌎 Base/Região", regioes)
            if sel_regiao != "Todas":
                df_filtrado = df_filtrado[
                    df_filtrado[col_base_final].astype(str).str.strip() == sel_regiao
                ]

    with col_f2:
        if col_monitor and col_monitor in df_filtrado.columns:
            monitores = ["Todos"] + sorted(
                [
                    str(m).strip()
                    for m in df_filtrado[col_monitor].dropna().unique()
                    if str(m).strip()
                ]
            )
            sel_monitor = st.selectbox("👔 Monitor", monitores)
            if sel_monitor != "Todos":
                df_filtrado = df_filtrado[
                    df_filtrado[col_monitor].astype(str).str.strip() == sel_monitor
                ]

    with col_f3:
        if col_tecnico and col_tecnico in df_filtrado.columns:
            tecnicos = ["Todos"] + sorted(
                [
                    str(t).strip()
                    for t in df_filtrado[col_tecnico].dropna().unique()
                    if str(t).strip()
                ]
            )
            sel_tecnico = st.selectbox("👤 Técnico", tecnicos)
            if sel_tecnico != "Todos":
                df_filtrado = df_filtrado[
                    df_filtrado[col_tecnico].astype(str).str.strip() == sel_tecnico
                ]

    busca = st.text_input(
        "🔎 Busca livre",
        placeholder="Digite para pesquisar em toda a base...",
    )

    if busca:
        # Busca Vetorizada de Alta Performance
        mask_busca = pd.Series(False, index=df_filtrado.index)
        for col in df_filtrado.columns:
            mask_busca |= (
                df_filtrado[col].astype(str).str.contains(busca, case=False, na=False)
            )
        df_filtrado = df_filtrado[mask_busca]

    # Resultado
    pct_apos = (len(df_filtrado) / pendentes * 100) if pendentes > 0 else 0
    st.markdown(
        f'<div class="resultado-base">'
        f'<span class="resultado-label">🚗 Resultado:</span>'
        f'<span class="resultado-badge" style="background:#DBEAFE;color:#1E40AF;border-color:#3B82F6;">'
        f"{fmt_int(len(df_filtrado))} pendentes</span>"
        f'<span class="resultado-count">{pct_apos:.0f}% dos {fmt_int(pendentes)} originais</span>'
        f"</div>",
        unsafe_allow_html=True,
    )

    if df_filtrado.empty:
        render_insight("🔍 Nenhum registro encontrado com os filtros.", tipo="alerta")
        return

    # Tabela
    colunas_prio = [
        c
        for c in [
            col_contrato,
            col_cliente,
            col_tecnico,
            col_monitor,
            col_base_final,
            col_tipo,
            col_status,
            col_inicio,
        ]
        if c is not None and c in df_filtrado.columns
    ]
    outras = [c for c in df_filtrado.columns if c not in colunas_prio]
    df_exibir = df_filtrado[colunas_prio + outras].copy()

    # Spinner caso a tabela seja muito grande
    with st.spinner("Montando tabela de pendentes..."):
        st.dataframe(df_exibir, use_container_width=True, hide_index=True, height=500)

    # Exportação
    render_section("📥", "Exportar Lista para Rota")

    col_exp1, col_exp2, col_info = st.columns([1, 1, 2])
    timestamp = datetime.now().strftime("%Y%m%d_%H%M")

    with col_exp1:
        st.download_button(
            "🚗 **Baixar Excel**",
            data=gerar_excel(df_exibir, "Pendentes_Rota"),
            file_name=f"rota_pendentes_{timestamp}.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            use_container_width=True,
            type="primary",
        )

    with col_exp2:
        csv = df_exibir.to_csv(index=False, sep=";", encoding="utf-8-sig").encode(
            "utf-8-sig"
        )
        st.download_button(
            "📄 **Baixar CSV**",
            data=csv,
            file_name=f"rota_pendentes_{timestamp}.csv",
            mime="text/csv",
            use_container_width=True,
        )

    with col_info:
        st.caption(
            f"📅 **{datetime.now().strftime('%d/%m/%Y às %H:%M')}** · "
            f"🚗 **{fmt_int(len(df_exibir))}** registros para rota"
        )

    # Rodapé
    st.markdown(
        '<div style="text-align:center;color:#94A3B8;font-size:11px;'
        'padding:24px 0 8px 0;margin-top:32px;border-top:1px solid #F1F5F9;">'
        '🚗 <b style="color:#012869;">Primeiro Atendimento</b> · '
        "Ferramenta operacional TOTALE"
        "</div>",
        unsafe_allow_html=True,
    )


if __name__ == "__main__":
    main()

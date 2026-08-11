from __future__ import annotations

import html
import textwrap
import re
import unicodedata
from io import BytesIO
from typing import Any, Dict, List, Literal, Optional

import numpy as np
import pandas as pd
import plotly.express as px
import plotly.graph_objects as go
import streamlit as st
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter
from streamlit_gsheets import GSheetsConnection

# ==========================================================
# CONFIGURAÇÃO E CONSTANTES
# ==========================================================
st.set_page_config(
    page_title="Volumetria",
    page_icon="📊",
    layout="wide",
    initial_sidebar_state="expanded",
)


class Config:
    URL_GSHEETS = "https://docs.google.com/spreadsheets/d/1LQKDcLshC6XSXLBVWaEYSpxrro6uydyU9pwDLc38pEg/edit"
    META_EXECUCAO = 0.80
    META_EXECUTADAS_TECNICO = 7

    COL_STATUS = "STATUS CONTRATO"
    COL_TOTAL = "TOTAL DE TAREFAS"
    COL_TECNICO = "TÉCNICO"
    COL_MONITOR = "MONITOR"
    COL_REGIAO = "REGIÃO"

    CORES_STATUS = {
        "Executada": "#10B981",
        "Não Executada": "#EF4444",
        "Pendente": "#F59E0B",
    }

    STATUS_ORDEM = ["Executada", "Não Executada", "Pendente"]
    CONTRATO_VALORES_VAZIOS = {"", "NAN", "NONE", "N/A", "NA", "-", "0", "NULL"}
    REGIOES_PRINCIPAIS = ["LESTE", "GRU", "ABCDM"]


# ==========================================================
# TEMAS DE CARDS
# ==========================================================
TEMAS_CARD: Dict[str, Dict[str, str]] = {
    "amarelo": {"fundo": "#FEF9C3", "texto": "#854D0E", "borda": "#EAB308", "titulo": "#A16207"},
    "azul":    {"fundo": "#F0F9FF", "texto": "#0369A1", "borda": "#0EA5E9", "titulo": "#075985"},
    "verde":   {"fundo": "#F0FDF4", "texto": "#15803D", "borda": "#22C55E", "titulo": "#166534"},
    "roxo":    {"fundo": "#FAF5FF", "texto": "#7E22CE", "borda": "#A855F7", "titulo": "#6B21A8"},
    "cinza":   {"fundo": "#F8FAFC", "texto": "#334155", "borda": "#94A3B8", "titulo": "#64748B"},
    "escuro":  {"fundo": "#1E293B", "texto": "#FFFFFF", "borda": "#475569", "titulo": "#E2E8F0"},
    "vermelho":{"fundo": "#FEF2F2", "texto": "#B91C1C", "borda": "#EF4444", "titulo": "#991B1B"},
    "laranja": {"fundo": "#FFF7ED", "texto": "#C2410C", "borda": "#F97316", "titulo": "#9A3412"},
}

CORES_REGIAO: Dict[str, Dict[str, str]] = {
    "LESTE":  {"bg": "#DBEAFE", "text": "#1E40AF", "border": "#3B82F6"},
    "GRU":    {"bg": "#D1FAE5", "text": "#065F46", "border": "#10B981"},
    "ABCDM":  {"bg": "#EDE9FE", "text": "#5B21B6", "border": "#8B5CF6"},
    "OUTRAS": {"bg": "#F1F5F9", "text": "#475569", "border": "#94A3B8"},
}

RENOMEAR_COLUNAS: Dict[str, str] = {
    "TÉCNICO": "Técnico", "MONITOR": "Monitor", "REGIÃO": "Região",
    "STATUS CONTRATO": "Status", "TOTAL DE TAREFAS": "Total de O.S.",
    "Executada": "Executadas", "Não Executada": "Não Exec.",
    "Pendente": "Pendentes", "Baixadas": "Baixadas",
    "Total Alocado": "Total Alocado", "Taxa Execução": "Taxa Exec.",
    "Taxa Quebra": "Taxa Quebra", "Projeção": "Projeção",
    "Nao_Executadas": "Não Exec.", "Executadas": "Executadas",
    "Pendentes": "Pendentes", "Total_Alocado": "Total Alocado",
    "Técnicos": "Técnicos", "OS/Técnico": "OS/Técnico",
    "Exec/Técnico": "Exec/Técnico",
}

COLUNAS_INTEIRAS = [
    "Executada", "Não Executada", "Pendente", "Baixadas",
    "Total Alocado", "Projeção", "Executadas", "Não Exec.",
    "Pendentes", "Técnicos",
]


# ==========================================================
# UTILITÁRIOS
# ==========================================================
class Utils:
    @staticmethod
    def remover_acentos(valor) -> str:
        if pd.isna(valor):
            return ""
        return unicodedata.normalize("NFKD", str(valor)).encode("ASCII", "ignore").decode("ASCII")

    @staticmethod
    def normalizar_chave(serie: pd.Series) -> pd.Series:
        """Normaliza texto para comparações, preservando valores nulos como vazio."""
        s = serie.copy()
        return (
            s.where(s.notna(), "")
            .astype(str)
            .str.strip()
            .str.upper()
            .apply(Utils.remover_acentos)
        )

    @staticmethod
    def normalizar_login(serie: pd.Series) -> pd.Series:
        """Normaliza logins e remove o sufixo .0 comum em dados vindos do Excel."""
        return Utils.normalizar_chave(serie).str.replace(r"\.0$", "", regex=True)

    @staticmethod
    def buscar_coluna(df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
        cols_map = {Utils.normalizar_chave(pd.Series([c]))[0]: c for c in df.columns}
        for alias in aliases:
            chave = Utils.normalizar_chave(pd.Series([alias]))[0]
            if chave in cols_map:
                return cols_map[chave]
        return None

    @staticmethod
    def classificar_status(status_os: pd.Series) -> pd.Series:
        """Padroniza os status da O.S. em três categorias do painel."""
        s = Utils.normalizar_chave(status_os)
        nao_exec = s.str.contains(r"NAO\s*EXECUT", regex=True, na=False)
        exec_ = s.str.contains(r"EXECUT", regex=True, na=False) & ~nao_exec
        return pd.Series(
            np.select(
                [exec_, nao_exec],
                ["Executada", "Não Executada"],
                default="Pendente",
            ),
            index=status_os.index,
            dtype="object",
        )

    @staticmethod
    def contrato_valido(serie: pd.Series) -> pd.Series:
        norm = serie.astype(str).str.strip().str.upper().apply(Utils.remover_acentos)
        return ~norm.isin(Config.CONTRATO_VALORES_VAZIOS)

    @staticmethod
    def resolver_renomeacao(df: pd.DataFrame, mapa: Dict[str, str]) -> Dict[str, str]:
        existentes = set(df.columns)
        usados: set[str] = set()
        resultado: Dict[str, str] = {}
        for col in df.columns:
            if col not in mapa:
                continue
            novo = mapa[col]
            if novo == col or novo in existentes or novo in usados:
                continue
            resultado[col] = novo
            usados.add(novo)
        return resultado

    @staticmethod
    def calcular_metricas_grupo(agrupado: pd.DataFrame) -> pd.DataFrame:
        ag = agrupado.copy()
        for c in ["Executadas", "Nao_Executadas", "Pendentes", "Total_Alocado"]:
            if c in ag.columns:
                ag[c] = pd.to_numeric(ag[c], errors="coerce").fillna(0).astype(int)
        ag["Baixadas"] = ag["Executadas"] + ag["Nao_Executadas"]
        ag["Taxa Execução"] = np.where(ag["Baixadas"] > 0, ag["Executadas"] / ag["Baixadas"], 0.0)
        ag["Taxa Quebra"] = np.where(ag["Baixadas"] > 0, ag["Nao_Executadas"] / ag["Baixadas"], 0.0)
        ag["Projeção"] = (ag["Total_Alocado"] * ag["Taxa Execução"]).round(0).astype(int)
        return ag


# ==========================================================
# CARREGAMENTO
# ==========================================================
class DataLoader:
    @staticmethod
    @st.cache_data(show_spinner=False)
    def ler_arquivo(file_bytes: bytes, filename: str) -> pd.DataFrame:
        """Lê Excel ou CSV e devolve uma base independente do formato de origem."""
        if not file_bytes:
            raise ValueError("O arquivo enviado está vazio.")

        nome = filename.lower()
        if nome.endswith(".xlsx"):
            return pd.read_excel(BytesIO(file_bytes), engine="openpyxl")
        if nome.endswith(".xls"):
            # O pandas seleciona o engine apropriado (xlrd) para .xls.
            return pd.read_excel(BytesIO(file_bytes))
        if nome.endswith(".csv"):
            try:
                return pd.read_csv(BytesIO(file_bytes), sep=None, engine="python", encoding="utf-8-sig")
            except UnicodeDecodeError:
                return pd.read_csv(BytesIO(file_bytes), sep=None, engine="python", encoding="latin1")
        raise ValueError("Formato não suportado. Use .xlsx, .xls ou .csv.")

    @staticmethod
    @st.cache_data(ttl=600, show_spinner=False)
    def buscar_hierarquia_gsheets() -> pd.DataFrame:
        try:
            conn = st.connection("gsheets", type=GSheetsConnection)
            raw = conn.read(spreadsheet=Config.URL_GSHEETS)
            if raw is None or raw.empty:
                return pd.DataFrame()
            col_login = Utils.buscar_coluna(raw, ["LOGIN", "MATRÍCULA", "ID"])
            col_tec = Utils.buscar_coluna(raw, ["TÉCNICO", "NOME"])
            col_mon = Utils.buscar_coluna(raw, ["MONITOR", "GESTOR"])
            if not col_login:
                return pd.DataFrame()
            df_gs = pd.DataFrame({
                "__LOGIN_KEY": Utils.normalizar_login(raw[col_login]),
                "__TEC_GS": raw[col_tec] if col_tec else "",
                "__MON_GS": raw[col_mon] if col_mon else "",
            })
            return df_gs.drop_duplicates("__LOGIN_KEY")
        except Exception:
            return pd.DataFrame()

    @staticmethod
    def preparar_base(df: pd.DataFrame, df_gs: pd.DataFrame) -> pd.DataFrame:
        df = df.copy()
        diag = {"Inicial": len(df)}

        col_con = Utils.buscar_coluna(df, ["CONTRATO", "Nº CONTRATO", "NUMERO CONTRATO", "NUM CONTRATO"])
        if col_con:
            valida = Utils.contrato_valido(df[col_con])
            rem = (~valida).sum()
            df = df[valida].copy()
            diag["Removidos por contrato vazio"] = int(rem)
            if rem > 0:
                st.toast(f"🗑️ {rem} linha(s) removida(s) por contrato vazio.", icon="⚠️")
        else:
            st.warning("⚠️ Coluna de contrato não encontrada.")

        col_atv = Utils.buscar_coluna(df, ["STATUS DA ATIVIDADE"])
        if col_atv:
            susp = Utils.normalizar_chave(df[col_atv]).str.contains("SUSP", na=False)
            df = df[~susp].copy()

        col_os1 = Utils.buscar_coluna(df, ["STATUS DA O.S 1", "STATUS OS 1"])
        if not col_os1:
            st.error("Coluna 'Status da O.S 1' não encontrada!")
            st.stop()
        df[Config.COL_STATUS] = Utils.classificar_status(df[col_os1])

        col_qtd = Utils.buscar_coluna(df, ["TOTAL DE TAREFAS", "QUANTIDADE"])
        if col_qtd:
            df[Config.COL_TOTAL] = pd.to_numeric(df[col_qtd], errors="coerce").fillna(1)
        else:
            df[Config.COL_TOTAL] = 1
        df[Config.COL_TOTAL] = df[Config.COL_TOTAL].clip(lower=0)

        col_log = Utils.buscar_coluna(df, ["LOGIN DO TÉCNICO", "LOGIN", "USUÁRIO"])
        if col_log and not df_gs.empty:
            df["__LOGIN_KEY"] = Utils.normalizar_login(df[col_log])
            df = df.merge(df_gs, on="__LOGIN_KEY", how="left")
        for c in ("__TEC_GS", "__MON_GS"):
            if c not in df.columns:
                df[c] = np.nan

        col_tec_b = Utils.buscar_coluna(df, ["TÉCNICO", "NOME"]) or col_log
        col_mon_b = Utils.buscar_coluna(df, ["MONITOR", "GESTOR"])

        base_tec = df[col_tec_b] if col_tec_b else pd.Series(np.nan, index=df.index)
        base_mon = df[col_mon_b] if col_mon_b else pd.Series(np.nan, index=df.index)

        tec_gs = df["__TEC_GS"].where(df["__TEC_GS"].notna(), "")
        mon_gs = df["__MON_GS"].where(df["__MON_GS"].notna(), "")
        df[Config.COL_TECNICO] = tec_gs.mask(tec_gs.astype(str).str.strip().eq(""), base_tec)
        df[Config.COL_MONITOR] = mon_gs.mask(mon_gs.astype(str).str.strip().eq(""), base_mon)
        df[Config.COL_TECNICO] = df[Config.COL_TECNICO].where(
            df[Config.COL_TECNICO].notna() & df[Config.COL_TECNICO].astype(str).str.strip().ne(""),
            "NÃO MAPEADO",
        )
        df[Config.COL_MONITOR] = df[Config.COL_MONITOR].where(
            df[Config.COL_MONITOR].notna() & df[Config.COL_MONITOR].astype(str).str.strip().ne(""),
            "SEM MONITOR",
        )

        col_cid = Utils.buscar_coluna(df, ["CIDADE", "LOCALIDADE"])
        cidade = Utils.normalizar_chave(df[col_cid]) if col_cid else pd.Series("", index=df.index)
        df[Config.COL_REGIAO] = np.select(
            [
                cidade.isin(["SAO PAULO"]),
                cidade.isin(["GUARULHOS", "ARUJA", "MOGI DAS CRUZES", "SUZANO", "ITAQUAQUECETUBA", "FERRAZ DE VASCONCELOS", "POA"]),
                cidade.isin(["SANTO ANDRE", "SAO BERNARDO DO CAMPO", "SAO CAETANO DO SUL", "DIADEMA", "MAUA", "RIBEIRAO PIRES", "RIO GRANDE DA SERRA"]),
            ],
            ["LESTE", "GRU", "ABCDM"],
            default="OUTRAS",
        )

        diag["Final"] = len(df)
        df.attrs["diagnostico"] = diag
        return df


# ==========================================================
# CÁLCULOS
# ==========================================================
def calcular_kpis(df: pd.DataFrame) -> Dict[str, Any]:
    """Calcula os KPIs gerais da base filtrada."""
    if df.empty:
        return {
            "total": 0, "executadas": 0, "nao_executadas": 0,
            "pendentes": 0, "baixadas": 0, "taxa": 0.0,
            "quebra": 0.0, "projecao": 0,
        }
    k_tot = int(round(df[Config.COL_TOTAL].sum()))
    k_exe = int(df[df[Config.COL_STATUS] == "Executada"][Config.COL_TOTAL].sum())
    k_nex = int(df[df[Config.COL_STATUS] == "Não Executada"][Config.COL_TOTAL].sum())
    k_pen = int(df[df[Config.COL_STATUS] == "Pendente"][Config.COL_TOTAL].sum())
    k_bai = k_exe + k_nex
    k_tx = k_exe / k_bai if k_bai > 0 else 0.0
    k_proj = int(k_exe + k_tx * k_pen)
    return {
        "total": k_tot, "executadas": k_exe, "nao_executadas": k_nex,
        "pendentes": k_pen, "baixadas": k_bai, "taxa": k_tx,
        "quebra": 1.0 - k_tx, "projecao": k_proj,
    }


def calcular_volumetria_por_tecnico(kpis: Dict[str, Any], n: int) -> Dict[str, Any]:
    """Calcula médias por técnico sem produzir números artificiais quando n=0."""
    n = max(int(n), 0)
    meta = Config.META_EXECUTADAS_TECNICO
    divisor = n if n > 0 else None
    os_t = kpis["total"] / divisor if divisor else 0.0
    exe_t = kpis["executadas"] / divisor if divisor else 0.0
    nex_t = kpis["nao_executadas"] / divisor if divisor else 0.0
    pen_t = kpis["pendentes"] / divisor if divisor else 0.0
    bai_t = kpis["baixadas"] / divisor if divisor else 0.0
    proj_t = kpis["projecao"] / divisor if divisor else 0.0
    ating = exe_t / meta if meta else 0.0
    return {
        "n": n, "os_tec": os_t, "exe_tec": exe_t, "nex_tec": nex_t,
        "pen_tec": pen_t, "bai_tec": bai_t, "proj_tec": proj_t,
        "atingimento": ating, "atingiu_meta": exe_t >= meta,
        "gap_meta": exe_t - meta,
    }


def calcular_volumetria(df: pd.DataFrame, grupos: List[str]) -> pd.DataFrame:
    tabela = (
        df.groupby(grupos + [Config.COL_STATUS], observed=True)[Config.COL_TOTAL]
        .sum().unstack(Config.COL_STATUS, fill_value=0).reset_index()
    )
    for s in Config.STATUS_ORDEM:
        if s not in tabela.columns:
            tabela[s] = 0
        tabela[s] = tabela[s].astype(int)
    tabela["Baixadas"] = tabela["Executada"] + tabela["Não Executada"]
    tabela["Total Alocado"] = tabela["Baixadas"] + tabela["Pendente"]
    tabela["Taxa Execução"] = np.where(tabela["Baixadas"] > 0, tabela["Executada"] / tabela["Baixadas"], 0.0)
    tabela["Taxa Quebra"] = 1.0 - tabela["Taxa Execução"]
    tabela["Projeção"] = (tabela["Executada"] + tabela["Taxa Execução"] * tabela["Pendente"]).astype(int)
    if Config.COL_TECNICO in df.columns:
        n_tec = df.groupby(grupos, observed=True)[Config.COL_TECNICO].nunique().reset_index(name="Técnicos")
        tabela = tabela.merge(n_tec, on=grupos, how="left")
        tabela["Técnicos"] = tabela["Técnicos"].fillna(0).astype(int)
        tabela["OS/Técnico"] = np.where(tabela["Técnicos"] > 0, tabela["Total Alocado"] / tabela["Técnicos"], 0.0)
        tabela["Exec/Técnico"] = np.where(tabela["Técnicos"] > 0, tabela["Executada"] / tabela["Técnicos"], 0.0)
    return tabela.sort_values("Total Alocado", ascending=False)


def gerar_excel(df: pd.DataFrame, nome_aba: str) -> bytes:
    """Gera Excel formatado, com filtros, congelamento do cabeçalho e larguras adaptativas."""
    output = BytesIO()
    export = df.copy()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export.to_excel(writer, index=False, sheet_name=nome_aba[:31])
        ws = writer.sheets[nome_aba[:31]]
        hf = PatternFill("solid", fgColor="0F172A")

        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for cell in ws[1]:
            cell.fill = hf
            cell.font = Font(color="FFFFFF", bold=True)

        for i, col in enumerate(export.columns, 1):
            valores = export[col].head(500).astype(str)
            maior = max([len(str(col))] + [len(v) for v in valores], default=10)
            ws.column_dimensions[get_column_letter(i)].width = min(max(maior + 2, 12), 40)

    output.seek(0)
    return output.getvalue()


# ==========================================================
# VISÕES DE TÉCNICOS
# ==========================================================
def _agrupar_tecnicos(df: pd.DataFrame) -> pd.DataFrame:
    """Agrega O.S. por técnico usando o TOTAL DE TAREFAS como peso."""
    if df.empty:
        return pd.DataFrame()

    base = df.copy()
    base[Config.COL_TOTAL] = pd.to_numeric(base[Config.COL_TOTAL], errors="coerce").fillna(0).clip(lower=0)
    base["__EXEC"] = np.where(base[Config.COL_STATUS].eq("Executada"), base[Config.COL_TOTAL], 0)
    base["__NAO_EXEC"] = np.where(base[Config.COL_STATUS].eq("Não Executada"), base[Config.COL_TOTAL], 0)
    base["__PEND"] = np.where(base[Config.COL_STATUS].eq("Pendente"), base[Config.COL_TOTAL], 0)

    grupos = [Config.COL_TECNICO, Config.COL_MONITOR, Config.COL_REGIAO]
    ag = (
        base.groupby(grupos, observed=True)
        .agg(
            Executadas=("__EXEC", "sum"),
            Nao_Executadas=("__NAO_EXEC", "sum"),
            Pendentes=("__PEND", "sum"),
            Total_Alocado=(Config.COL_TOTAL, "sum"),
        )
        .reset_index()
    )
    return Utils.calcular_metricas_grupo(ag)


def criar_visao_tecnicos_escalados(df: pd.DataFrame) -> pd.DataFrame:
    df_at = df[pd.to_numeric(df[Config.COL_TOTAL], errors="coerce").fillna(0) >= 1].copy()
    ag = _agrupar_tecnicos(df_at)
    if ag.empty:
        return pd.DataFrame()
    ren = Utils.resolver_renomeacao(ag, RENOMEAR_COLUNAS)
    ag = ag.rename(columns=ren)
    ct = RENOMEAR_COLUNAS.get("Taxa Execução", "Taxa Exec.")
    if ct in ag.columns:
        ag = ag.sort_values(ct, ascending=False)
    return ag.reset_index(drop=True)


def criar_visao_tecnicos_montados(df: pd.DataFrame, total_montados: int) -> pd.DataFrame:
    ag = _agrupar_tecnicos(df)
    if ag.empty:
        return pd.DataFrame()
    ren = Utils.resolver_renomeacao(ag, RENOMEAR_COLUNAS)
    ag = ag.rename(columns=ren)
    ct = RENOMEAR_COLUNAS.get("Taxa Execução", "Taxa Exec.")
    if ct in ag.columns:
        ag = ag.sort_values(ct, ascending=False)
    ag.attrs["total_montados_fixo"] = total_montados
    return ag.reset_index(drop=True)


def _resumo_por_monitor(df_tec: pd.DataFrame, col_tec: str) -> pd.DataFrame:
    cm = RENOMEAR_COLUNAS.get("MONITOR", "Monitor")
    cr = RENOMEAR_COLUNAS.get("REGIÃO", "Região")
    ce, cn, cp, cb, ct2, cta, cpr = "Executadas", "Não Exec.", "Pendentes", "Baixadas", "Total Alocado", RENOMEAR_COLUNAS.get("Taxa Execução", "Taxa Exec."), RENOMEAR_COLUNAS.get("Projeção", "Projeção")
    agg: Dict[str, Any] = {}
    for c in [ce, cn, cp, cb, ct2, cpr]:
        if c in df_tec.columns:
            agg[c] = "sum"
    agg[col_tec] = "count"
    gc = [c for c in [cm, cr] if c in df_tec.columns]
    if not gc:
        return pd.DataFrame()
    res = df_tec.groupby(gc, observed=True).agg(agg).reset_index().rename(columns={col_tec: "Técnicos"})
    if cb in res.columns and ce in res.columns:
        res[cta] = np.where(res[cb] > 0, res[ce] / res[cb], 0.0)
    return res.sort_values("Técnicos", ascending=False).reset_index(drop=True)


# ==========================================================
# ESTILOS CSS
# ==========================================================
def aplicar_estilo():
    st.markdown("""
    <style>
    @import url('https://fonts.googleapis.com/css2?family=Manrope:wght@400;500;600;700;800&family=Inter:wght@400;500;600;700;800&display=swap');

    html, body, [class*="css"], .stApp {
        font-family: 'Manrope', 'Inter', -apple-system, sans-serif !important;
    }
    [data-testid="stDataFrame"] * {
        font-variant-numeric: tabular-nums !important;
    }

    .hero-corp {
        background: linear-gradient(135deg, #012869 0%, #1E40AF 50%, #F37C04 100%);
        padding: 32px 40px; border-radius: 16px; color: white;
        box-shadow: 0 10px 40px rgba(1,40,105,0.25); margin-bottom: 24px;
        position: relative; overflow: hidden;
    }
    .hero-corp::before {
        content:''; position:absolute; top:-50%; right:-10%;
        width:400px; height:400px; background:rgba(255,255,255,0.05); border-radius:50%;
    }
    .hero-title { font-size:34px; font-weight:800; margin:0; letter-spacing:-0.8px; }
    .hero-subtitle { font-size:15px; opacity:0.92; margin:6px 0 0; font-weight:400; }

    /* ── CARD DE TÉCNICOS ── */
    .tec-card {
        border-radius: 16px;
        overflow: hidden;
        box-shadow: 0 8px 32px rgba(15, 23, 42, 0.12);
        border: 1px solid #E2E8F0;
    }
    .tec-card-head {
        padding: 20px 24px 16px;
        display: flex;
        align-items: center;
        justify-content: space-between;
        gap: 12px;
    }
    .tec-card-icon {
        width: 48px; height: 48px;
        border-radius: 14px;
        display: flex; align-items: center; justify-content: center;
        font-size: 22px;
        flex-shrink: 0;
    }
    .tec-card-info { flex: 1; }
    .tec-card-tag {
        font-size: 0.6rem; font-weight: 800;
        text-transform: uppercase; letter-spacing: 1.2px;
        opacity: 0.7;
    }
    .tec-card-titulo {
        font-size: 0.95rem; font-weight: 700;
        letter-spacing: -0.2px; margin: 2px 0 0;
        line-height: 1.25;
    }
    .tec-card-numero {
        text-align: right; flex-shrink: 0;
    }
    .tec-card-n-val {
        font-size: 3.2rem; font-weight: 800;
        line-height: 1; letter-spacing: -2px;
        font-variant-numeric: tabular-nums;
    }
    .tec-card-n-lab {
        font-size: 0.6rem; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.8px;
        opacity: 0.65; margin-top: 2px; text-align: right;
    }
    .tec-card-body {
        background: #FFFFFF;
        padding: 14px 20px 16px;
        display: grid;
        grid-template-columns: 1fr 1fr 1fr;
        gap: 8px;
    }
    .tec-card-metrica {
        background: #F8FAFC;
        border: 1px solid #F1F5F9;
        border-radius: 10px;
        padding: 10px 12px;
        text-align: center;
    }
    .tec-card-m-lab {
        font-size: 0.58rem; font-weight: 700;
        text-transform: uppercase; letter-spacing: 0.5px;
        color: #64748B; margin-bottom: 3px;
    }
    .tec-card-m-val {
        font-size: 1.2rem; font-weight: 800;
        letter-spacing: -0.3px; font-variant-numeric: tabular-nums;
        color: #0F172A; line-height: 1.15;
    }
    .tec-card-m-sub {
        font-size: 0.6rem; font-weight: 600;
        color: #94A3B8; margin-top: 1px;
    }
    .tec-card-foot {
        background: #F8FAFC;
        border-top: 1px solid #E2E8F0;
        padding: 10px 20px;
        display: flex; align-items: center;
        justify-content: space-between; gap: 10px;
    }
    .tec-barra {
        flex: 1; height: 8px;
        background: #E2E8F0; border-radius: 999px;
        overflow: hidden;
    }
    .tec-barra-fill {
        height: 100%; border-radius: 999px;
        transition: width 0.4s ease;
    }
    .tec-meta-txt {
        font-size: 0.68rem; font-weight: 700;
        white-space: nowrap; font-variant-numeric: tabular-nums;
    }

    /* ── FAIXA DE DIFERENÇA ── */
    .diff-strip {
        background: #F8FAFC; border: 1px solid #E2E8F0;
        border-radius: 12px; padding: 12px 20px;
        display: flex; align-items: center;
        flex-wrap: wrap; gap: 16px;
        font-size: 0.82rem; margin: 6px 0 4px;
    }
    .diff-strip-label {
        font-weight: 700; color: #64748B;
        text-transform: uppercase; letter-spacing: 0.6px;
        font-size: 0.68rem;
    }
    .diff-strip-val {
        font-weight: 800; font-size: 1.05rem;
    }
    .diff-strip-detail {
        color: #64748B;
    }
    .diff-strip-right {
        margin-left: auto; color: #475569; font-weight: 600;
    }

    /* ── KPI / BADGES / RESULTADO ── */
    .kpi-card {
        padding: 1.4rem 1.6rem; border-radius: 1rem;
        border-left: 5px solid;
        box-shadow: 0 4px 12px rgba(0,0,0,0.06);
        transition: transform 0.15s ease, box-shadow 0.15s ease;
    }
    .kpi-card:hover { transform: translateY(-3px); box-shadow: 0 8px 20px rgba(0,0,0,0.1); }
    .kpi-val { font-size: 1.95rem; font-weight: 800; line-height: 1.1; margin: 0.3rem 0; letter-spacing: -0.5px; font-variant-numeric: tabular-nums; }
    .kpi-lab { font-size: 0.72rem; text-transform: uppercase; font-weight: 700; letter-spacing: 0.08em; }
    .kpi-sub { font-size: 0.78rem; margin-top: 0.2rem; font-weight: 500; }

    .visao-badge { display: inline-flex; align-items: center; gap: 6px; padding: 6px 16px; border-radius: 999px; font-size: 0.78rem; font-weight: 700; letter-spacing: 0.5px; text-transform: uppercase; border: 2px solid; margin-bottom: 12px; }
    .badge-escalados { background:#D1FAE5; color:#065F46; border-color:#10B981; }
    .badge-montados  { background:#FEF3C7; color:#92400E; border-color:#F59E0B; }

    .resultado-base { background: linear-gradient(135deg, #0F172A 0%, #1E3A5F 100%); padding: 1rem 1.5rem; border-radius: 0.75rem; margin-bottom: 1.5rem; display: flex; align-items: center; flex-wrap: wrap; gap: 0.6rem; box-shadow: 0 4px 16px rgba(15,23,42,0.15); }
    .resultado-base-label { color: #94A3B8; font-size: 0.8rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; }
    .resultado-base-regiao { padding: 0.3rem 0.9rem; border-radius: 999px; font-size: 0.82rem; font-weight: 700; border: 2px solid; }
    .resultado-base-count { color: #64748B; font-size: 0.72rem; margin-left: auto; font-weight: 600; font-variant-numeric: tabular-nums; }

    .section-header { display: flex; align-items: center; gap: 0.6rem; margin: 1.5rem 0 0.8rem; padding-bottom: 0.4rem; border-bottom: 2px solid #E2E8F0; }
    .section-header h3 { margin: 0; font-size: 1.1rem; color: #0F172A; font-weight: 700; letter-spacing: -0.2px; }

    .montados-box { background: linear-gradient(135deg, #FEF3C7 0%, #FDE68A 100%); border: 2px solid #F59E0B; border-radius: 12px; padding: 14px 16px; margin: 8px 0 4px 0; }
    .montados-box-title { font-size: 0.7rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.08em; color: #92400E; margin-bottom: 6px; display: flex; align-items: center; gap: 5px; }

    .stButton > button { font-family: 'Manrope', 'Inter', sans-serif !important; font-weight: 600 !important; border-radius: 0.5rem !important; }
    .stTabs [data-baseweb="tab-list"] { gap: 8px; }
    .stTabs [data-baseweb="tab"] { font-weight: 600; border-radius: 8px 8px 0 0; }

    /* Estilização para deixar o st.radio como botões de alternância modernos (Segmented Controls) */
    div[data-testid="stRadio"] > div {
        flex-direction: row !important;
        gap: 0 !important;
        background: #F1F5F9;
        border-radius: 12px;
        padding: 4px;
        margin-bottom: 12px;
    }
    div[data-testid="stRadio"] label {
        flex: 1 !important;
        text-align: center !important;
        padding: 8px 16px !important;
        border-radius: 10px !important;
        font-weight: 700 !important;
        font-size: 0.82rem !important;
        letter-spacing: 0.3px !important;
        transition: all 0.2s ease !important;
        cursor: pointer !important;
        margin: 0 !important;
        border: none !important;
        background: transparent !important;
    }
    div[data-testid="stRadio"] label[data-checked="true"], 
    div[data-testid="stRadio"] label:has(input:checked) {
        background: #0F172A !important;
        color: #FFFFFF !important;
        box-shadow: 0 4px 12px rgba(15, 23, 42, 0.15) !important;
    }
    div[data-testid="stRadio"] label p {
        color: inherit !important;
        font-weight: inherit !important;
    }
    div[data-testid="stRadio"] input[type="radio"] {
        display: none !important;
    }
    </style>
    """, unsafe_allow_html=True)


# ==========================================================
# COMPONENTES VISUAIS
# ==========================================================
def render_resultado_base(regioes: List[str], total: int):
    badges = "".join(
        f'<span class="resultado-base-regiao" style="background:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["bg"]};color:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["text"]};border-color:{CORES_REGIAO.get(r, CORES_REGIAO["OUTRAS"])["border"]}">{html.escape(str(r))}</span>'
        for r in sorted(regioes)
    )
    st.markdown(f'<div class="resultado-base"><span class="resultado-base-label">📋 Resultado da Base:</span>{badges}<span class="resultado-base-count">{total:,} registros</span></div>', unsafe_allow_html=True)


def render_kpi(col, label: str, value: str, sub: str, tema: str = "azul"):
    t = TEMAS_CARD.get(tema, TEMAS_CARD["azul"])
    col.markdown(f'<div class="kpi-card" style="background:{t["fundo"]};border-left-color:{t["borda"]};"><div class="kpi-lab" style="color:{t["titulo"]}">{label}</div><div class="kpi-val" style="color:{t["texto"]}">{value}</div><div class="kpi-sub" style="color:{t["titulo"]}">{sub}</div></div>', unsafe_allow_html=True)


def render_section(titulo: str):
    st.markdown(f'<div class="section-header"><h3>{titulo}</h3></div>', unsafe_allow_html=True)


def _fmt_br(n: float, casas: int = 1) -> str:
    return f"{n:,.{casas}f}".replace(",", "X").replace(".", ",").replace("X", ".")


def render_card_tecnicos(
    col,
    *,
    titulo: str,
    tag: str,
    fonte: str,
    n_tecnicos: int,
    vol: Dict[str, Any],
    kpis: Dict[str, Any],
    variante: Literal["escalados", "montados"] = "escalados",
):
    """
    Card visual completo mostrando:
    • Número grande de técnicos
    • 6 métricas de produtividade por técnico
    • Barra de atingimento da meta exec/técnico
    """
    if variante == "escalados":
        head_bg = "linear-gradient(135deg, #064E3B 0%, #047857 55%, #10B981 100%)"
        head_fg = "#ECFDF5"
        accent = "#10B981"
        icon_bg = "rgba(255,255,255,0.15)"
        icon_emoji = "🟢"
        fill_ok, fill_nok = "#10B981", "#F59E0B"
    else:
        head_bg = "linear-gradient(135deg, #78350F 0%, #B45309 55%, #F59E0B 100%)"
        head_fg = "#FFFBEB"
        accent = "#F59E0B"
        icon_bg = "rgba(255,255,255,0.15)"
        icon_emoji = "🟡"
        fill_ok, fill_nok = "#10B981", "#EF4444"

    meta = Config.META_EXECUTADAS_TECNICO
    pct_meta = min(max(vol["atingimento"], 0.0), 1.5)
    barra_pct = min(pct_meta * 100 / 1.5, 100)
    cor_barra = fill_ok if vol["atingiu_meta"] else fill_nok
    selo = "✅ Meta atingida" if vol["atingiu_meta"] else "⚠️ Abaixo"
    cor_selo = "#047857" if vol["atingiu_meta"] else "#B45309"
    gap = vol["gap_meta"]
    gap_txt = f"{'+' if gap >= 0 else ''}{_fmt_br(gap)} vs meta {meta}"

    card_html = textwrap.dedent(f"""
    <div class="tec-card">
        <div class="tec-card-head" style="background:{head_bg};color:{head_fg};">
            <div class="tec-card-icon" style="background:{icon_bg}">{icon_emoji}</div>
            <div class="tec-card-info">
                <div class="tec-card-tag">{tag}</div>
                <div class="tec-card-titulo">{titulo}</div>
            </div>
            <div class="tec-card-numero">
                <div class="tec-card-n-val">{n_tecnicos:,}</div>
                <div class="tec-card-n-lab">técnicos</div>
            </div>
        </div>

        <div class="tec-card-body">
            <div class="tec-card-metrica">
                <div class="tec-card-m-lab">O.S. / Técnico</div>
                <div class="tec-card-m-val">{_fmt_br(vol['os_tec'])}</div>
                <div class="tec-card-m-sub">{kpis['total']:,} total</div>
            </div>
            <div class="tec-card-metrica">
                <div class="tec-card-m-lab">Exec. / Técnico</div>
                <div class="tec-card-m-val" style="color:{accent}">{_fmt_br(vol['exe_tec'])}</div>
                <div class="tec-card-m-sub">{kpis['executadas']:,} exec.</div>
            </div>
            <div class="tec-card-metrica">
                <div class="tec-card-m-lab">Projeção / Téc.</div>
                <div class="tec-card-m-val">{_fmt_br(vol['proj_tec'])}</div>
                <div class="tec-card-m-sub">{kpis['projecao']:,} proj.</div>
            </div>
            <div class="tec-card-metrica">
                <div class="tec-card-m-lab">Não Exec. / Téc.</div>
                <div class="tec-card-m-val">{_fmt_br(vol['nex_tec'])}</div>
                <div class="tec-card-m-sub">quebra {kpis['quebra']:.1%}</div>
            </div>
            <div class="tec-card-metrica">
                <div class="tec-card-m-lab">Pendentes / Téc.</div>
                <div class="tec-card-m-val">{_fmt_br(vol['pen_tec'])}</div>
                <div class="tec-card-m-sub">{kpis['pendentes']:,} pend.</div>
            </div>
            <div class="tec-card-metrica">
                <div class="tec-card-m-lab">Taxa Execução</div>
                <div class="tec-card-m-val">{kpis['taxa']:.1%}</div>
                <div class="tec-card-m-sub">meta {Config.META_EXECUCAO:.0%}</div>
            </div>
        </div>

        <div class="tec-card-foot">
            <span class="tec-meta-txt" style="color:{cor_selo}">{selo}</span>
            <div class="tec-barra">
                <div class="tec-barra-fill" style="width:{barra_pct:.1f}%;background:{cor_barra};"></div>
            </div>
            <span class="tec-meta-txt" style="color:#334155">{gap_txt}</span>
        </div>
    </div>
    """).strip()
    # O Markdown do Streamlit encerra blocos HTML em determinadas quebras de linha.
    # Removemos linhas em branco e a indentação para garantir que todo o card
    # seja interpretado como um único bloco HTML, evitando que o conteúdo
    # interno apareça literalmente na tela.
    card_html = re.sub(r"\n\s*\n", "\n", card_html)
    card_html = "\n".join(line.strip() for line in card_html.splitlines() if line.strip()).strip()
    col.markdown(card_html, unsafe_allow_html=True)


def render_faixa_diferenca(n_esc: int, n_mon: int, vol_esc: Dict, vol_mon: Dict):
    """Faixa comparativa entre os dois cards."""
    dif = n_mon - n_esc
    cor = "#B91C1C" if dif > 0 else ("#047857" if dif < 0 else "#64748B")
    sinal = "+" if dif > 0 else ""
    st.markdown(f"""
    <div class="diff-strip">
        <span class="diff-strip-label">Diferença</span>
        <span class="diff-strip-val" style="color:{cor}">{sinal}{dif} técnico(s)</span>
        <span class="diff-strip-detail">
            Montados ({n_mon}) − Escalados ({n_esc})
        </span>
        <span class="diff-strip-right">
            Exec/téc: Escalados <strong>{_fmt_br(vol_esc['exe_tec'])}</strong>
            &nbsp;×&nbsp;
            Montados <strong>{_fmt_br(vol_mon['exe_tec'])}</strong>
        </span>
    </div>
    """, unsafe_allow_html=True)


# ==========================================================
# DATAFRAME ESTILIZADO
# ==========================================================
def render_dataframe(
    df: pd.DataFrame,
    titulo: str = "",
    icone: str = "📊",
    badge: str = "",
    fmt: Optional[Dict[str, Any]] = None,
    color_col: Optional[str] = None,
    color_meta: Optional[float] = None,
    height: int | Literal["auto", "stretch", "content"] = "auto",
    adicionar_totais: bool = True,
):
    df_d = df.copy()
    mapa = Utils.resolver_renomeacao(df_d, RENOMEAR_COLUNAS)
    df_d = df_d.rename(columns=mapa)

    ce = mapa.get("Executada", "Executadas")
    cn = mapa.get("Não Executada", "Não Exec.")
    cp = mapa.get("Pendente", "Pendentes")
    cb = mapa.get("Baixadas", "Baixadas")
    ct = mapa.get("Total Alocado", "Total Alocado")
    cta = mapa.get("Taxa Execução", "Taxa Exec.")
    cq = mapa.get("Taxa Quebra", "Taxa Quebra")
    cpr = mapa.get("Projeção", "Projeção")

    for co in COLUNAS_INTEIRAS:
        cd = mapa.get(co, co)
        if cd in df_d.columns:
            df_d[cd] = pd.to_numeric(df_d[cd], errors="coerce").fillna(0).astype(int)

    ranking: Dict[float, float] = {}
    if cta in df_d.columns and len(df_d):
        valores = pd.to_numeric(df_d[cta], errors="coerce")
        n = len(valores.dropna())
        if n:
            # Percentil por valor evita que empates recebam cores inconsistentes.
            ranks = valores.rank(method="average", ascending=False, pct=True)
            ranking = {
                float(v): float(p)
                for v, p in zip(valores, ranks)
                if pd.notna(v) and pd.notna(p)
            }

    if adicionar_totais and len(df_d):
        tr: Dict[str, Any] = {c: (0 if pd.api.types.is_numeric_dtype(df_d[c]) else "") for c in df_d.columns}
        cn2 = df_d.columns[1] if len(df_d.columns) > 1 else df_d.columns[0]
        tr[cn2] = "TOTAL GERAL"
        for c in [ce, cn, cp, cb, ct, cpr, "Técnicos"]:
            if c in df_d.columns:
                tr[c] = int(df_d[c].sum())
        if cb in df_d.columns and ce in df_d.columns:
            bt = df_d[cb].sum()
            et = df_d[ce].sum()
            if cta in df_d.columns:
                tr[cta] = et / bt if bt else 0.0
            if cq in df_d.columns:
                tr[cq] = 1 - et / bt if bt else 0.0
        nt = tr.get("Técnicos", 0) or 0
        for cc in ["OS/Técnico", "Exec/Técnico"]:
            if cc in df_d.columns:
                ref = ct if "OS" in cc else ce
                tr[cc] = tr.get(ref, 0) / nt if nt else 0.0
        df_d = pd.concat([df_d, pd.DataFrame([tr])], ignore_index=True)

    data = pd.Timestamp.now().strftime("%d/%m/%Y")
    st.markdown(f'<div style="background:#0B1739;padding:16px 24px;border-radius:12px 12px 0 0;color:#F9FAFB;margin-bottom:0;border-bottom:2px solid #1E3A8A;font-family:Manrope,Inter,sans-serif;"><span style="font-weight:700;font-size:0.85rem;letter-spacing:1.2px;">{icone}  {titulo.upper()} — {data}</span></div>', unsafe_allow_html=True)

    sty = df_d.style

    def fp(v):
        if v == "" or v is None or pd.isna(v):
            return ""
        try:
            return f"{float(v)*100:.1f}%".replace(".", ",")
        except (ValueError, TypeError):
            return str(v)

    def fi(v):
        if v == "" or v is None or pd.isna(v):
            return ""
        try:
            return f"{int(v):,}".replace(",", ".")
        except (ValueError, TypeError):
            return str(v)

    def fd(v):
        if v == "" or v is None or pd.isna(v):
            return ""
        try:
            return f"{float(v):.1f}".replace(".", ",")
        except (ValueError, TypeError):
            return str(v)

    fm: Dict[str, Any] = {}
    if cta in df_d.columns: fm[cta] = fp
    if cq in df_d.columns: fm[cq] = fp
    for c in [ce, cn, cp, cb, ct, cpr, "Técnicos"]:
        if c in df_d.columns: fm[c] = fi
    for c in ["OS/Técnico", "Exec/Técnico"]:
        if c in df_d.columns: fm[c] = fd
    if fm:
        sty = sty.format(fm)

    if cta in df_d.columns:
        def _ct(v):
            if v == "" or pd.isna(v): return ""
            try:
                vv = float(v)
                p = ranking.get(vv, 0.5)
                if p <= 0.2: bg, fg = "#D1FAE5", "#047857"
                elif p <= 0.4: bg, fg = "#ECFDF5", "#059669"
                elif p <= 0.6: bg, fg = "#FEF3C7", "#B45309"
                elif p <= 0.8: bg, fg = "#FFEDD5", "#C2410C"
                else: bg, fg = "#FEE2E2", "#B91C1C"
                return f"background-color:{bg};color:{fg};font-weight:700;text-align:center;"
            except (ValueError, TypeError): return ""
        sty = sty.map(_ct, subset=[cta])

    if cq in df_d.columns:
        def _cq(v):
            if v == "" or pd.isna(v): return ""
            try:
                vv = float(v)
                if vv >= 0.25: return "color:#B91C1C;font-weight:600;"
                elif vv >= 0.15: return "color:#B45309;font-weight:600;"
                return "color:#047857;font-weight:600;"
            except (ValueError, TypeError): return ""
        sty = sty.map(_cq, subset=[cq])

    if ce in df_d.columns:
        sty = sty.set_properties(**{"background-color": "#FEF9C3", "color": "#854D0E", "font-weight": "700"}, subset=[ce])
    if ct in df_d.columns:
        sty = sty.set_properties(**{"background-color": "#D1FAE5", "color": "#065F46", "font-weight": "700"}, subset=[ct])
    if cpr in df_d.columns:
        sty = sty.set_properties(**{"background": "#1E293B", "color": "#FFFFFF", "font-weight": "800"}, subset=[cpr])

    def _et(row):
        for v in row:
            if isinstance(v, str) and "TOTAL GERAL" in str(v).upper():
                return ["background-color:#E2E8F0;font-weight:800;color:#0F172A;border-top:2px solid #64748B;"] * len(row)
        return [""] * len(row)
    sty = sty.apply(_et, axis=1)

    sty = sty.set_table_styles([
        {"selector": "thead th", "props": [("background-color", "#1E293B"), ("color", "#F1F5F9"), ("font-weight", "600"), ("text-align", "center"), ("padding", "13px 12px"), ("border", "none"), ("font-size", "0.72rem"), ("text-transform", "uppercase"), ("letter-spacing", "1px")]},
        {"selector": "tbody td", "props": [("padding", "12px 15px"), ("border-bottom", "1px solid #F1F5F9"), ("font-size", "0.85rem"), ("text-align", "center"), ("color", "#334155"), ("font-variant-numeric", "tabular-nums")]},
        {"selector": "tbody td:nth-child(1)", "props": [("text-align", "left"), ("font-weight", "600"), ("color", "#475569"), ("padding-left", "18px")]},
        {"selector": "tbody td:nth-child(2)", "props": [("text-align", "left"), ("font-weight", "600"), ("color", "#0F172A"), ("padding-left", "15px")]},
        {"selector": "tbody tr:nth-child(even) td", "props": [("background-color", "#FAFBFC")]},
        {"selector": "", "props": [("border-radius", "0 0 12px 12px"), ("overflow", "hidden"), ("box-shadow", "0 4px 16px rgba(15,23,42,0.1)"), ("border", "1px solid #E2E8F0"), ("border-top", "none")]},
    ])

    st.dataframe(sty, use_container_width=True, hide_index=True, height=height)


# ==========================================================
# TABELAS DE TÉCNICOS (ABAS DETALHADAS)
# ==========================================================
def renderizar_volumetria_tecnicos(df: pd.DataFrame, total_montados_fixo: int):
    render_section("🧑‍🔧 Detalhamento por Técnico")

    df_esc = criar_visao_tecnicos_escalados(df)
    df_mon = criar_visao_tecnicos_montados(df, total_montados_fixo)
    ct = RENOMEAR_COLUNAS.get("TÉCNICO", "Técnico")
    ne = df_esc[ct].nunique() if ct in df_esc.columns else 0
    nm = total_montados_fixo
    dif = max(0, nm - ne)

    a1, a2 = st.tabs(["🟢 Escalados — Base Importada", "🟡 Montados — Valor Fixo"])

    with a1:
        st.markdown('<div class="visao-badge badge-escalados">🟢 Visão 1 · Escalados · Fonte: Base Importada</div><p style="font-size:0.85rem;color:#475569;margin:0 0 12px;">Técnicos com ao menos <strong>1 O.S. alocada</strong> na base importada.</p>', unsafe_allow_html=True)
        if df_esc.empty:
            st.warning("⚠️ Nenhum técnico escalado encontrado.")
        else:
            render_dataframe(df_esc, titulo=f"Técnicos Escalados ({ne})", icone="🟢", height=520)
            c1, _ = st.columns([1, 5])
            with c1:
                st.download_button("📥 Baixar Escalados", gerar_excel(df_esc, "Escalados"), "tecnicos_escalados.xlsx", use_container_width=True)
            with st.expander("📊 Resumo por Monitor — Escalados"):
                r = _resumo_por_monitor(df_esc, ct)
                if not r.empty:
                    st.dataframe(r, use_container_width=True, hide_index=True)

    with a2:
        st.markdown(f'<div class="visao-badge badge-montados">🟡 Visão 2 · Montados · Fixo: {nm}</div><p style="font-size:0.85rem;color:#475569;margin:0 0 12px;">Todos os técnicos da base. Diferença de <strong>{dif}</strong> indica montados sem rota.</p>', unsafe_allow_html=True)
        if df_mon.empty:
            st.warning("⚠️ Nenhum técnico encontrado.")
        else:
            nb = df_mon[ct].nunique() if ct in df_mon.columns else 0
            render_dataframe(df_mon, titulo=f"Técnicos Montados (fixo: {nm} | base: {nb})", icone="🟡", height=520)
            c3, _ = st.columns([1, 5])
            with c3:
                st.download_button("📥 Baixar Montados", gerar_excel(df_mon, "Montados"), "tecnicos_montados.xlsx", use_container_width=True)
            with st.expander("📊 Resumo por Monitor — Montados"):
                r = _resumo_por_monitor(df_mon, ct)
                if not r.empty:
                    st.dataframe(r, use_container_width=True, hide_index=True)


# ==========================================================
# GRÁFICOS
# ==========================================================
def plot_status_pie(df):
    res = df.groupby(Config.COL_STATUS)[Config.COL_TOTAL].sum()
    fig = go.Figure(data=[go.Pie(labels=res.index, values=res.values, hole=0.5, marker=dict(colors=[Config.CORES_STATUS.get(s) for s in res.index]), textinfo="label+percent", textfont_size=13)])
    fig.update_layout(height=370, margin=dict(t=40, b=10, l=10, r=10), title=dict(text="Distribuição de Status", font=dict(size=15)), showlegend=False)
    return fig


def plot_ranking_monitor(df):
    tab = calcular_volumetria(df, [Config.COL_MONITOR]).nlargest(15, "Total Alocado")
    fig = px.bar(tab, x="Taxa Execução", y=Config.COL_MONITOR, orientation="h", color="Taxa Execução", color_continuous_scale="RdYlGn", range_color=[0.4, 0.9], text=tab["Taxa Execução"].apply(lambda v: f"{v:.0%}"))
    fig.add_vline(x=Config.META_EXECUCAO, line_dash="dash", line_color="#0F172A", annotation_text=f"Meta {Config.META_EXECUCAO:.0%}", annotation_position="top")
    fig.update_traces(textposition="outside")
    fig.update_layout(height=420, title=dict(text="Taxa de Execução por Monitor (Top 15)", font=dict(size=15)), margin=dict(l=10, r=10, t=50, b=10), yaxis=dict(autorange="reversed"))
    return fig


def plot_comparativo(vol_esc, vol_mon):
    df_p = pd.DataFrame([
        {"Métrica": "O.S./Téc.", "Visão": "Escalados", "Valor": vol_esc["os_tec"]},
        {"Métrica": "O.S./Téc.", "Visão": "Montados", "Valor": vol_mon["os_tec"]},
        {"Métrica": "Exec./Téc.", "Visão": "Escalados", "Valor": vol_esc["exe_tec"]},
        {"Métrica": "Exec./Téc.", "Visão": "Montados", "Valor": vol_mon["exe_tec"]},
        {"Métrica": "Proj./Téc.", "Visão": "Escalados", "Valor": vol_esc["proj_tec"]},
        {"Métrica": "Proj./Téc.", "Visão": "Montados", "Valor": vol_mon["proj_tec"]},
    ])
    fig = px.bar(df_p, x="Métrica", y="Valor", color="Visão", barmode="group", color_discrete_map={"Escalados": "#10B981", "Montados": "#F59E0B"}, text=df_p["Valor"].apply(lambda v: f"{v:.1f}"), title="Produtividade por Técnico — Escalados × Montados")
    fig.add_hline(y=Config.META_EXECUTADAS_TECNICO, line_dash="dash", line_color="#0F172A", annotation_text=f"Meta {Config.META_EXECUTADAS_TECNICO}", annotation_position="top right")
    fig.update_traces(textposition="outside")
    fig.update_layout(height=360, margin=dict(t=50, b=20, l=10, r=10), yaxis_title="Média / técnico", legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1))
    return fig


# ==========================================================
# MAIN
# ==========================================================
def main():
    aplicar_estilo()

    st.markdown("""
    <div class="hero-corp">
        <div style="position:relative;z-index:2;">
            <h1 class="hero-title">📊 Gestão de Volumetria</h1>
            <p class="hero-subtitle">Análise executiva de performance e projeções operacionais</p>
        </div>
    </div>
    """, unsafe_allow_html=True)

    if "base_data" not in st.session_state:
        st.session_state.base_data = None

    # ── SIDEBAR ──────────────────────────────────────────
    with st.sidebar:
        st.header("⚙️ Configurações")
        if st.button("🔄 Reiniciar Painel", use_container_width=True):
            st.session_state.base_data = None
            st.session_state.pop("input_montados", None)
            st.session_state.pop("_montados_init", None)
            st.rerun()

        st.divider()

        st.markdown('<div class="montados-box"><div class="montados-box-title">🟡 Técnicos Montados</div></div>', unsafe_allow_html=True)

        sug = 0
        if st.session_state.base_data is not None:
            sug = int(st.session_state.base_data[Config.COL_TECNICO].nunique())

        if st.session_state.base_data is not None and not st.session_state.get("_montados_init"):
            st.session_state.input_montados = sug
            st.session_state._montados_init = True
        if "input_montados" not in st.session_state:
            st.session_state.input_montados = 0

        total_montados_fixo: int = st.number_input(
            "Nº fixo de técnicos montados",
            min_value=0, max_value=9999, step=1,
            help="Número fixo de técnicos montados na escala do dia.",
            key="input_montados",
        )

        st.markdown(f'<div style="background:#FEF3C7;border:1px solid #F59E0B;border-radius:8px;padding:8px 12px;margin-top:4px;font-size:0.8rem;color:#92400E;font-weight:600;">📋 Montados: <strong>{total_montados_fixo}</strong></div>', unsafe_allow_html=True)
        st.divider()

        if st.session_state.base_data is not None:
            diag = st.session_state.base_data.attrs.get("diagnostico", {})
            st.write(f"📥 **Entrada:** {diag.get('Inicial', 0)}")
            st.write(f"🗑️ **Vazios:** {diag.get('Removidos por contrato vazio', 0)}")
            st.write(f"📈 **Processadas:** {diag.get('Final', 0)}")

    # ── UPLOAD ───────────────────────────────────────────
    if st.session_state.base_data is None:
        render_section("📁 Importação de Dados")
        u = st.file_uploader("Selecione a base (Excel/CSV)", type=["xlsx", "xls", "csv"])
        if u:
            try:
                with st.spinner("Processando..."):
                    arquivo = DataLoader.ler_arquivo(u.getvalue(), u.name)
                    if arquivo.empty:
                        st.error("O arquivo não possui registros.")
                        return

                    st.session_state.base_data = DataLoader.preparar_base(
                        arquivo,
                        DataLoader.buscar_hierarquia_gsheets(),
                    )
                    st.rerun()
            except Exception as exc:
                st.error(f"Não foi possível processar o arquivo: {exc}")
                st.exception(exc)
        return

    df_full = st.session_state.base_data

    # ── FILTROS ──────────────────────────────────────────
    with st.sidebar:
        st.header("🎯 Filtros")
        mons = sorted(df_full[Config.COL_MONITOR].dropna().astype(str).unique())
        sel_m = st.multiselect("Monitor", mons, default=mons)
        regs = sorted(df_full[Config.COL_REGIAO].dropna().astype(str).unique())
        sel_r = st.multiselect("Região", regs, default=regs)

        tecnicos = sorted(df_full[Config.COL_TECNICO].dropna().astype(str).unique())
        sel_t = st.multiselect("Técnico", tecnicos, default=tecnicos)

        statuses = Config.STATUS_ORDEM
        sel_s = st.multiselect("Status", statuses, default=statuses)

        if st.button("↩️ Limpar filtros", use_container_width=True):
            st.rerun()

    df = df_full[
        df_full[Config.COL_MONITOR].isin(sel_m)
        & df_full[Config.COL_REGIAO].isin(sel_r)
        & df_full[Config.COL_TECNICO].isin(sel_t)
        & df_full[Config.COL_STATUS].isin(sel_s)
    ]
    if df.empty:
        st.warning("Nenhum dado selecionado.")
        return

    # ── RESULTADO DA BASE ────────────────────────────────
    render_resultado_base(sorted(df[Config.COL_REGIAO].unique()), len(df))

    # ── KPIs DE VOLUME ───────────────────────────────────
    kpis = calcular_kpis(df)

    c1, c2, c3, c4 = st.columns(4)
    render_kpi(c1, "Total Alocado", f"{kpis['total']:,}", f"{kpis['pendentes']:,} pendentes", "azul")
    render_kpi(c2, "Executadas", f"{kpis['executadas']:,}", f"Taxa: {kpis['taxa']:.1%}", "verde")
    render_kpi(c3, "Projeção Final", f"{kpis['projecao']:,}", "Baseado na taxa atual", "escuro")
    render_kpi(c4, "Meta", f"{Config.META_EXECUCAO:.0%}", "Referência", "amarelo")

    st.markdown("")
    s1, s2, s3 = st.columns(3)
    render_kpi(s1, "Não Executadas", f"{kpis['nao_executadas']:,}", f"Quebra: {kpis['quebra']:.1%}", "vermelho")
    render_kpi(s2, "Baixadas", f"{kpis['baixadas']:,}", "Exec + Não Exec", "roxo")
    render_kpi(s3, "Pendentes", f"{kpis['pendentes']:,}", f"{kpis['pendentes']/kpis['total']:.1%} do total" if kpis["total"] else "0%", "cinza")

    # ==========================================================
    # ★ CARDS DE TÉCNICOS — DOIS PAINÉIS LADO A LADO ★
    # ==========================================================
    n_escalados = int(df.loc[df[Config.COL_TOTAL] >= 1, Config.COL_TECNICO].nunique())
    n_montados = int(total_montados_fixo)

    vol_esc = calcular_volumetria_por_tecnico(kpis, n_escalados)
    vol_mon = calcular_volumetria_por_tecnico(kpis, n_montados)

    render_section("👷 Nº de Técnicos — Escalados × Montados")

    p1, p2 = st.columns(2)

    render_card_tecnicos(
        p1,
        titulo="Técnicos Escalados",
        tag="Painel 1 · Importação",
        fonte="Base importada (técnicos com ≥ 1 O.S.)",
        n_tecnicos=n_escalados,
        vol=vol_esc,
        kpis=kpis,
        variante="escalados",
    )

    render_card_tecnicos(
        p2,
        titulo="Técnicos Montados",
        tag="Painel 2 · Valor Fixo",
        fonte="Número digitado no painel lateral",
        n_tecnicos=n_montados,
        vol=vol_mon,
        kpis=kpis,
        variante="montados",
    )

    # Faixa de diferença entre os dois
    render_faixa_diferenca(n_escalados, n_montados, vol_esc, vol_mon)

    # ==========================================================
    # ★ MÉDIAS POR TÉCNICO — ESCALADOS × MONTADOS ★
    # ==========================================================
    render_section("📊 Médias por Técnico — Escalados × Montados")

    def _media_card(col, titulo, valor, detalhe, tema):
        render_kpi(col, titulo, _fmt_br(valor), detalhe, tema)

    m1, m2, m3, m4 = st.columns(4)
    _media_card(m1, "OS / Escalado", vol_esc["os_tec"], f"{n_escalados:,} técnicos escalados", "verde")
    _media_card(m2, "OS / Montado", vol_mon["os_tec"], f"{n_montados:,} técnicos montados", "amarelo")
    _media_card(m3, "Executadas / Escalado", vol_esc["exe_tec"], f"Meta: {_fmt_br(Config.META_EXECUTADAS_TECNICO)}", "verde")
    _media_card(m4, "Executadas / Montado", vol_mon["exe_tec"], f"Meta: {_fmt_br(Config.META_EXECUTADAS_TECNICO)}", "amarelo")

    m5, m6, m7, m8 = st.columns(4)
    _media_card(m5, "Pendentes / Escalado", vol_esc["pen_tec"], "Média de pendências", "azul")
    _media_card(m6, "Pendentes / Montado", vol_mon["pen_tec"], "Média de pendências", "cinza")
    _media_card(m7, "Projeção / Escalado", vol_esc["proj_tec"], "Projeção média final", "escuro")
    _media_card(m8, "Projeção / Montado", vol_mon["proj_tec"], "Projeção média final", "roxo")

    # Tabela comparativa das médias
    medias = pd.DataFrame({
        "Indicador": [
            "Técnicos", "OS / Técnico", "Executadas / Técnico",
            "Não Executadas / Técnico", "Pendentes / Técnico",
            "Baixadas / Técnico", "Projeção / Técnico", "Atingimento da Meta"
        ],
        "Escalados": [
            n_escalados, vol_esc["os_tec"], vol_esc["exe_tec"],
            vol_esc["nex_tec"], vol_esc["pen_tec"], vol_esc["bai_tec"],
            vol_esc["proj_tec"], vol_esc["atingimento"]
        ],
        "Montados": [
            n_montados, vol_mon["os_tec"], vol_mon["exe_tec"],
            vol_mon["nex_tec"], vol_mon["pen_tec"], vol_mon["bai_tec"],
            vol_mon["proj_tec"], vol_mon["atingimento"]
        ],
    })
    medias["Diferença (Montados - Escalados)"] = medias["Montados"] - medias["Escalados"]
    medias.loc[medias["Indicador"] == "Atingimento da Meta", ["Escalados", "Montados", "Diferença (Montados - Escalados)"]] *= 100

    st.dataframe(
        medias.style.format({
            "Escalados": lambda x: f"{x:,.1f}".replace(",", "X").replace(".", ",").replace("X", "."),
            "Montados": lambda x: f"{x:,.1f}".replace(",", "X").replace(".", ",").replace("X", "."),
            "Diferença (Montados - Escalados)": lambda x: f"{x:+,.1f}".replace(",", "X").replace(".", ",").replace("X", "."),
        }),
        use_container_width=True,
        hide_index=True,
    )

    # Gráfico comparativo
    st.plotly_chart(plot_comparativo(vol_esc, vol_mon), use_container_width=True)

    # ── GRÁFICOS GERAIS ─────────────────────────────────
    render_section("📈 Visão Geral")
    g1, g2 = st.columns([1, 2])
    g1.plotly_chart(plot_status_pie(df), use_container_width=True)
    g2.plotly_chart(plot_ranking_monitor(df), use_container_width=True)

    # ── ABAS PRINCIPAIS ─────────────────────────────────
    t1, t2, t3, t4 = st.tabs(["👥 Equipes", "🧑‍🔧 Técnicos", "🔍 Detalhe Monitor", "📋 Base"])

    with t1:
        te = calcular_volumetria(df, [Config.COL_REGIAO, Config.COL_MONITOR])
        render_dataframe(te, titulo="Volumetria por Equipe", icone="👥", color_col="Taxa Execução", color_meta=Config.META_EXECUCAO)
        st.download_button("📥 Baixar Equipes", gerar_excel(te, "Equipes"), "equipes.xlsx")

    with t2:
        renderizar_volumetria_tecnicos(df, total_montados_fixo)

    with t3:
        md = st.selectbox("Monitor", sel_m, key="monitor_detalhe")
        tt = calcular_volumetria(df[df[Config.COL_MONITOR] == md], [Config.COL_TECNICO])
        tt = tt.sort_values(["Executada", "Taxa Execução", "Total Alocado"], ascending=[False, False, False]).reset_index(drop=True)
        render_dataframe(tt, titulo=f"Técnicos — {md}", icone="🧑‍🔧", color_col="Taxa Execução", color_meta=Config.META_EXECUCAO, height=500)
        st.download_button("📥 Baixar Técnicos", gerar_excel(tt, "Tecnicos"), f"tecnicos_{md}.xlsx")

    with t4:
        render_dataframe(df.head(500), titulo="Base de Dados (prévia — 500 linhas)", icone="📋", badge=f"{len(df)} total", height=600)


if __name__ == "__main__":
    main()